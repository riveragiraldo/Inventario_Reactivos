#Diferentes vistas y/o APIS que interactuan el front con back

import locale
from plistlib import UID
import secrets
from django.shortcuts import render
from django.shortcuts import get_object_or_404, redirect
from django.http import JsonResponse, HttpResponseBadRequest, HttpResponse
from .models import *
from django.db.models import Q
from django.contrib import messages
from decimal import Decimal
from django.views.generic import ListView, View, CreateView, UpdateView
from django.db.models import F
from django.views import View
import json
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
import openpyxl
from openpyxl.worksheet.hyperlink import Hyperlink
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
from openpyxl.styles.colors import WHITE
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as ExcelImage
from datetime import datetime
from django.template.loader import get_template
from io import BytesIO
from reportlab.pdfgen import canvas
from reportlab.lib import utils
from reportlab.lib.pagesizes import letter
from reportlab.platypus import Image
from django.contrib.staticfiles import finders
from PIL import Image as PILImage
from django.urls import reverse
from django.utils.http import urlencode
import time
from django.core.exceptions import ObjectDoesNotExist
from reportlab.lib.pagesizes import letter, landscape
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
import re
import warnings
from urllib.parse import urlparse, urlunparse

from django.conf import settings
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.http import HttpResponseRedirect, QueryDict
from django.shortcuts import resolve_url
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
# from django.utils.deprecation import RemovedInDjango50Warning
from django.utils.http import url_has_allowed_host_and_scheme, urlsafe_base64_decode
from django.utils.translation import gettext_lazy as _
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.debug import sensitive_post_parameters
from django.views.generic.base import TemplateView
from django.views.generic.edit import FormView
from django.contrib.auth import REDIRECT_FIELD_NAME, get_user_model
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import (
    AuthenticationForm,
    PasswordChangeForm,
    PasswordResetForm,
    SetPasswordForm,
)
from django.contrib.auth.tokens import default_token_generator
from django.contrib.sites.shortcuts import get_current_site
from django.core.exceptions import ImproperlyConfigured, ValidationError
from django.http import HttpResponseRedirect, QueryDict
from django.shortcuts import resolve_url
from django.urls import reverse_lazy
from django.utils.decorators import method_decorator
# from django.utils.deprecation import RemovedInDjango50Warning
from django.utils.http import url_has_allowed_host_and_scheme, urlsafe_base64_decode
from django.utils.translation import gettext_lazy as _
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.debug import sensitive_post_parameters
from django.views.generic.base import TemplateView
from django.views.generic.edit import FormView
from django.utils.dateparse import parse_date
from datetime import datetime, timedelta, date
from .forms import  FormularioUsuario, SolicitudForm,ConfiguracionSistemaForm, CustomPasswordResetForm, SolicitudesExternasForm
from django.core.mail import send_mail
from django.utils.translation import gettext as _
from django.contrib.auth.tokens import default_token_generator
from django.template.loader import render_to_string
from django.utils.html import strip_tags
import base64
from urllib.parse import urlencode
from functools import wraps
from django.utils.timezone import make_aware
from django.template.defaultfilters import date as django_date
from django.http import Http404
from openpyxl.styles import Alignment
from django.conf import settings
from django.utils.timezone import localtime
from django.core.files.storage import FileSystemStorage
from django import forms
from django.utils import timezone
from django.utils.timezone import now
from django.http import FileResponse
from django.core.mail import EmailMessage
from django.core.mail import EmailMultiAlternatives
from django.utils.html import strip_tags
from collections import defaultdict


from apscheduler.schedulers.background import BlockingScheduler, BackgroundScheduler
from time import sleep
from captcha.fields import CaptchaField, CaptchaTextInput
from captcha.models import CaptchaStore
from django.utils.html import format_html
from babel.dates import format_datetime, format_time
import threading
import os  # Importa el módulo os para trabajar con rutas de archivo
from openpyxl.styles import NamedStyle



# # Enviar correo electrónico administrativo
# @login_required(login_url='/UniCLab/accounts/login/')
# def enviar_correo(request):
#     if request.method == 'POST':
#         form = CorreoForm(request.POST, request.FILES)
#         if form.is_valid():
#             destino = form.cleaned_data['destino']
#             print(destino)
#             laboratorio = form.cleaned_data['laboratorio']
#             usuario = form.cleaned_data['usuario']
#             asunto = form.cleaned_data['asunto']
#             contenido = form.cleaned_data['mensaje']
#             adjunto = form.cleaned_data['adjunto']
#             if adjunto and adjunto.size > 5 * 1024 * 1024:  # 2 MB en bytes
#                 nombre=adjunto.name
#                 peso=adjunto.size / (1024 * 1024)
#                 mensaje = f'No se ha podido enviar el mensaje porque el tamaño del archivo adjunto con nombre {nombre}, con tamaño de {peso:.2f} MB es superior al máximo permitido (5 MB).'
#                 messages.error(request, mensaje)  # Agregar mensaje de error
#                 return redirect('reactivos:enviar_correo')
        

#             # Lógica para determinar a quién enviar el correo y construir la lista de destinatarios
#             if destino=='USUARIO_ESPECIFICO':
#                 destinatarios=[usuario,]
#             elif destino=='TODOS' and laboratorio=='TODOS':
#                 destinatarios = UserModel.objects.filter(is_active=True).values_list('email', flat=True)
#             elif destino=='TODOS' and laboratorio!='TODOS':
#                 destinatarios = UserModel.objects.filter(is_active=True, lab=laboratorio).values_list('email', flat=True)
#             elif destino!='TODOS' and laboratorio=='TODOS':
#                 destinatarios = UserModel.objects.filter(is_active=True, rol=destino).values_list('email', flat=True)
#             elif destino!='TODOS' and laboratorio!='TODOS':
#                 destinatarios = UserModel.objects.filter(is_active=True, rol=destino, lab=laboratorio).values_list('email', flat=True)
                
#             else:
#                 destinatarios = ['andresrgiraldo@gmail.com']

#             # Lógica para construir el mensaje de correo

#             # Crear de contexto para el mensaje de correo
#             # Obtener el dominio
#             url = settings.BASE_URL
#             # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
#             fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
#             # Formatear el valor con separadores de miles
#             locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')

#             # Contexto del template del diseño del correo 
#             context = {
#                 'url': url,
#                 'fecha_actual': fecha_actual,
#                 'message': contenido,
#             }

#             # Template
#             html_message = render_to_string('admin/enviar_correo_admin.html', context)
#             plain_message = strip_tags(html_message)
#             from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
#             recipient_list = destinatarios
#             subject = f'Mensaje desde {url} -  {asunto}'

#             # Adjuntar el archivo al correo si se proporciona
#             if adjunto:
#                 # Si hay archivos adjuntos, enviar con mensaje HTML y adjunto
#                 email = EmailMultiAlternatives(subject, plain_message, from_email, recipient_list)
#                 email.attach_alternative(html_message, "text/html")  # Agregar versión HTML del mensaje
#                 email.attach(adjunto.name, adjunto.read(), adjunto.content_type)
#                 email.send()
#                 print('Se ha enviado el correo con mensaje HTML y adjunto')
#             else:
#                 print('Se ha enviado el correo')
#                 send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=html_message,)
            
#             mensaje=f'Se ha enviado el mensaje de manera correcta'
#             messages.success(request, mensaje)  # Agregar mensaje de éxito
#             return redirect('reactivos:enviar_correo')
#         else:
#             mensaje=f'{form.errors}'
#             messages.error(request, mensaje)  # Agregar mensaje de error
#             return redirect('reactivos:enviar_correo')
#     else:
#         form = CorreoForm()

#     return render(request, 'admin/enviar_correo.html', {'form': form})


# Función para crear enlace de descarga de manual de usuario
def descargar_manual(request):
    configuracion_sistema = ConfiguracionSistema.objects.first()  # Obtiene el primer registro

    if configuracion_sistema:
        manual = configuracion_sistema.manual
        response = FileResponse(manual)
        return response
    
# Función para servir la imagen del logo institucional
def logo_institucional(request):
    configuracion_sistema = ConfiguracionSistema.objects.first()

    if configuracion_sistema and configuracion_sistema.logo_institucional:
        logo = configuracion_sistema.logo_institucional
        return FileResponse(logo.open(), content_type='image')
    else:
        # Lógica para manejar el caso en que no se encuentre una imagen
        # Por ejemplo, puedes devolver una imagen predeterminada o un error.
        return HttpResponse("No se ha encontrado un logo institucional")

# Función que envía correo a coordinadores de laboratorio de reactivos vencidos o próximos a vencer
def enviar_correo_alerta_vencimiento(subject, recipient_list, message, reactivos):
    # Crear de contexto para el mensaje de correo
    # Obtener el dominio
    url=settings.BASE_URL
    # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
    fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
    # Formatear el valor con separadores de miles
    locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
                  
    # Contexto del template del diseño del correo 
    context = {
        'url': url,
        'fecha_actual': fecha_actual,
        'message':message,
        'reactivos':reactivos,
        
    }
    # Template
    message = render_to_string('reactivos/alerta_vencimiento_reactivos.html', context)
    plain_message = strip_tags(message)
    from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
    recipient_list = recipient_list
    subject = subject
    send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)


# Tareas programadas para alertas de fecha de vencimientos
def check_edate():
    
    # Obtenemos la fecha actual
    today = now().date()
    print(str(datetime.now()))

    # Iteramos a través de los laboratorios
    for lab in Laboratorios.objects.all():
        
        # Filtramos los reactivos de este laboratorio cuya fecha de vencimiento (edate) es menor o igual a la fecha actual
        reactivos_vencidos = Inventarios.objects.filter(edate__lte=today, is_active=True, edate__isnull=False, lab=lab)

        if reactivos_vencidos:
                      
            # Consulta para obtener los usuarios con el rol "COORDINADOR" que pertenecen a este laboratorio
            coordinadores = User.objects.filter(rol__name='COORDINADOR', lab=lab)
            
            
            # Ahora, construyamos el correo para este laboratorio
            subject = f'Alerta de vencimiento de reactivos en {lab.name}'
            recipient_list = [coordinador.email for coordinador in coordinadores]
            
            message = f'Estimado(s) {", ".join([f"{coordinador.first_name} {coordinador.last_name}" for coordinador in coordinadores])},\n'
            message += f'Los siguientes reactivos están vencidos en el laboratorio {lab.name}:\n\n'
            print(f'Enviando correo a {recipient_list}')
                       
            enviar_correo_alerta_vencimiento(subject, recipient_list, message, reactivos_vencidos)
        else:
            print(f'Para el laboratorio {lab} no hay reactivos vencidos')
        
# Función que revisa los reactivos próximos a vencer
def check_upcoming_edate():
    
    # Obtener la fecha actual
    today = datetime.now().date()
    # Obtener la configuración del sistema (debería haber solo una instancia)
    configuracion = ConfiguracionSistema.objects.first()
    if configuracion:
        # Obten el valor de 'tiempo_vencimiento_reactivos' de la configuración
        delta = configuracion.tiempo_vencimiento_reactivos
    else:
        # Calculamos la fecha de final (90 días en el futuro)
        delta=90
    # Calculamos la fecha de inicio (mañana)
    start_date = today + timedelta(days=1)   
    
    end_date = today + timedelta(days=delta)

    # Iteramos a través de los laboratorios
    for lab in Laboratorios.objects.all():
        
        # Filtramos los reactivos de este laboratorio cuya fecha de vencimiento (edate) está entre la fecha de inicio y final
        reactivos_proximos_a_vencer = Inventarios.objects.filter(edate__gte=start_date, edate__lte=end_date, is_active=True, edate__isnull=False, lab=lab)

        if reactivos_proximos_a_vencer:
            # Consulta para obtener los usuarios con el rol "COORDINADOR" que pertenecen a este laboratorio
            coordinadores = User.objects.filter(rol__name='COORDINADOR', lab=lab)
                      
            # Ahora, construyamos el correo para este laboratorio
            subject = f'Alerta de reactivos próximos a vencer en {lab.name}'
            recipient_list = [coordinador.email for coordinador in coordinadores]
            
            message = f'Estimado(s) {", ".join([f"{coordinador.first_name} {coordinador.last_name}" for coordinador in coordinadores])},\n'
            message += f'Los siguientes reactivos están próximos a vencer en un lapso de los próximos {delta} días, en el laboratorio {lab.name}:\n\n'
            print(f'Enviar correo a {recipient_list} laboratorio {lab}')
            enviar_correo_alerta_vencimiento(subject, recipient_list, message, reactivos_proximos_a_vencer)
        else:
            print(f'Para el laboratorio {lab} no hay reactivos próximos a vencer')


scheduler = BackgroundScheduler()

def programar_tareas():
    
    configuracion = ConfiguracionSistema.objects.first()  # Suponiendo que solo hay una configuración en la base de datos.
    
    global scheduler
    if configuracion:
        
        start_date=configuracion.fecha_incio         
        scheduler.add_job(check_upcoming_edate, 'interval', days=configuracion.intervalo_tiempo, start_date=start_date, )
        five_minutes=timedelta(minutes=5)
        print(start_date)
        start_date=start_date+five_minutes
        print(start_date)
        scheduler.add_job(check_edate, 'interval', days=configuracion.intervalo_tiempo, start_date=start_date,)
        one_minutes=timedelta(minutes=1)
        print(start_date)
        start_date=start_date+one_minutes
        print(start_date)
        scheduler.add_job(depurar_eventos_antiguos, 'interval', days=1, start_date=start_date,)
        print(scheduler.state)
        if  configuracion.programacion_activa:
            if scheduler.state==0:
                scheduler.start()
            elif scheduler.state==2:
                scheduler.state=1

        else:
            print(scheduler.state)
            if scheduler.state==1:
                scheduler.pause()
        print(scheduler.state)
# programar_tareas()

# Vista para depuración de eventos
def depurar_eventos_antiguos():
    # Depurar eventos antiguos
    # Obtén la configuración del sistema
    configuracion = ConfiguracionSistema.objects.first()  # Suponiendo que solo hay una configuración en la base de datos.

    if configuracion:
        # Calcula la fecha límite para eventos antiguos
        fecha_limite = timezone.now() - timedelta(days=configuracion.tiempo_eventos)

        # Elimina eventos antiguos
        Eventos.objects.filter(fecha_evento__lt=fecha_limite).delete()
        print('Se han depurado los eventos')

    # Depurar solicitudes Internas
    # Obtener la configuración del sistema, suponiendo que existe un único registro.
    configuracion_sistema = ConfiguracionSistema.objects.first()
    # Comprobar si se obtuvo la configuración.
    if configuracion_sistema:
        # Usar el campo tiempo_solicitudes de la configuración como max_antiguedad.
        max_antiguedad = timedelta(days=configuracion_sistema.tiempo_solicitudes)
    else:
        # Si no se pudo obtener la configuración, usar un valor predeterminado (30 días, por ejemplo).
        max_antiguedad = timedelta(days=30)
    fecha_limite = timezone.now() - max_antiguedad
    # Obtén las solicitudes antiguas que cumplen con el criterio de antigüedad
    solicitudes_antiguas = Solicitudes.objects.filter(tramitado=True, fecha_tramite__lt=fecha_limite)
    # Eliminar los archivos adjuntos asociados a las solicitudes antiguas
    for solicitud in solicitudes_antiguas:
        if solicitud.archivos_adjuntos:
            # Asegúrate de que el archivo realmente se elimine del disco.
            solicitud.archivos_adjuntos.delete()
    # Eliminar las solicitudes antiguas
    solicitudes_antiguas.delete()
    print('Se han depurado las solicitudes internas')

    # Depurar solicitudes Externas
    # Obtener la configuración del sistema, suponiendo que existe un único registro.
    configuracion_sistema = ConfiguracionSistema.objects.first()
    # Comprobar si se obtuvo la configuración.
    if configuracion_sistema:
        # Usar el campo tiempo_solicitudes de la configuración como max_antiguedad.
        max_antiguedad = timedelta(days=configuracion_sistema.tiempo_solicitudes)
    else:
        # Si no se pudo obtener la configuración, usar un valor predeterminado (30 días, por ejemplo).
        max_antiguedad = timedelta(days=30)
    fecha_limite = timezone.now() - max_antiguedad
    # Obtén las solicitudes antiguas que cumplen con el criterio de antigüedad
    solicitudes_antiguas = SolicitudesExternas.objects.filter(is_view=True, registration_date__lt=fecha_limite)
    # Eliminar los archivos adjuntos asociados a las solicitudes antiguas
    for solicitud in solicitudes_antiguas:
        if solicitud.attach:
            # Asegúrate de que el archivo realmente se elimine del disco.
            solicitud.attach.delete()
    # Eliminar las solicitudes antiguas
    solicitudes_antiguas.delete()
    print('Se han depurado las solicitudes externas')

from django.shortcuts import get_object_or_404

def crear_evento(tipo_evento, usuario_evento):
    # Verifica si el tipo de evento existe
    tipo_evento_obj, created = TipoEvento.objects.get_or_create(name=tipo_evento)

    # Si no existía, se creó un nuevo objeto tipo_evento
    if created:
        print(f"Se ha creado un nuevo tipo de evento: {tipo_evento}")

    evento = Eventos.objects.create(
        tipo_evento=tipo_evento_obj,
        usuario_evento=usuario_evento,
    )


#-------------------------------------------------#


UserModel = get_user_model()

#Esta vista valida los grupos de usuarios a los que pertenece la persona que está tratando de acceder
#a la vista
def check_group_permission(groups_required):
    def decorator(view_func):
        @method_decorator(login_required)
        def _wrapped_view(self, request, *args, **kwargs):
            grupos_usuario = self.request.user.groups.all().values('name')
            is_of_group = False
            
            for grupo in grupos_usuario:
                for grupo_requerido in groups_required:
                    if grupo['name'] == grupo_requerido:
                        is_of_group = True
                        break
            
            if is_of_group or self.request.user.is_superuser:
                return view_func(self, request, *args, **kwargs)
            else:
                messages.error(request, 'No tiene permisos para acceder a esta vista, desea acceder con credenciales distintas')
                return HttpResponseRedirect(reverse('reactivos:login'))

        return _wrapped_view
    return decorator


# Vista para la visualización del web template
@login_required
def webtemplate(request):
    laboratorio = request.user.lab
    
    
    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,
        
    }
    return render(request, 'webtemplate.html', context)

# Función que envía alertas de Stock Mínimo, o stock =0
def enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente):
    # Crear de contexto para el mensaje de correo
    # Obtener el dominio
    protocol = 'https' if request.is_secure() else 'http'
    domain = request.get_host()
    url=f'{protocol}://{domain}'
    # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
    fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
    # Formatear el valor con separadores de miles
    locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
    usuario=f'{request.user.first_name} {request.user.last_name}'
    laboratorio=inventario_existente.lab
                
    # Contexto del template del diseño del correo 
    context = {
        'url': url,
        'fecha_actual': fecha_actual,
        'reactivo':reactivo,
        'marca':marca,
        'referencia':referencia,
        'usuario':usuario,
        'mensaje':mensaje,
        'alerta':alerta,
        'cantidad':cantidad,
        'mensaje':mensaje,
        'laboratorio':laboratorio,
        'unidad':unidad,
    }
    # Template
    message = render_to_string('reactivos/alerta_reactivos.html', context)
    plain_message = strip_tags(message)
    from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
    rol_coordinador=get_object_or_404(Rol, name='COORDINADOR')
    rol_tecnico=get_object_or_404(Rol, name='TECNICO')
    coordinators=User.objects.filter(lab=inventario_existente.lab, rol=rol_coordinador, is_active=True)
    technicians=User.objects.filter(lab=inventario_existente.lab, rol=rol_tecnico, is_active=True)
    recipient_list = [request.user.email]
    recipient_list.extend(coordinator.email for coordinator in coordinators)
    recipient_list.extend(technician.email for technician in technicians)
    subject = f"Alerta en reactivo {reactivo} - {alerta} - {laboratorio}"
    send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)

# Función que correo al realizar una solicitud
def enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email, email_type ):
    # Crear de contexto para el mensaje de correo
    # Obtener el dominio
    protocol = 'https' if request.is_secure() else 'http'
    domain = request.get_host()
    suffix=suffix
    url_ppal=f'{protocol}://{domain}'
    url=f'{protocol}://{domain}/{suffix}'
    # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
    fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
    # Formatear el valor con separadores de miles
    locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
    solicitud=get_object_or_404(Solicitudes, id=id)
    laboratorio=solicitud.created_by.lab.name
                
    # Contexto del template del diseño del correo 
    context = {
        'url': url,
        'url_ppal':url_ppal,
        'solicitud':solicitud,
        'initial_message':initial_message,
    
        
    }
    # Template
    message = render_to_string('solicitudes/registro_exitoso_solicitud.html', context)
    plain_message = strip_tags(message)
    from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
     
    recipient_list = [shipping_email]
    tipo_solicitud=get_object_or_404(TipoSolicitud, name='OTRA')

    if solicitud.tipo_solicitud.id==tipo_solicitud.id:
        asunto=solicitud.name
    else:
        
        asunto=solicitud.tipo_solicitud.name
    
    url=f'{protocol}://{domain}'
    subject = f"{email_type} - {asunto} en {url_ppal} - {laboratorio}"
    send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)

# Vista que maneja la configuración del sistema

def configuraciones(request):
    
    configuracion = ConfiguracionSistema.objects.first()
    

    if not configuracion:
        # Si no existe ningún registro en la tabla, crearemos uno nuevo
        if request.method == 'POST':
            form = ConfiguracionSistemaForm(request.POST, request.FILES)
            if form.is_valid():
                instance = form.save(commit=False)
                # Completa el campo 'url' con el protocolo y el dominio principal
                protocol = 'https' if request.is_secure() else 'http'
                domain = request.get_host()
                instance.url = f'{protocol}://{domain}'
                instance.save()
                # Crea un evento de modificar configuraciones
                tipo_evento = 'MODIFICAR CONFIGURACIONES'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                programar_tareas()
                mensaje=f'Se han creado las configuraciones de manera correcta.'
                messages.success(request, mensaje)
                return redirect('reactivos:configuraciones')
        else:
            form = ConfiguracionSistemaForm()
    else:
        # Si ya existe un registro, actualizaremos el primero con los nuevos valores
        if request.method == 'POST':
            form = ConfiguracionSistemaForm(request.POST, request.FILES, instance=configuracion)
            if form.is_valid():
                instance = form.save(commit=False)
                # Completa el campo 'url' con el protocolo y el dominio principal
                protocol = 'https' if request.is_secure() else 'http'
                domain = request.get_host()
                instance.url = f'{protocol}://{domain}'
                instance.save()
                # Crea un evento de modificar configuraciones
                tipo_evento = 'MODIFICAR CONFIGURACIONES'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)
                programar_tareas()
                mensaje=f'Se han actualizado las configuraciones de manera correcta.'
                messages.success(request, mensaje)
                return redirect('reactivos:configuraciones')
        else:
            form = ConfiguracionSistemaForm(instance=configuracion)
    if request.user.lab:
        laboratorio = request.user.lab.name
    else:
        laboratorio = 'No Lab'
    usuario = request.user
    context = {
        'usuario': usuario,
        'laboratorio': laboratorio,
        'form': form,
        }

    return render(request, 'admin/configuraciones.html', context)


# Vista para la validación de solicitudes externas, 

class PreSolDirLab(View):  # Utiliza LoginRequiredMixin como clase base
    template_name = 'dir_lab/pre_solicitudes_externas.html'  # Nombre de la plantilla

    def get(self, request,*args,**kwargs):
        
             
        context = {
            
        }
        return render(request, self.template_name, context)

import requests
# solicitudes externas  
class SolDirLab(View):
    template_name = 'dir_lab/solicitudes_externas.html'

    def get_recipient_list(self, lab):
        # Obtener correos de usuarios con rol 'COORDINADOR' o 'ADMINISTRADOR' en el laboratorio especificado
        recipient_list = User.objects.filter(
            lab=lab
        ).values_list('email', flat=True)
        return list(recipient_list)
    
    def enviar_correo_asincrono(self, recipient_list, subject, message, attach_path):
        try:
            enviar_correo(recipient_list, subject, message, attach_path)
        except Exception as e:
            print(f'Error al enviar correo: {e}')

    def get(self, request, *args, **kwargs):
        form = SolicitudesExternasForm()
        return render(request, self.template_name, {'form': form})
    
    
            

    def post(self, request, *args, **kwargs):
        form = SolicitudesExternasForm(request.POST, request.FILES)
        
        try:
            if form.is_valid():
                # Obtener el token
                token = request.POST.get('access_token', '')
                
                # URL de verificación del token de Google
                google_token_url = 'https://www.googleapis.com/oauth2/v2/userinfo'

                # Parámetros de la solicitud
                params = {'access_token': token}

                # Realizar la solicitud a la API de Google
                response = requests.get(google_token_url, params=params)

                # Verificar el estado de la respuesta
                if response.status_code == 200:
                    # La solicitud fue exitosa, verificar los datos en la respuesta
                    data = response.json()
                    if 'error_description' in data:
                        # El token no es válido, ha expirado u otro error
                        print(f'Token inválido: {data["error_description"]}')
                        return JsonResponse({'success': False, 'errors': 'La sesión no es válida o ha caducado, debe autenticarse nuevamente para realizar la solicitud'})

                    else:
                        # El token es válido, puedes acceder a los datos en 'data'
                        print('Token válido')
                        # Guardar la solicitud si el formulario es válido
                        solicitud = form.save()
                        # Obtener la lista de destinatarios
                        recipient_list = self.get_recipient_list(solicitud.lab)
                        # Agregar el correo de la solicitud a la lista de destinatarios
                        recipient_list.append(solicitud.email)

                        subject=f'Registro de solicitud externa {solicitud.subject} en {solicitud.lab.name}'
                        header=f'<p>El presente mensaje es para informarte que desde la web principal de Dirección de Laboratorios se ha registrado una solicitud externa dirigida al laboratorio {solicitud.lab.name}, los datos de la solcitud son los siguientes:</p>'
                        body=f'<p><b>Laboratorio: </b>{solicitud.lab.name}<br><b>Remitente: </b>{solicitud.name}<br><b>Correo electrónico: </b>{solicitud.email}<br><b>Teléfono de contacto: </b>{solicitud.mobile_number}<br><b>Área: </b>{solicitud.department}<br><b>Asunto: </b>{solicitud.subject}<br><b>Mensaje: </b>{solicitud.message}</p>'
                        if solicitud.attach:
                            footer=f'<p>Adjunto encontrarás la información cargada desde el formulario de solicitudes.</p>'
                        else:
                            footer=f''
                        message=header+body+footer

                        if solicitud.attach:
                            attach_path=solicitud.attach.path
                        else:
                            attach_path=None
                        # Crear un hilo y ejecutar enviar_correo en segundo plano
                        correo_thread = threading.Thread(
                            target=self.enviar_correo_asincrono,
                            args=(recipient_list, subject, message, attach_path),
                        )
                        correo_thread.start()
                        # enviar_correo(recipient_list, subject, message,attach_path)
                        mensaje=f'La solicitud se ha registrado de manera correcta, se ha enviado un correo electrónico a {solicitud.email} con los datos de la solicitud.'

                        # Devuelve una respuesta JSON de éxito
                        return JsonResponse({'success': True, 'message': mensaje})

                else:
                    # La solicitud no fue exitosa, manejar el error
                    print(f'Error al verificar el token: {response.status_code}')
                    return JsonResponse({'success': False, 'errors': 'La sesión no es válida o ha caducado, debe autenticarse nuevamente para realizar la solicitud'})

                
            else:
                # Devuelve una respuesta JSON con los errores de validación
                return JsonResponse({'success': False, 'errors': form.errors})
        except Exception as e:
            # Imprimir el error en la consola o logs
            print(e)
            mensaje=f'Error: {e}'
            # Devolver una respuesta de error o redirigir a una página de error
            return HttpResponseBadRequest(f'Error interno del servidor:{mensaje}')

class SolicitudesExternasListView(LoginRequiredMixin,ListView):
    model = SolicitudesExternas
    template_name = "solicitudes/listado_solicitudes_externas.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR','COORDINADOR','TECNICO'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        lab = self.request.GET.get('lab')
        keyword = self.request.GET.get('keyword')  
        
        

        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        request.session['filtered_lab'] = lab
        request.session['filtered_keyword'] = keyword
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Calcular la fecha hace un mes hacia atrás
        one_month_ago = today - timedelta(days=30)

        # Agregar la fecha al contexto
        context['one_month_ago'] = one_month_ago

        # Agregar la fecha de hoy al contexto
        context['today'] = today
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        
        # Obtener la lista de inventarios
        entradas = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha        
               
        context['object_list'] = entradas
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()

        lab = self.request.GET.get('lab')
        keyword = self.request.GET.get('keyword')
        
        

        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=''

        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(registration_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(registration_date__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(registration_date__gte=start_date,date_create__lte=end_date)
        
        # Filtrar por  campos que contengan la palabra clave en su nombre
        if keyword:
            queryset = queryset.filter(Q(name__icontains=keyword) | Q(subject__icontains=keyword) | Q(message__icontains=keyword) | Q(attach__icontains=keyword))
        
        if lab:
            queryset = queryset.filter(lab=lab)

        queryset = queryset.order_by('id')
        return queryset

@login_required
def export_to_excel_solicitud_externa(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario    
    start_date = request.session.get('filtered_start_date')
    end_date = request.session.get('filtered_end_date')
    lab = request.session.get('filtered_lab')
    keyword = request.session.get('filtered_keyword')
    # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
    if lab=='0':
         lab=''
    
    
    # Validar y convertir las fechas
    try:
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
        pass

    queryset = SolicitudesExternas.objects.all()
    #Filtra según los valores previos de filtro en los selectores
        
    # Realiza la filtración de acuerdo a las fechas
    if start_date:
        queryset = queryset.filter(registration_date__gte=start_date)
    if end_date:
        queryset = queryset.filter(registration_date__lte=end_date)
    elif start_date and end_date:
        queryset = queryset.filter(registration_date__gte=start_date,date_create__lte=end_date)
    
    # Filtrar por  campos que contengan la palabra clave en su nombre
    if keyword:
        queryset = queryset.filter(Q(name__icontains=keyword) | Q(subject__icontains=keyword) | Q(message__icontains=keyword) | Q(attach__icontains=keyword))
    
    if lab:
        queryset = queryset.filter(lab=lab)
    
    queryset = queryset.order_by('id')
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    sheet.merge_cells('C1:F1')

    sheet['C1'] = 'Listado de solicitudes externas'
    sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Id'
    sheet['B4'] = 'Fecha de solicitud'
    sheet['C4'] = 'Asunto'
    sheet['D4'] = 'Mensaje'
    sheet['E4'] = 'Archivos Adjuntos'
    sheet['F4'] = 'Laboratorio a la que va dirigida la solicitud'
    sheet['G4'] = 'Remitente'
    sheet['H4'] = 'Correo electrónico'
    sheet['I4'] = 'Teléfono'
    sheet['J4'] = 'Área'

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1

    cell_A1 = sheet['C1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 5
    sheet.column_dimensions['A'].width = 5

    # Establecer el ancho de la columna B a 26
    sheet.column_dimensions['B'].width = 26

    # Establecer el ancho de la columna C a 34
    sheet.column_dimensions['C'].width = 34

    # Establecer el ancho de la columna D a 40
    sheet.column_dimensions['D'].width = 40

    # Establecer el ancho de la columna E a 30
    sheet.column_dimensions['E'].width = 30

    # Establecer el ancho de la columna F a 27
    sheet.column_dimensions['F'].width = 27

    # Establecer el ancho de la columna G a 21
    sheet.column_dimensions['G'].width = 21

    # Establecer el ancho de la columna H a 30
    sheet.column_dimensions['H'].width = 30

    # Establecer el ancho de la columna I a 11
    sheet.column_dimensions['I'].width = 11

    # Establecer el ancho de la columna J a 20
    sheet.column_dimensions['J'].width = 20


    # Define una alineación que tenga la vertical en la parte superior y la horizontal a la izquierda
    # Crear un objeto de alineación
    alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    
    
    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 11):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        if item.attach:
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url = f'{protocol}://{domain}{item.attach.url}'

            # Obtener solo el nombre del archivo de la ruta completa
            nombre_archivo = os.path.basename(item.attach.name)

            # Establecer un estilo para la celda con el nombre del archivo
            estilo_celda = sheet.cell(row=row, column=5)
            estilo_celda.value = nombre_archivo
            estilo_celda.font = Font(color="0000FF", underline="single")  # Azul y subrayado

            # Crear un objeto Hyperlink con la URL
            hyperlink = Hyperlink(target=url, ref=f'A{row}')

            # Aplicar el hipervínculo a la celda
            estilo_celda.hyperlink = hyperlink
        else:
            sheet.cell(row=row, column=5).value = ""

        sheet.cell(row=row, column=1).value = (item.id)
        sheet.cell(row=row, column=2).value = str((item.registration_date).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=3).value = str(item.subject)
        sheet.cell(row=row, column=4).value = item.message
        # sheet.cell(row=row, column=5).value = download
        sheet.cell(row=row, column=6).value = item.lab.name
        sheet.cell(row=row, column=7).value = item.name
        sheet.cell(row=row, column=8).value = item.email
        sheet.cell(row=row, column=9).value = str(item.mobile_number)
        sheet.cell(row=row, column=10).value =item.department

               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 11):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).alignment = alignment

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 10
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_solicitudes_externas.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Vista que maneja el elimina la solicitud externa
@login_required
def eliminar_solicitud_externa(request, solicitud_code):
    try:
        solicitud_id = int(base64.urlsafe_b64decode(solicitud_code).decode())
        
    except (ValueError, SolicitudesExternas.DoesNotExist):
        return HttpResponse('Solicitud no existe', 400)

    solicitud = get_object_or_404(SolicitudesExternas, pk=solicitud_id)
    solicitud.delete()
    # crear evento 
    tipo_evento = 'ELIMINAR SOLICITUD EXTERNA'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    return HttpResponse('Solicitud eliminada correctamente', 200)

# Vista que cambia la solicitud a leída
@login_required
def solicitud_leida(request, solicitud_code):
    try:
        solicitud_id = int(base64.urlsafe_b64decode(solicitud_code).decode())
        
    except (ValueError, SolicitudesExternas.DoesNotExist):
        return HttpResponse('Solicitud no existe', 400)

    solicitud = get_object_or_404(SolicitudesExternas, pk=solicitud_id)
    solicitud.is_view=True
    solicitud.save()
    # Evento
    tipo_evento = 'MARCAR SOLICITUD EXTERNA COMO LEIDA'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    return HttpResponse('Solicitud marcada como leída correctamente', 200)

# Vista que cambia la solicitud a leída
@login_required
def solicitud_no_leida(request, solicitud_code):
    try:
        solicitud_id = int(base64.urlsafe_b64decode(solicitud_code).decode())
        
    except (ValueError, SolicitudesExternas.DoesNotExist):
        return HttpResponse('Solicitud no existe', 400)

    solicitud = get_object_or_404(SolicitudesExternas, pk=solicitud_id)
    solicitud.is_view=False
    solicitud.save()
    # Evento
    tipo_evento = 'MARCAR SOLICITUD EXTERNA COMO  NO LEIDA'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    return HttpResponse('Solicitud marcada como no leída correctamente', 200)


# Enviar Correo
def enviar_correo(recipient_list, subject, message,attach_path):
    attach = None  # Asigna un valor predeterminado
    # Lógica para construir el mensaje de correo
    if attach_path:
        attach = open(attach_path, 'rb')
    message = format_html(message)
    # Obtener el dominio
    configuracion = ConfiguracionSistema.objects.first()
    url = configuracion.url
    # Configura la localización a español
    locale.setlocale(locale.LC_TIME, 'es_ES.UTF-8')
    
    # Obtén la fecha y hora actual en el formato "dd de mes de aaaa, hh:mm AM/PM"
    fecha_actual = format_datetime(localtime(), 'dd MMMM yyyy, hh:mm a', locale='es_ES')
    
    # Contexto del template del diseño del correo 
    context = {
        'url': url,
        'fecha_actual': fecha_actual,
        'message':message,
        }
    
    # Template
    html_message = render_to_string('dir_lab/enviar_correo.html', context)
    plain_message = strip_tags(html_message)
    from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
    recipient_list = recipient_list
    subject = f'{subject} - {url}'

    try:
        email = EmailMultiAlternatives(subject, plain_message, from_email, recipient_list)
        email.attach_alternative(html_message, 'text/html')  # Agrega el contenido HTML
        if attach:
            email.attach_file(attach_path)
        email.send()
        print('Se ha enviado el correo con mensaje HTML y adjunto')
    except Exception as e:
        print(f'Error al enviar el correo: {e}')
    finally:
        if attach:
            attach.close()

    mensaje=f'Se ha enviado el mensaje de manera correcta'
    print(mensaje)

# Vista para la creación del index, 

class Index(LoginRequiredMixin, View):  # Utiliza LoginRequiredMixin como clase base
    template_name = 'reactivos/index.html'  # Nombre de la plantilla

    def get(self, request,*args,**kwargs):
        laboratorio = request.user.lab   
             
        context = {
            'usuarios': User.objects.all(),
            'laboratorio': laboratorio,
        }
        return render(request, self.template_name, context)
    
# Recargar Captcha
def recargar_captcha(request):
    if request.method == 'POST':
        # Generar una nueva clave de captcha
        captcha_key = CaptchaStore.generate_key()

        return JsonResponse({'captcha_key': captcha_key})

    return JsonResponse({'error': 'Método de solicitud no permitido.'})

# Vista para la visualización de enlaces, 

class Enlaces(LoginRequiredMixin, View):  # Utiliza LoginRequiredMixin como clase base
    template_name = 'reactivos/enlaces.html'  # Nombre de la plantilla

    def get(self, request,*args,**kwargs):
        laboratorio = request.user.lab
        category = self.kwargs.get('category')  # Obtener el valor de la URL   
             
        context = {
            'usuarios': User.objects.all(),
            'laboratorio': laboratorio,
            'category': category,  # Agregar 'category' al contexto
        }
        return render(request, self.template_name, context)
# La vista "crear_unidades" se encarga de gestionar la creación de unidades. Esta vista toma los datos del formulario 
# existente en el template "crear_unidades.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Unidades". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la unidad
# ya existe en la base de datos antes de crearla. Si la unidad es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Unidades". Si la unidad ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.

class CrearUnidades(LoginRequiredMixin, View):
    template_name = 'reactivos/crear_unidades.html'
    template_exito = 'reactivos:crear_estado'

    @check_group_permission(groups_required=['COORDINADOR','ADMINISTRADOR'])
    def get(self, request, *args, **kwargs):
        laboratorio = self.request.user.lab
        context = {
            'usuarios': User.objects.all(),
            'laboratorio': laboratorio,
            
        }
        return render(request, self.template_name, context)

    @check_group_permission(groups_required=['COORDINADOR','ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        name = request.POST.get('name')

        if Unidades.objects.filter(name=name).exists():
            unidad = Unidades.objects.get(name=name)
            unidad_id = unidad.id
            mensaje=f'Ya existe una unidad con nombre {name} id: {str(unidad_id)}'
            messages.error(request, mensaje)
            return HttpResponse(mensaje,400)            

        unidad = Unidades.objects.create(
            name=name,
            created_by=request.user,
            last_updated_by=request.user,
        )
        unidad_id = unidad.id        

        context = {'unidad_id': unidad.id, 'unidad_name': unidad.name}
        
        tipo_evento = 'CREAR UNIDAD'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        mensaje=f'Se ha creado exitosamente la unidad con nombre {name} id: {str(unidad_id)}'
        messages.success(request, mensaje)
        return HttpResponse(mensaje,200)

# Verificar estado de autenticación
def check_auth_status(request):
    if request.user.is_authenticated:
        return JsonResponse({}, status=200)
    else:
        return JsonResponse({}, status=401)

    
# La vista "crear_tipo de solicitudes" se encarga de gestionar la creación de tipo de solictudes. Esta vista toma los datos del formulario 
# existente en el template "crear_tipo_solicitudes.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "TipoSolicitud". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si el tipo de solicitud
# ya existe en la base de datos antes de crearla. Si el tipo de solictud es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "TipoSolicitud". Si la esta ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.

class CrearTipoSolicitud(LoginRequiredMixin, View):
    template_name = 'solicitudes/crear_tipo_solicitud.html'

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def get(self, request, *args, **kwargs):
        laboratorio = self.request.user.lab
        context = {
            'usuarios': User.objects.all(),
            'laboratorio': laboratorio,
            
        }
        return render(request, self.template_name, context)

    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def post(self, request, *args, **kwargs):
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        if TipoSolicitud.objects.filter(name=name).exists():
            tipo_solicitud = TipoSolicitud.objects.get(name=name)
            tipo_solicitud_id = tipo_solicitud.id
            messages.error(
                request, 'Ya existe un tipo de solicitus con nombre ' + name + ' id: ' + str(tipo_solicitud))
            return HttpResponse('Error al insertar en la base de datos', status=400)

        tipo_solicitud = TipoSolicitud.objects.create(
            name=name,
            created_by=request.user,
            last_updated_by=request.user,
        )
        tipo_solicitud_id = tipo_solicitud.id
        # Crea un evento de crear tipo de solicitud
        tipo_evento = 'CREAR TIPO DE SOLICITUD'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente el tipo de solicitud con nombre ' + name + ' id: ' + str(tipo_solicitud_id))

        context = {'tipo_solicitud_id': tipo_solicitud.id, 'tipo_solicitud_name': tipo_solicitud.name}
        
        return HttpResponse('Operación exitosa', status=200)
    
# La vista "solicitudes" se encarga de gestionar el registro de solictudes. toma los datos del formuario de registro
# de solicitudes, los registra en la base de datos y envía correo electrónico al administrador del sistema


class RegistrarSolicitud(LoginRequiredMixin, CreateView):
    model = Solicitudes
    form_class = SolicitudForm
    template_name = 'solicitudes/registrar_solicitud.html'
    success_url = '/'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # Agregar el usuario en sesión y el laboratorio al contexto
        context['usuario'] = self.request.user
        context['laboratorio'] = self.request.user.lab
        return context

    def post(self, request, *args, **kwargs):
        form = self.form_class(request.POST, request.FILES)

        # Validar el tamaño del archivo adjunto
        archivos_adjuntos = request.FILES.get('archivos_adjuntos')
        if archivos_adjuntos and archivos_adjuntos.size > 5 * 1024 * 1024:  # 2 MB en bytes
            nombre=archivos_adjuntos.name
            peso=archivos_adjuntos.size / (1024 * 1024)
            mensaje = f'No se ha podido registrar la solicitud porque el tamaño del archivo adjunto con nombre {nombre}, con tamaño de {peso:.2f} MB es superior al máximo permitido (5 MB).'
            messages.error(request, mensaje)  # Agregar mensaje de error
            return redirect('reactivos:registrar_solicitud')
        
        if form.is_valid():
            solicitud = form.save(commit=False)  # No guardar aún, solo crear una instancia
            solicitud.created_by = request.user
            solicitud.last_updated_by = request.user
            solicitud.save()  # guarda la instancia con los archivos
            form.save()
            # Crea un evento de registrar solicitud
            tipo_evento = 'REGISTRAR SOLICITUD'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)

            mensaje = f'La solicitud se ha registrado correctamente, se a enviado un correo al administrador del sistema el radicado de su solicitud es: {solicitud.id:04}'
            messages.success(request, mensaje)  # Agregar mensaje de éxito
            
            # Enviar correo al usuario que registra la solicitud
            id = solicitud.id
            solicitud_code = base64.urlsafe_b64encode(str(id).encode()).decode()
            suffix = f'UniCLab/solicitudes/estado_solicitud/{solicitud_code}'
            shipping_email=solicitud.created_by.email
            email_type=f'Registro de solicitud'
            initial_message= f'El presente mensaje de correo electrónico es porque recientemente se ha registrado una solicitud en el aplicativo de inventario y gestión de reactivos, con la siguiente información:'
            # enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email, email_type )
            # envío de correo electrónico en segundo plano
        
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
            target=enviar_correo_solicitud, args=(request, suffix, id, initial_message,shipping_email, email_type),)
            correo_thread.start()
            # Enviar correo al administrador del aplicativo
            # Obtener el correo del administrador desde la base de datos
            configuracion = ConfiguracionSistema.objects.first()
            if configuracion:
                admin_email = configuracion.correo_administrador
            else:
                admin_email = 'uniclab_man@unal.edu.co'

            id = solicitud.id
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url=f'{protocol}://{domain}/UniCLab/solicitudes/listado_solicitudes'
            solicitud_code = base64.urlsafe_b64encode(str(id).encode()).decode()
            suffix = f'UniCLab/solicitudes/responder_solicitud/{solicitud_code} o visita: {url}'
            shipping_email=admin_email
            email_type=f'Registro de solicitud'
            initial_message= f'Recientemente se ha registrado una solicitud en el aplicativo de inventario y gestión de reactivos, con la siguiente información:'
            # enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email,email_type )
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
            target=enviar_correo_solicitud, args=(request, suffix, id, initial_message,shipping_email, email_type),)
            correo_thread.start()
        else:
            mensaje = f'La solicitud no se ha podido enviar, por favor consulte el administrador del sistema, error:.'
            error = form.errors
            mensaje = f'{mensaje} {error}'
            messages.error(request, mensaje)  # Agregar mensaje de error

        return redirect('reactivos:registrar_solicitud')
    
# Vista que maneja el detalle de la solicitud
@login_required
def estado_solicitud(request, solicitud_code):
    try:
        solicitud_id = int(base64.urlsafe_b64decode(solicitud_code).decode())
        
    except (ValueError, Solicitudes.DoesNotExist):
        raise Http404("La solicitud no existe")

    solicitud = get_object_or_404(Solicitudes, pk=solicitud_id)
    laboratorio = request.user.lab
    usuario = request.user
    context = {
        'usuario': usuario,
        'laboratorio': laboratorio,

        'solicitud': solicitud
    }
    return render(request, 'solicitudes/estado_solicitud.html', context)


# Vista que maneja la la respuesta de la solicitud
@login_required
def responder_solicitud(request, solicitud_code):
    try:
        solicitud_id = int(base64.urlsafe_b64decode(solicitud_code).decode())
        
    except (ValueError, Solicitudes.DoesNotExist):
        raise Http404("La solicitud no existe")
    solicitud = get_object_or_404(Solicitudes, pk=solicitud_id)

    if request.method == 'POST':
        observations = request.POST.get('observations')
        observations = estandarizar_nombre(observations)
        solicitud.observaciones=observations
        solicitud.tramitado=True
        solicitud.last_updated_by=request.user
        solicitud.usuario_tramita=request.user
        solicitud.fecha_tramite=timezone.now()
        solicitud.save()
        # Crea un evento de responder solicitud
        tipo_evento = 'RESPONDER SOLICITUD'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        # Enviar correo al usuario que registra la solicitud
        id = solicitud.id
        solicitud_code = base64.urlsafe_b64encode(str(id).encode()).decode()
        suffix = f'UniCLab/solicitudes/estado_solicitud/{solicitud_code}'
        shipping_email=solicitud.created_by.email
        initial_message= f'Cordial saludo se ha dado respuesta a una solicitud realizada por usted en el aplicativo de inventario y gestión de reactivos, con la siguiente información:'
        email_type=f'Respuesta a solicitud'
        # enviar_correo_solicitud(request, suffix, id, initial_message,shipping_email, email_type )
        # envío de correo electrónico en segundo plano
        
        # Crear un hilo y ejecutar enviar_correo en segundo plano
        correo_thread = threading.Thread(
        target=enviar_correo_solicitud, args=(request, suffix, id, initial_message,shipping_email, email_type),)
        correo_thread.start()

        # Realiza la depuración automática de registros
        # Registros antiguos: fecha_tramite es más antigua que los días configurados en configuración del sistema.

        # Obtener la configuración del sistema, suponiendo que existe un único registro.
        configuracion_sistema = ConfiguracionSistema.objects.first()

        # Comprobar si se obtuvo la configuración.
        if configuracion_sistema:
            # Usar el campo tiempo_solicitudes de la configuración como max_antiguedad.
            max_antiguedad = timedelta(days=configuracion_sistema.tiempo_solicitudes)
        else:
            # Si no se pudo obtener la configuración, usar un valor predeterminado (30 días, por ejemplo).
            max_antiguedad = timedelta(days=30)

        fecha_limite = timezone.now() - max_antiguedad

        # Obtén las solicitudes antiguas que cumplen con el criterio de antigüedad
        solicitudes_antiguas = Solicitudes.objects.filter(tramitado=True, fecha_tramite__lt=fecha_limite)

        # Eliminar los archivos adjuntos asociados a las solicitudes antiguas
        for solicitud in solicitudes_antiguas:
            if solicitud.archivos_adjuntos:
                # Asegúrate de que el archivo realmente se elimine del disco.
                solicitud.archivos_adjuntos.delete()

        # Eliminar las solicitudes antiguas
        solicitudes_antiguas.delete()
        
        return HttpResponse(f'Se ha dado respuesta a la solicitud de manera correcta y se ha enviado la notificación al usuario.',200)
    
    laboratorio = request.user.lab
    usuario = request.user
    context = {
        'usuario': usuario,
        'laboratorio': laboratorio,
        'solicitud': solicitud,
        'tipo_solictudes':TipoSolicitud.objects.all(),
    }
    return render(request, 'solicitudes/responder_solicitud.html', context)



# La vista "crear_estado" se encarga de gestionar la creación de estados en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_estado.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Estados". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si el estado
# ya existe en la base de datos antes de crearlo. Si el estado es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Estados". Si el estado ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_estado(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        # Verifica si ya existe un registro con el mismo nombre del estado
        if Estados.objects.filter(name=name).exists():
            estado = Estados.objects.get(name=name)
            estado_id = estado.id
            estado_name = estado.name
            messages.error(request, 'Ya existe un estado con nombre ' +
                           estado_name+' id: '+str(estado_id))
            return HttpResponse('Ya existe un registro en la base de datos', status=409)

        estado = Estados.objects.create(

            name=name,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado            

        )
        estado_id = estado.id
        estado_name = estado.name

        messages.success(
            request, 'Se ha creado exitosamente la presentación con nombre '+estado_name+' id: '+str(estado_id))
        # Crea un evento de crear estado
        tipo_evento = 'CREAR ESTADO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
         'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_estado.html', context)

# La vista "crear_almacenamiento_interno" se encarga de gestionar la creación de clasificación almacenamiento_interno. Esta vista toma los datos del formulario 
# existente en el template "crear_almacenamiento_interno.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "almacenamiento_internoC". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la clasifciación almacenamiento_interno
# ya existe en la base de datos antes de crearla. Si la clasificación es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "almacenamiento_internoC". Si la clasificación ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_almacenamiento_interno(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        description = request.POST.get('description')
        description = estandarizar_nombre(description)

        # Verifica si ya existe un registro con el mismo nombre de la marca
        if AlmacenamientoInterno.objects.filter(name=name).exists():
            almacenamiento_interno = AlmacenamientoInterno.objects.get(name=name)
            almacenamiento_interno_id = almacenamiento_interno.id
            messages.error(
                request, 'Ya existe un Almacenamiento Interno con nombre '+name+' id: '+str(almacenamiento_interno_id))
            return HttpResponse('Ya existe un registro en la base de datos', status=409)

        almacenamiento_interno = AlmacenamientoInterno.objects.create(

            name=name,
            description=description,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado

        )
        almacenamiento_interno_id = almacenamiento_interno.id
        # Crea un evento de crear almacenamiento interno
        tipo_evento = 'CREAR ALMACENAMIENTO INTERNO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente el Almacenamiento Interno con nombre '+name+' id: '+str(almacenamiento_interno_id))
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_almacenamiento_interno.html', context)

# La vista "crear_clase_almacenamiento" se encarga de gestionar la creación de Clase de almacenamiento. Esta vista toma los datos del formulario 
# existente en el template "crear_clase_almacenamiento.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "ClaseAlmacenamiento". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la Clase de Almacenamiento
# ya existe en la base de datos antes de crearla. Si la codificación es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "ClaseAlmacenamiento". Si la codificación ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_clase_almacenamiento(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        description = request.POST.get('description')
        description = estandarizar_nombre(description)

        # Verifica si ya existe un registro con el mismo nombre de la marca
        if ClaseAlmacenamiento.objects.filter(name=name).exists():
            clase_almacenamiento = ClaseAlmacenamiento.objects.get(name=name)
            clase_almacenamiento_id = clase_almacenamiento.id
            messages.error(
                request, 'Ya existe una clase de almacenamiento con nombre '+name+' id: '+str(clase_almacenamiento_id))
            return HttpResponse('Ya existe un registro en la base de datos', status=409)

        clase_almacenamiento = ClaseAlmacenamiento.objects.create(

            name=name,
            description=description,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado

        )
        clase_almacenamiento_id = clase_almacenamiento.id
        # Crea un evento de crear clase de almacenamiento
        tipo_evento = 'CREAR CLASE DE ALMACENAMIENTO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente la clase de alamcenamiento con nombre '+name+' id: '+str(clase_almacenamiento_id))
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_clase_almacenamiento.html', context)

#La vista "crear_responsable" se encarga de gestionar la creación de responsables en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_responsables.html" y realiza las operaciones necesarias en la base de datos utilizando el modelo 
# "Responsables". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la el responsable ya existe en la 
# base de datos antes de crearlo, para ello, realiza verificación por nombre, correo electrónico y teléfono. Si este es único, se crea 
# un nuevo registro en la tabla correspondiente utilizando el modelo "Responsables". Si el responsable ya existe, se muestra un mensaje 
# de error o se toma la acción apropiada según los requisitos del sistema.
@login_required
def crear_responsable(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        phone = request.POST.get('phone')
        prefix = request.POST.get('prefix')
        cc = request.POST.get('cc')

        #Obtener Acepta Política
        acceptDataProcessing = request.POST.get('acceptDataProcessing')
        if acceptDataProcessing=="Acepta":
            acceptDataProcessing=True
        elif acceptDataProcessing=="":
            acceptDataProcessing=False

        # Añadir la secuencia de escape "\+" al prefijo
        if prefix.startswith("+"):
            prefix = "\\" + prefix
        # Eliminar los caracteres de escape
        prefix = prefix.strip("\\")

        phone = prefix + phone
        mail = request.POST.get('mail')
        mail = estandarizar_nombre(mail)

        # Verifica si ya existe un registro con el mismo número de cédula, telefono o email de la marca
        if Responsables.objects.filter(cc=cc).exists():
            responsablecc = Responsables.objects.get(cc=cc)
            responsable_name = responsablecc.name
            responsable_mail = responsablecc.mail
            messages.error(
                request, 'Ya existe un responsable el número de cédula registrada, su nombre: '+responsable_name+', correo: '+responsable_mail)
            return HttpResponse('Error al insertar en la base de datos', status=400)
           

        
        if Responsables.objects.filter(phone=phone).exists():
            responsablename = Responsables.objects.get(phone=phone)
            responsable_name = responsablename.name
            responsable_mail = responsablename.mail
            messages.error(
                request, 'Ya existe una responsable con el telefono registrado, su nombre: '+responsable_name+', correo: '+responsable_mail)
            return HttpResponse('Error al insertar en la base de datos', status=400)

        if Responsables.objects.filter(mail=mail).exists():
            responsablename = Responsables.objects.get(mail=mail)
            responsable_name = responsablename.name
            responsable_mail = responsablename.mail
            messages.error(
                request, 'Ya existe una responsable con el email registrado, su nombre: '+responsable_name+', correo: '+responsable_mail)
            return HttpResponse('Error al insertar en la base de datos', status=400)

        responsable = Responsables.objects.create(

            cc=cc,
            name=name,
            phone=phone,
            mail=mail,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
            acceptDataProcessing=acceptDataProcessing,
        )
        # Crea un evento de crear responsable
        tipo_evento = 'CREAR RESPONSABLE'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente el siguiente responsable cc '+cc+' nombre: '+name+', correo: '+mail)
        return HttpResponse('Se ha creado exitosamente el siguiente responsable: '+name, status=200)
        

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_responsable.html', context)

# La vista "crear_marca" se encarga de gestionar la creación de marcas en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_marca.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Marcas". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la marca
# ya existe en la base de datos antes de crearla. Si la marca es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Marcas". Si la marca ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_marca(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        # Verifica si ya existe un registro con el mismo nombre de la marca
        if Marcas.objects.filter(name=name).exists():
            marca = Marcas.objects.get(name=name)
            marca_id = marca.id
            messages.error(
                request, 'Ya existe una marca con nombre '+name+' id: '+str(marca_id))
            return HttpResponse('Ya existe un registro en la base de datos', status=409)


        marca = Marcas.objects.create(

            name=name,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
            

        )
        marca_id = marca.id
        # Crea un evento de crear marca
        tipo_evento = 'CREAR MARCA'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente la marca con nombre '+name+' id: '+str(marca_id))
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_marca.html', context)

# La vista "crear_facultad" se encarga de gestionar la creación de estados en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_facultad.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Facultades". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la facultad
# ya existe en la base de datos antes de crearlo. Si la facultad es único, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Facultades". Si la facultad ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_facultad(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        # Verifica si ya existe un registro con el mismo nombre del estado
        if Facultades.objects.filter(name=name).exists():
            facultad = Facultades.objects.get(name=name)
            facultad_id = facultad.id
            facultad_name = facultad.name
            messages.error(request, 'Ya existe una facultad con nombre ' +
                           facultad_name+' id: '+str(facultad_id))
            return HttpResponse('ya existe un registro en la base de datos', status=409)

        facultad = Facultades.objects.create(

            name=name,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado

        )
        facultad_id = facultad.id
        facultad_name = facultad.name
        # Crea un evento de crear facultad
        tipo_evento = 'CREAR FACULTAD'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(request, 'Se ha creado exitosamente la facultad con nombre ' +
                         facultad_name+' id: '+str(facultad_id))
        return HttpResponse('Operación exitosa', status=201)
    
    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,
        }
    return render(request, 'reactivos/crear_facultad.html', context)

# La vista "crear_destino" se encarga de gestionar la creación de ubicaciones en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_ubicaciones.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Ubicaciones". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la ubicación
# ya existe en la base de datos antes de crearlo. Si este es único, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Ubicaciones". Si la ubicación ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_ubicacion(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        facultad_id = request.POST.get('facultad')
        if not facultad_id.isdigit():
            messages.error(request, 'Por favor seleccione una Facultad primero')
            return HttpResponse('Error al insertar en la base de datos', status=400)

        # Obtiene la instancia de la facultad
        facultad = get_object_or_404(Facultades, id=facultad_id)

        # Verifica si ya existe un registro con el mismo nombre de la asignatura y la misma facultad
        if Ubicaciones.objects.filter(Q(name=name) & Q(facultad=facultad)).exists():
            messages.error(request, 'Ya existe  una ubicación con nombre: '+name+', facultad: '+str(facultad))
            return HttpResponse('Error al insertar en la base de datos', status=400)

        asignatura = Ubicaciones.objects.create(
            name=name,
            facultad=facultad,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
        )
        # Crea un evento de crear asignatura
        tipo_evento = 'CREAR ASIGNATURA'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        
        messages.success(request, 'Se ha creado exitosamente la asignatura/ubicación con nombre: '+name+', facultad: '+str(facultad))
        return HttpResponse('Inserción exitosa', status=200)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,
        'facultades': Facultades.objects.all()
    }
    return render(request, 'reactivos/crear_ubicacion.html', context)

# La vista "crear_destino" se encarga de gestionar la creación de destinos en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_destino.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Destinos". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si el destino
# ya existe en la base de datos antes de crearlo. Si este es único, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Destinos". Si el destino ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_destino(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        
        # Verifica si ya existe un registro con el mismo nombre del destino
        if Destinos.objects.filter(name=name).exists():
            destino = Destinos.objects.get(name=name)
            destino_id = destino.id
            messages.error(request, 'Ya existe un destino llamado ' +
                           name+' con id: '+str(destino_id))
            return HttpResponse('ya existe un registro en la base de datos', status=409)
        destino = Destinos.objects.create(

            name=name,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado

        )
        destino_id = destino.id
        # Crea un evento de crear destino
        tipo_evento = 'CREAR DESTINO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente el destino con nombre '+name+' con id: '+str(destino_id))
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_destino.html', context)

# La vista "crear_laboratorio" se encarga de gestionar la creación de estados en la db. Esta vista toma los datos del formulario 
# existente en el template "crear_laboratorio.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Laboratorios". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si el laboratorio
# ya existe en la base de datos antes de crearlo. Si el laboratorio es único, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Laboratorios". Si el estado ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.
@login_required
def crear_laboratorio(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        name = estandarizar_nombre(name)

        # Verifica si ya existe un registro con el mismo nombre del laboratorio
        if Laboratorios.objects.filter(name=name).exists():
            laboratorio = Laboratorios.objects.get(name=name)
            laboratorio_id = laboratorio.id
            laboratorio_name = laboratorio.name
            messages.error(request, 'Ya existe un laboratorio con nombre ' +
                           laboratorio_name+' id: '+str(laboratorio_id))
            return HttpResponse('ya existe un registro en la base de datos', status=409)

        laboratorio = Laboratorios.objects.create(

            name=name,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado

        )
        # Crea un evento de crear laboratorio
        tipo_evento = 'CREAR LABORATORIO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        laboratorio_id = laboratorio.id
        laboratorio_name = laboratorio.name

        messages.success(request, 'Se ha creado exitosamente el laboratorio con nombre ' +
                         laboratorio_name+' id: '+str(laboratorio_id))
        return HttpResponse('Operación exitosa', status=201)

    laboratorio = request.user.lab

    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

    }
    return render(request, 'reactivos/crear_laboratorio.html', context)

# La vista "crear_walmacen" es responsable de la creación de ubicaciones en el almacén dentro de la base de datos. Los 
# datos se obtienen del formulario presente en el template "crear_walmacen.html", y se realizan las operaciones necesarias 
# en la base de datos utilizando el modelo "Almacenamiento". El objetivo es asegurar la unicidad de los registros, lo cual 
# implica verificar si la ubicación en el almacén ya existe antes de crearla, considerando la clave foránea "lab". Si la 
# ubicación es única para un laboratorio específico, se crea un nuevo registro en la tabla correspondiente utilizando el 
# modelo "Almacenamiento". En caso de que la ubicación ya exista dentro del laboratorio, se muestra un mensaje de error o 
# se toma la acción apropiada según los requisitos del sistema.
@login_required
def crear_walmacen(request):
    if request.method == 'POST':
        laboratorio = request.POST.get('lab')
        if not laboratorio:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un laboratorio válido, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        description = request.POST.get('description')
        description = estandarizar_nombre(description)
        lab = request.POST.get('lab')
        nlab = lab

        try:
            namelab = Laboratorios.objects.get(name=lab)
            lab = namelab
        except Laboratorios.DoesNotExist:
            messages.error(request, "El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            lab = None
            return HttpResponse("No se encuentra en la base de datos", status=404)

       # Verifica si ya existe un registro con el mismo nombre y laboratorio



        if Almacenamiento.objects.filter(name=name, lab=lab).exists():
            w_location = Almacenamiento.objects.get(name=name, lab=lab)
            wlocation_id = w_location.id
            messages.error(request, "Ya existe una ubicación en almacén con nombre "+name+' id: '+str(wlocation_id))
            return HttpResponse('Ya existe un registro en la base de datos', status=409)
        
        wubicaciones = Almacenamiento.objects.create(
            name=name,
            description=description,
            lab=lab,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
        )

        wubicacion_id = wubicaciones.id
        # Crea un evento de ubicación en almacén
        tipo_evento = 'CREAR UBICACION EN ALMACEN'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente la ubicacion en almacén con nombre '+name+' id: '+str(wubicacion_id))
        return HttpResponse('Operación exitosa', status=201)
    laboratorio = request.user.lab

    context = {
        'laboratorio':laboratorio,
        'laboratorios': Laboratorios.objects.all(),
    }
    return render(request, 'reactivos/crear_walmacen.html', context)

# La vista "crear_reactivo" se encarga de gestionar la creación de un reactivo. Esta vista toma los datos del formulario 
# existente en el template "crear_reactivo.html" y realiza las operaciones necesarias en la base de datos para almacenar 
# la información del reactivo. Esto puede incluir la validación de los datos ingresados, la creación de un nuevo registro 
# en la tabla correspondiente y cualquier otra gestión requerida para asegurar la integridad de los datos en la base de datos.
@login_required
def crear_reactivo(request):
    
    if request.method == 'POST':        

        code = request.POST.get('code')
        code = estandarizar_nombre(code)
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        cas = request.POST.get('cas')
        cas = estandarizar_nombre(cas)
        
        state = request.POST.get('state')
        if not state.isdigit():
            messages.error(request, 'Por favor seleccione un estado primero')
            HttpResponse('Por favor seleccione un estado primero', status=400)
        state = get_object_or_404(Estados, id=state)

        unit = request.POST.get('unit')
        if not unit.isdigit():
            messages.error(request, 'Por favor seleccione una unidad primero')
            HttpResponse('Por favor seleccione una unidad primero', status=400)
        unit = get_object_or_404(Unidades, id=unit)

        clase_almacenamiento = request.POST.get('clase_almacenamiento')
        if not clase_almacenamiento.isdigit():
            messages.error(request, 'Por favor seleccione una clase de almacenamiento primero')
            HttpResponse('Por favor seleccione una clase de almacenamiento primero', status=400)
        clase_almacenamiento = get_object_or_404(ClaseAlmacenamiento, id=clase_almacenamiento)

        almacenamiento_interno = request.POST.get('almacenamiento_interno')
        if not almacenamiento_interno.isdigit():
            messages.error(request, 'Por favor seleccione una almacenamiento interno primero')
            HttpResponse('Por favor seleccione una almacenamiento interno primero', status=400)
        almacenamiento_interno = get_object_or_404(AlmacenamientoInterno, id=almacenamiento_interno)

        if Reactivos.objects.filter(name=name).exists():
            reactivo = Reactivos.objects.get(name=name)
            reactivo_name = reactivo.name
            messages.error(
                request, 'Ya existe un reactivo con el nombre registrado: '+reactivo_name)
            return HttpResponse('Ya existe un reactivo con el nombre registrado: ' + reactivo_name, status=400)

        if Reactivos.objects.filter(code=code).exists():
            reactivo = Reactivos.objects.get(code=code)
            reactivo_name = reactivo.name
            messages.error(
                request, 'Ya existe un reactivo con el código registrado: '+reactivo_name)
            return HttpResponse('Ya existe un reactivo con el código registrado: ' + reactivo_name, status=400)
        
        if cas!='NO APLICA':

            if Reactivos.objects.filter(cas=cas).exists():
                reactivo = Reactivos.objects.get(cas=cas)
                reactivo_name = reactivo.name
                messages.error(
                    request, 'Ya existe un reactivo con el CAS registrado: '+reactivo_name)
                return HttpResponse('Ya existe un reactivo con el cas registrado: ' + reactivo_name, status=400)
            
        reactivo = Reactivos.objects.create(
            code=code,
            name=name,
            unit=unit,
            cas=cas,
            state=state,
            clase_almacenamiento=clase_almacenamiento,
            almacenamiento_interno=almacenamiento_interno,
            created_by=request.user,  # Asignar el usuario actualmente autenticado
            last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
        )

        # Crear de contexto para el mensaje de correo
        # Obtener el dominio
        protocol = 'https' if request.is_secure() else 'http'
        domain = request.get_host()
        url=f'{protocol}://{domain}'
        # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
        fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
        # Formatear el valor con separadores de miles
        locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
        user=f'{request.user.first_name} {request.user.last_name}'
                    
        # Contexto del template del diseño del correo 
        context = {
            'url': url,
            'fecha_actual': fecha_actual,
            'reactivo':reactivo,
            'user':user,
        }
        # Template
        message = render_to_string('reactivos/registro_exitoso_reactivo.html', context)
        plain_message = strip_tags(message)
        from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
        current_lab=request.user.lab
        rol=get_object_or_404(Rol,name='COORDINADOR')
        coordinators=User.objects.filter(lab=current_lab, rol=rol, is_active=True)
        recipient_list = [request.user.email]
        recipient_list.extend(coordinator.email for coordinator in coordinators)
        subject = f"Creación exitosa de reactivo {reactivo.name} - {protocol}://{domain}"
        # send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)           
        
        # envío de correo electrónico en segundo plano
        # Archivo adjunto nulo
        attach_path=None
        # Crear un hilo y ejecutar enviar_correo en segundo plano
        correo_thread = threading.Thread(
        target=enviar_correo, args=(recipient_list, subject, message, attach_path),)
        correo_thread.start()

        # Crea un evento de crear reactivo
        tipo_evento = 'CREAR REACTIVO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)


        # Mensaje de éxito al usuario
        messages.success(request, 'Se ha creado exitosamente el reactivo: '+name)
        return HttpResponse('Reactivo creado correctamente: '+name, status=200)
    
    
    else:
        
        # Obtiene el último registro existente
        ultimo_registro = Reactivos.objects.order_by('-id').first()
        
        
        if ultimo_registro:
            # Obtiene el último código y lo divide en partes
            partes = ultimo_registro.code.split('-')

            # Obtiene el último número en el código
            ultimo_numero = partes[-1]

            try:
                # Intenta convertir el último número a entero y aumentarlo en 1
                nuevo_numero = int(ultimo_numero) + 1

                # Reemplaza el último número en las partes
                partes[-1] = str(nuevo_numero)

                # Construye el nuevo código
                nuevo_codigo = '-'.join(partes)

            except ValueError:
                # Si el último número no es un entero, se usa un valor predeterminado
                nuevo_codigo = '1-001-1'

        laboratorio = request.user.lab
        
        context = {
            'laboratorio': laboratorio,
            'laboratorios': Laboratorios.objects.all(),
            'unidades': Unidades.objects.all(),
            'estados': Estados.objects.all(),
            'almacenamiento_internos': AlmacenamientoInterno.objects.all(),
            'clase_almacenamientos': ClaseAlmacenamiento.objects.all(),
            'nuevo_codigo': nuevo_codigo,  # Pasar el nuevo código al contexto
        }

    
    return render(request, 'reactivos/crear_reactivo.html', context)

# La vista "registrar_entrada" se encarga de gestionar el registro de transacciones de entrada en el aplicativo de insumos en la base de 
# datos. Los datos se obtienen del formulario presente en el template "registrar_entrada.html". Utilizando el modelo "Entradas", se 
# realizan las operaciones necesarias. La vista verifica la existencia de los campos de entrada que son foráneos en la base de datos. 
# Si alguno de estos campos es tomado como un nombre o un ID, se convierten en una instancia del modelo foráneo correspondiente. 
# Luego, se verifica en la tabla "Inventarios" si existe algún registro que coincida con los campos "lab", "name", "trademark" y 
# "reference". Si hay una coincidencia, se suma el valor del campo "weight" del registro existente. En caso contrario, se crea un nuevo 
# registro con los valores correspondientes.Finalmente, se realiza la creación del registro de entrada en la base de datos utilizando 
# el modelo "Entradas".
@login_required
def registrar_entrada(request):

    if request.method == 'POST':
        laboratorio = request.POST.get('lab')
        if not laboratorio:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un laboratorio válido, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)
        
        name = request.POST.get('name')
        nReactivo = name
        try:
            nameReactivo = Reactivos.objects.get(name=name)
            name = nameReactivo
        except Reactivos.DoesNotExist:
            messages.error(request, "El reactivo "+nReactivo +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            name = None
            return HttpResponse("El reactivo "+nReactivo +" no se encuentra en la base de datos, favor crearlo primero.", status=400)

        facultad = request.POST.get('facultad')
        location = request.POST.get('location')
        if facultad=='' and location =='':
            location = None
        else:
            facultad=get_object_or_404(Facultades, name=facultad)
            nlocation = location
            try:
                nameLocation = Ubicaciones.objects.get(name=location, facultad=facultad)
                location = nameLocation

            except Ubicaciones.DoesNotExist:
                messages.error(request, "La ubicación "+nlocation +
                               " no se encuentra en la base de datos, favor crearlo primero.")
                location = None
                return HttpResponse("La ubicación "+nlocation +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        correo = request.POST.get('correo')
        
        manager = request.POST.get('manager')
        nmanager = manager
        try:
            nameManager = Responsables.objects.get(name=manager, mail=correo)
            manager = nameManager
        except Responsables.DoesNotExist:
            messages.error(request, "El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            manager = None
            return HttpResponse("El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
                
        trademark_id = request.POST.get('trademark')
        if not trademark_id.isdigit():
            messages.error(request, 'Por favor seleccione una marca primero')
        try:
            nameMarca = Marcas.objects.get(id=trademark_id)
            trademark = nameMarca
        except ObjectDoesNotExist:
            trademark = None
            messages.error(request, 'Por favor seleccione una marca primero')
            return HttpResponse(" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        
        wlocation_id = request.POST.get('wlocation')
        if not wlocation_id.isdigit():
            messages.error(request, 'Por favor seleccione una ubicación en almacén primero')
            return HttpResponse(" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        try:
            nameWlocation = Almacenamiento.objects.get(id=wlocation_id)
            wlocation = nameWlocation
        except ObjectDoesNotExist:
            wlocation = None
            messages.error(request, 'Por favor seleccione una ubicación en almacén primero')
            return HttpResponse("Por favor seleccione una ubicación en almacén primero", status=400)       
        
        destination_id = request.POST.get('destination')

        if not destination_id.isdigit():
            messages.error(request, 'Por favor seleccione un destino primero')
            return HttpResponse("Por favor seleccione una destino en almacén primero", status=400)
        try:
            namedestino = Destinos.objects.get(id=destination_id)
            destination = namedestino
        except ObjectDoesNotExist:
            destination = None
            messages.error(request, 'Por favor seleccione un destino primero')
            return HttpResponse("Por favor seleccione un destino primero", status=400)
        
        lab = request.POST.get('lab')
        nlab = lab
        try:
            namelab = Laboratorios.objects.get(name=lab)
            lab = namelab
        except Laboratorios.DoesNotExist:
            messages.error(request, "El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            lab = None
            return HttpResponse("El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        reference = request.POST.get('reference')
        reference = estandarizar_nombre(reference)

        #verificar que el valor sea positivo
        price = request.POST.get('price')
        if price:
            price=float(price)
            if price<0:
                messages.error(request, 'Solo se permiten registros con precios positivos')
                return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
        else:
            price=None
        
        #Obtener minStockControl
        minStockControl = request.POST.get('minStockControl')
        if minStockControl=="Activo":
            minStockControl=True
        elif minStockControl=="Inactivo":
            minStockControl=False

        
        #verificar que el valor sea positivo
        minstock = request.POST.get('minstock')
        if minstock=='':
            minstock=0

        minstock_number=float(minstock)
        if minstock_number<0:
            messages.error(request, 'Solo se permiten registros con stock mínimo positivo')
            return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
        
        
        # Verificar si el reactivo ya existe en la tabla de inventarios
        try:
            inventario_existente = Inventarios.objects.filter(
                name=name, trademark=trademark, reference=reference, lab=lab).first()
            #Verificar que el peso sea positvo
            weight = request.POST.get('weight')
            weight_number=float(weight)
            if weight_number<=0:
                messages.error(request, 'Solo se permiten registros de cantidades positivas')
                return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)


            edate = request.POST.get('edate')
            # Convertir la cadena de texto en un objeto de fecha
            edate = parse_date(edate)

            # Obtener la fecha actual
            today = datetime.now().date()
            # Obtener la fecha futura (mañana)
            tomorrow = today + timedelta(days=1)

            # Obtener la fecha máxima permitida (31/12/2100)
            max_date = datetime(2100, 12, 31).date()

            # Verificar si la fecha de vencimiento es válida
            if not tomorrow <= edate <= max_date:
		        # Fecha no válida, mostrar el mensaje de error
                tomorrow = tomorrow.strftime('%d/%m/%Y')
                mensaje='Por favor ingrese una fecha válida entre '+str(tomorrow)+' y 31/12/2100'
                messages.error(request, mensaje)
                return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)

            if inventario_existente:
                # Si el reactivo ya existe y está activo(cantidad>0), sumar el peso obtenido del formulario al peso existente
                if inventario_existente.is_active == True:
                    # Si el reactivo ya existe, sumar el peso obtenido del formulario al peso existente
                    weight = request.POST.get('weight')
                    inventario_existente.weight += int(weight)
                    inventario_existente.edate = request.POST.get('edate')
                    inventario_existente.minstock = request.POST.get('minstock')
                    inventario_existente.wlocation = wlocation
                    inventario_existente.minStockControl = minStockControl
                    minstock = request.POST.get('minstock')
                    if minstock=='':
                        minstock=0
                    inventario_existente.minstock = minstock
                    inventario_existente.edate = edate
                    inventario_existente.last_updated_by = request.user
                    inventario_existente.save()
                    # Crea un evento de editar inventario
                    tipo_evento = 'EDITAR INVENTARIO'
                    usuario_evento = request.user
                    crear_evento(tipo_evento, usuario_evento)
                # Si el reactivo ya existe y NO está activo(cantidad00), poner is_active=True y sumar el peso obtenido del formulario al peso existente    
                else:
                    inventario_existente.is_active= True
                    weight = request.POST.get('weight')
                    inventario_existente.weight += int(weight)
                    inventario_existente.edate = request.POST.get('edate')
                    inventario_existente.wlocation = wlocation
                    inventario_existente.minStockControl = minStockControl
                    minstock = request.POST.get('minstock')
                    if minstock=='':
                        minstock=0
                    inventario_existente.minstock = minstock
                    inventario_existente.edate = edate
                    inventario_existente.last_updated_by = request.user

                    inventario_existente.save()
                    # Crea un evento de editar inventario
                    tipo_evento = 'EDITAR INVENTARIO'
                    usuario_evento = request.user
                    crear_evento(tipo_evento, usuario_evento)
            else:
                # Si el reactivo no existe, crear un nuevo registro en la tabla de inventarios
                weight = request.POST.get('weight')

                trademark_id = request.POST.get('trademark')
                try:
                    nameMarca = Marcas.objects.get(id=trademark_id)
                    trademark = nameMarca
                except ObjectDoesNotExist:
                    trademark = None
                    return redirect('reactivos:registrar_entrada')

                name = request.POST.get('name')
                nameReactivo = Reactivos.objects.get(name=name)
                name = nameReactivo
                reference = request.POST.get('reference')
                reference = estandarizar_nombre(reference)

                
                
                minstock = request.POST.get('minstock')
                if minstock=='':
                    minstock=0
                
                inventario = Inventarios.objects.create(
                    name=name,
                    trademark=trademark,
                    weight=weight,
                    reference=reference,
                    lab=lab,
                    wlocation=wlocation,
                    minStockControl=minStockControl,
                    minstock=minstock,
                    edate=edate,
                    created_by=request.user,  # Asignar el usuario actualmente autenticado
                    last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
                
                )
                # Crea un evento de crear inventario
                tipo_evento = 'CREAR INVENTARIO'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

        except Inventarios.DoesNotExist:
            weight = request.POST.get('weight')
        
        try:
            nuevo_inventario = Inventarios.objects.filter(
                name=name, trademark=trademark, reference=reference, lab=lab).first()
            
        except Inventarios.DoesNotExist:
            weight = request.POST.get('weight')
                       
        if name:

            inventario=nuevo_inventario.id
            inventario=get_object_or_404(Inventarios,id=inventario)
            weight = request.POST.get('weight')
            order = request.POST.get('order')
            order = estandarizar_nombre(order)
            orderdate = request.POST.get('orderdate')
            orderdate = parse_date(orderdate)
            observations = request.POST.get('observations')
            observations = estandarizar_nombre(observations)
            unit = request.POST.get('unit')
            nproject = request.POST.get('nproject')
            nproject = estandarizar_nombre(nproject)         
            
            

            entrada = Entradas.objects.create(
                name=name,
                trademark=trademark,
                reference=reference,
                weight=weight,
                location=location,
                order=order,
                date_order=orderdate,
                manager=manager,
                observations=observations,
                nproject=nproject,
                price=price,
                destination=destination,
                lab=lab,
                inventario=inventario,
                created_by=request.user,  # Asignar el usuario actualmente autenticado
                last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
            )
            # Crear de contexto para el mensaje de correo
            # Obtener el dominio
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url=f'{protocol}://{domain}'
            # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
            fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
            # Formatear el valor con separadores de miles
            locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
            if entrada.price:
                formatted_price = locale.format_string('%.2f', float(entrada.price), grouping=True)
            else:
                formatted_price=''
            formatted_weight = locale.format_string('%.2f', float(entrada.weight), grouping=True)
            formatted_minstock = locale.format_string('%.2f', float(inventario.minstock), grouping=True)
            formatted_edate=(inventario.edate).strftime('%d/%m/%Y')
            if entrada.date_order:
                formatted_orderdate=(entrada.date_order).strftime('%d/%m/%Y')
            else:
                formatted_orderdate=None
            user=f'{request.user.first_name} {request.user.last_name}'
                        
            # Contexto del template del diseño del correo 
            context = {
                'url': url,
                'fecha_actual': fecha_actual,
                'formatted_price':formatted_price,
                'formatted_weight':formatted_weight,
                'formatted_edate':formatted_edate,
                'formatted_minstock':formatted_minstock,
                'formatted_orderdate':formatted_orderdate,
                'entrada':entrada,
                'inventario':inventario,
                'user':user,
            }
            # Template
            message = render_to_string('reactivos/registro_exitoso_entrada.html', context)
            plain_message = strip_tags(message)
            from_email = "Notificación Dirección de Laboratorios"  # Agrega el correo electrónico desde el cual se enviará el mensaje
            current_lab=lab
            rol=get_object_or_404(Rol,name='COORDINADOR')
            coordinators=User.objects.filter(lab=current_lab, rol=rol, is_active=True)
            recipient_list = [request.user.email]
            recipient_list.extend(coordinator.email for coordinator in coordinators)
            
            subject = f"Registro exitoso de entrada de inventario de {entrada.name} en {entrada.lab.name} - {protocol}://{domain}"
            # send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)           
            
            attach_path=None

            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
                target=enviar_correo,
                args=(recipient_list, subject, message, attach_path),
            )
            correo_thread.start()

            # enviar_correo(recipient_list, subject, message, attach_path)
            
            # Crea un evento de registrar entrada
            tipo_evento = 'REGISTRAR ENTRADA'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)
            
            messages.success(request, 'Se ha registrado de manera exitosa el ingreso del insumo: ' +
                             nReactivo+', cantidad '+weight+' '+unit)
            return HttpResponse('Se ha registrado de manera exitosa el ingreso del: ' +
                             nReactivo+', cantidad '+weight+' '+unit, status=200)
    laboratorio = request.user.lab
    # Obtener la fecha de hoy
    today = date.today()
    # Calcular la fecha ayer
    yesterday = today - timedelta(days=1)
    tomorrow = today + timedelta(days=1)        
    context = {
                'reactivos': Reactivos.objects.all(),
                'responsables': Responsables.objects.all(),
                'marcas': Marcas.objects.all(),
                'ubicaiones': Ubicaciones.objects.all(),
                'destinos':Destinos.objects.all(),
                'laboratorios':Laboratorios.objects.all(),
                'wubicaciones':Almacenamiento.objects.all(),
                'usuarios': User.objects.all(),
                'laboratorio': laboratorio,
                'usuarios': User.objects.all(),
                'tomorrow': tomorrow,
                'yesterday': yesterday,
            }
        
    return render(request, 'reactivos/registrar_entrada.html', context)

# La vista "registrar_salida" se encarga de gestionar el registro de transacciones de salida en el aplicativo de insumos de la base de 
# datos. Los datos se obtienen del formulario presente en el template "registrar_salida.html" y se utilizan el modelo "Salidas" para 
# realizar las operaciones correspondientes de inserción de registros.
# En esta vista se verifica la existencia de los campos de entrada que son foráneos en la base de datos. En caso de que alguno de estos 
# campos sea tomado como un nombre o un ID, se convierten en una instancia del modelo foráneo correspondiente. Luego, se realiza una 
# verificación en la tabla "Inventarios" para comprobar si existe algún registro que coincida con los campos "lab", "name", "trademark" 
# y "reference".
#Si se encuentra una coincidencia en "Inventarios", se resta el valor del campo "weight" del registro existente. Si no hay una 
# coincidencia, se impide el registro tanto en la tabla "Inventarios" como en la tabla "Salidas". Además, se verifica que el campo 
# "weight" sea mayor o igual para permitir la transacción. En caso contrario, se muestra un mensaje de error. Adicionalmente, si el 
# campo "weight" llega a cero después de realizar el registro, se muestra un mensaje indicando que el insumo ha alcanzado el valor de cero.
#Finalmente, se realiza la creación del registro de salida en la base de datos utilizando el modelo "Salidas".
@login_required
def registrar_salida(request):

    if request.method == 'POST': 
        warning=""

        laboratorio = request.POST.get('lab')
        if not laboratorio:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un laboratorio válido, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)

        #Verificar que el peso sea positvo
        weight = request.POST.get('weight')
        weight_number=float(weight)
        if weight_number<=0:
            messages.error(request, 'Solo se permiten registros de cantidades positivas')
            return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
        
        name = request.POST.get('name')
        nReactivo = name
        try:
            nameReactivo = Reactivos.objects.get(name=name)
            name = nameReactivo
        except Reactivos.DoesNotExist:
            messages.error(request, "El reactivo "+nReactivo +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            name = None
            return HttpResponse("El reactivo "+nReactivo +" no se encuentra en la base de datos, favor crearlo primero.", status=400)

        facultad = request.POST.get('facultad')
        print(facultad)
        facultad=get_object_or_404(Facultades, name=facultad)
        
        location = request.POST.get('location')
        nlocation = location
        try:
            nameLocation = Ubicaciones.objects.get(name=location,facultad=facultad)
            location = nameLocation
        except Ubicaciones.DoesNotExist:
            messages.error(request, "La ubicación "+nlocation +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            location = None
            return HttpResponse("La ubicación "+nlocation +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        manager = request.POST.get('manager')
        correo = request.POST.get('correo')
        nmanager = manager
        try:
            nameManager = Responsables.objects.get(name=manager, mail=correo)
            manager = nameManager
        except Responsables.DoesNotExist:
            messages.error(request, "El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            manager = None
            return HttpResponse("El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        trademark_id = request.POST.get('trademark')
        if not trademark_id.isdigit():
            messages.error(request, 'No fue posible realizar su registro: no seleccionó una marca, por favor verifique.')
            return HttpResponse("Error al insertar en la base de datos", status=400)
                     

        try:
            nameMarca = Marcas.objects.get(id=trademark_id)
            trademark = nameMarca
        except ObjectDoesNotExist:
            trademark = None
            messages.error(request, 'La marca no coincide, por favor revise.')
            return redirect('reactivos:registrar_salida')

        destination_id = request.POST.get('destination')
        if not destination_id.isdigit():
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un destino, por favor verifique.')
            return HttpResponse("Error al insertar en la base de datos", status=400)
        try:
            namedestino = Destinos.objects.get(id=destination_id)
            destination = namedestino
        except ObjectDoesNotExist:
            destination = None
            messages.error(request, 'Por favor seleccione un destino primero')
            return redirect('reactivos:registrar_salida')
        
        lab = request.POST.get('lab')
        nlab = lab
        try:
            namelab = Laboratorios.objects.get(name=lab)
            lab = namelab
        except Laboratorios.DoesNotExist:
            messages.error(request, "El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            lab = None
            return HttpResponse("El responsable "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        reference = request.POST.get('reference')
        if not reference:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó una referencia, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)
        
        # Verificar si el reactivo ya existe en la tabla de inventarios
        try:
            inventario_existente = Inventarios.objects.filter(
                name=name, trademark=trademark, reference=reference, lab=lab).first()

            if inventario_existente:
                
                weight = request.POST.get('weight')
                weight=Decimal(weight)
                unit = request.POST.get('unit')
                #Verificar si la cantidad actual sea mayor o igual a cantidad registrada
                if inventario_existente.weight>=weight:
                    inventario_existente.weight -= int(weight)
                    inventario_existente.last_updated_by = request.user
                    inventario_existente.save()
                    # Crea un evento de editar inventario
                    tipo_evento = 'EDITAR INVENTARIO'
                    usuario_evento = request.user
                    crear_evento(tipo_evento, usuario_evento)
                    #Verificación después de restar en la tabla llegue a cero y ponga en warning la alerta que 
                    # posteriormente se enviará al usuario informando
                    if inventario_existente.weight == 0:
                        inventario_existente.is_active = False  # Asignar False a la columna is_active
                        inventario_existente.last_updated_by = request.user
                        inventario_existente.save()
                        
                        #preparar alerta para mensaje al usuario
                        warning=", pero el inventario actual ha llegado a 0. Favor informar al coordinador de laboratorio."
                         # Envío de correo al coordinador del laboratorio
                        alerta='Cantidad reactivo ha llegado a cero'
                        reactivo=f'{inventario_existente.name}'
                        marca=f'{inventario_existente.trademark}'
                        referencia=f'{inventario_existente.reference}'
                        cantidad=f'{inventario_existente.weight}'
                        unidad=f'{inventario_existente.name.unit}'
                        mensaje=f'La cantidad de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cero en inventario ({cantidad} {unidad}).'
                        # enviar_correo_alerta(request, alerta, reactivo, marca, referencia, cantidad, unidad, mensaje, inventario_existente)
                        # envío de correo electrónico en segundo plano        
                        # Crear un hilo y ejecutar enviar_correo en segundo plano
                        correo_thread = threading.Thread(
                        target=enviar_correo_alerta, args=(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente),)
                        correo_thread.start()

                    if (inventario_existente.weight<=inventario_existente.minstock) and inventario_existente.minStockControl==True and inventario_existente.weight>0:
                        # preparar alerta para mensaje al usuario    
                        warning=", pero el inventario actual es menor o igual que el stock mínimo para este reactivo. Favor informar al coordinador de laboratorio."
                        # Envío de correo al coordinador del laboratorio
                        alerta='Cantidad reactivo por debajo de inventario mínimo'
                        reactivo=f'{inventario_existente.name}'
                        marca=f'{inventario_existente.trademark}'
                        referencia=f'{inventario_existente.reference}'
                        cantidad=f'{inventario_existente.weight}'
                        unidad=f'{inventario_existente.name.unit}'
                        mensaje=f'El reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cantidades críticas ({cantidad} {unidad}), por debajo del stock mínimo ({inventario_existente.minstock} {unidad}).'
                        # enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente)
                        
                        # envío de correo electrónico en segundo plano
        
                        # Crear un hilo y ejecutar enviar_correo en segundo plano
                        correo_thread = threading.Thread(
                        target=enviar_correo_alerta, args=(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente),)
                        correo_thread.start()

                else:
                    inventario_existente.weight=int(inventario_existente.weight)
                    messages.error(request, "No es posible realizar la salida del reactivo "+inventario_existente.name.name+": Inventario actual: " + str(inventario_existente.weight) + ", " + unit + " Cantidad solicitada: " + str(weight) + " " + unit)
                    return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
            else:
                
                messages.error(request, "Los valor seleccionados (Reactivo, Marca o referencia) no corresponden a un insumo existente en el inventario, verifique de nuevo")
                return HttpResponse("Error al insertar en la base de datos", status=400)

        except Inventarios.DoesNotExist:
            weight = request.POST.get('weight')

        try:
            nuevo_inventario = Inventarios.objects.filter(
                name=name, trademark=trademark, reference=reference, lab=lab).first()
            
        except Inventarios.DoesNotExist:
            weight = request.POST.get('weight')
                       
        if name:

            inventario=nuevo_inventario.id
            inventario=nuevo_inventario
            reference = request.POST.get('reference')
            weight = request.POST.get('weight')
            observations = request.POST.get('observations')
            observations = estandarizar_nombre(observations)
            unit = request.POST.get('unit')
            

            salida = Salidas.objects.create(
                inventario=inventario,
                name=name,
                trademark=trademark,
                reference=reference,
                weight=weight,
                location=location,
                manager=manager,
                observations=observations,
                destination=destination,
                lab=lab,
                created_by=request.user,  # Asignar el usuario actualmente autenticado
                last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
            )
            # Crear de contexto para el mensaje de correo
            # Obtener el dominio
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url=f'{protocol}://{domain}'
            # Obtener la fecha y hora actual en el formato "dd/mm/aaaa hh:mm:ss"
            fecha_actual = localtime().strftime('%d/%m/%Y %H:%M:%S')
            # Formatear el valor con separadores de miles
            locale.setlocale(locale.LC_ALL, 'es_ES.UTF-8')
            formatted_weight = locale.format_string('%.2f', float(salida.weight), grouping=True)
            user=f'{request.user.first_name} {request.user.last_name}'
                        
            # Contexto del template del diseño del correo 
            context = {
                'url': url,
                'fecha_actual': fecha_actual,
                'formatted_weight':formatted_weight,
                'salida':salida,
                'inventario':inventario,
                'user':user,
            }
            # Template
            message = render_to_string('reactivos/registro_exitoso_salida.html', context)
            plain_message = strip_tags(message)
            from_email = "noreply@unal.edu.co"  # Agrega el correo electrónico desde el cual se enviará el mensaje
            current_lab=request.user.lab
            rol=get_object_or_404(Rol, name='COORDINADOR')
            coordinators=User.objects.filter(lab=current_lab, rol=rol, is_active=True)
            recipient_list = [request.user.email]
            recipient_list.extend(coordinator.email for coordinator in coordinators)
            subject = f"Registro exitoso de salida de inventario de {salida.name} en {salida.lab.name} - {protocol}://{domain}"
            # send_mail(subject, plain_message, from_email, recipient_list, fail_silently=False, html_message=message)           
            # envío de correo electrónico en segundo plano
            # Archivo adjunto nulo
            attach_path=None
            # Crear un hilo y ejecutar enviar_correo en segundo plano
            correo_thread = threading.Thread(
            target=enviar_correo, args=(recipient_list, subject, message, attach_path),)
            correo_thread.start()


            # Crea un evento de registrar salida
            tipo_evento = 'REGISTRAR SALIDA'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)
            # Mensaje de confirmación al usuario
            messages.success(request, 'Se ha registrado de manera exitosa la salida del insumo del insumo: ' +
                             nReactivo+', cantidad '+weight+' '+unit+warning)
            return HttpResponse('Se ha registrado de manera exitosa la salida del insumo : ' +
                             nReactivo+', cantidad '+weight+' '+unit+warning, status=200)
    laboratorio = request.user.lab

    context = {
                'laboratorio':laboratorio,
                'usuarios': User.objects.all(),
                'reactivos': Reactivos.objects.all(),
                'responsables': Responsables.objects.all(),
                'marcas': Marcas.objects.all(),
                'ubicaiones': Ubicaciones.objects.all(),
                'destinos':Destinos.objects.all(),
                'referencias':Inventarios.objects.all(),
                'laboratorios':Laboratorios.objects.all(),
            }
        
    return render(request, 'reactivos/registrar_salida.html', context)

# La vista "editar_reactivo" se encarga de gestionar la edición  de un registro reactivo en la db. Esta vista toma los datos del formulario 
# existente en el template "editar_reactivo.html" y realiza las operaciones necesarias en la base de datos para editar 
# la información del reactivo. Esto puede incluir la validación de los datos ingresados, la edición de un nuevo registro 
# en la tabla correspondiente y cualquier otra gestión requerida para asegurar la integridad de los datos en la base de datos.
@login_required
def editar_reactivo(request, pk):
    reactivo = get_object_or_404(Reactivos, pk=pk)
    
    if request.method == 'POST':
        
        code = request.POST.get('code')
        code = estandarizar_nombre(code)
        name = request.POST.get('name')
        name = estandarizar_nombre(name)
        cas = request.POST.get('cas')
        cas = estandarizar_nombre(cas)
        
        state = request.POST.get('state')
        unit = request.POST.get('unit')
        clase_almacenamiento = request.POST.get('clase_almacenamiento')
        almacenamiento_interno = request.POST.get('almacenamiento_interno')

        if not state.isdigit():
            
            HttpResponse('Por favor seleccione un estado primero', status=400)
        state = get_object_or_404(Estados, id=state)     
        
        if not unit.isdigit():
           
            HttpResponse('Por favor seleccione una unidad primero', status=400)
        unit = get_object_or_404(Unidades, id=unit)
        
        if not clase_almacenamiento.isdigit():
            
            HttpResponse('Por favor seleccione una clase de almacenamiento primero', status=400)
        clase_almacenamiento = get_object_or_404(ClaseAlmacenamiento, id=clase_almacenamiento)

        
        if not almacenamiento_interno.isdigit():
            
            HttpResponse('Por favor seleccione una almacenamiento interno primero', status=400)
        almacenamiento_interno = get_object_or_404(AlmacenamientoInterno, id=almacenamiento_interno)
        if cas!='NO APLICA':
            if (name != reactivo.name and Reactivos.objects.filter(name=name).exists()) or \
               (code != reactivo.code and Reactivos.objects.filter(code=code).exists()) or \
               (cas != reactivo.cas and Reactivos.objects.filter(cas=cas).exists()):
                return HttpResponse('No fue posible realizar la edición ya que los valores de Nombre, Código o CAS coinciden con otros registros existentes.', status=400)                 
        else:
            if (name != reactivo.name and Reactivos.objects.filter(name=name).exists()) or \
               (code != reactivo.code and Reactivos.objects.filter(code=code).exists()):
                return HttpResponse('No fue posible realizar la edición ya que los valores de Nombre, Código o CAS coinciden con otros registros existentes.', status=400)             
        reactivo.code=code
        reactivo.cas=cas
        reactivo.name=name
        reactivo.unit=unit
        reactivo.state=state
        reactivo.clase_almacenamiento=clase_almacenamiento
        reactivo.almacenamiento_interno=almacenamiento_interno
        reactivo.last_updated_by=request.user  # Asignar el usuario actualmente autenticado


        reactivo.save()
        # Crea un evento de editar reactivo
        tipo_evento = 'EDITAR REACTIVO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        
        return HttpResponse('Reactivo editado correctamente: '+name, status=200)
    
    laboratorio = request.user.lab
    
    context = {
        'reactivo':reactivo,
        'laboratorio':laboratorio,
        'laboratorios': Laboratorios.objects.all(),
        'unidades': Unidades.objects.all(),
        'estados': Estados.objects.all(),
        'almacenamiento_internos': AlmacenamientoInterno.objects.all(),
        'clase_almacenamientos': ClaseAlmacenamiento.objects.all(),
    }
    
    return render(request, 'reactivos/editar_reactivo.html', context)






# Esta vista edita el registro de entrada, además según los valores edita a su vez el modelo Inventarios
# de manera que una afectación en la entrada, afecta directamente el inventario

@login_required
def editar_entrada(request, pk):
       
    # Recupera la entrada o muestra una página de error 404 si no se encuentra
    entrada = get_object_or_404(Entradas, pk=pk)
    warning = ''

     
    if request.method == 'POST':
        # Recolección y validación de datos paar edición de registro de entrada
        # lab
        lab = request.POST.get('lab')
        if not lab:
            return HttpResponse('No fue posible realizar la edición de su registro: no seleccionó un laboratorio válido, por favor verifique.', status=400)
        
        nlab = lab
        try:
            namelab = Laboratorios.objects.get(name=lab)
            lab = namelab
        except Laboratorios.DoesNotExist:
            lab = None
            return HttpResponse("El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        #name (reactivo)
        name = request.POST.get('name')
        nReactivo = name
        try:
            nameReactivo = Reactivos.objects.get(name=name)
            name = nameReactivo
        except Reactivos.DoesNotExist:
            name = None
            return HttpResponse("El reactivo "+nReactivo +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        #ubicación y asignatura
        facultad = request.POST.get('facultad')        
        location = request.POST.get('location')
        
        if facultad and location:
            facultad=get_object_or_404(Facultades, name=facultad)
            nlocation = location
            try:
                nameLocation = Ubicaciones.objects.get(name=location, facultad=facultad)
                location = nameLocation
            except Ubicaciones.DoesNotExist:
                location = None
                return HttpResponse("La ubicación "+nlocation +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        else:
            location=None
            print('Hola Mundo')
        # Responsable y correo
        correo = request.POST.get('correo')
        manager = request.POST.get('manager')
        nmanager = manager
        try:
            nameManager = Responsables.objects.get(name=manager, mail=correo)
            manager = nameManager
        except Responsables.DoesNotExist:
            manager = None
            return HttpResponse("El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        #Marca        
        trademark_id = request.POST.get('trademark')
        if not trademark_id.isdigit():
            return HttpResponse("Por favor selrccione una marca primero.", status=400)
        try:
            nameMarca = Marcas.objects.get(id=trademark_id)
            trademark = nameMarca
        except ObjectDoesNotExist:
            trademark = None
            return HttpResponse("La Marca no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        # Destino
        destination_id = request.POST.get('destination')
        if not destination_id.isdigit():
            messages.error(request, 'Por favor seleccione un destino primero')
            return HttpResponse("Por favor seleccione una destino en almacén primero", status=400)
        try:
            namedestino = Destinos.objects.get(id=destination_id)
            destination = namedestino
        except ObjectDoesNotExist:
            destination = None
            return HttpResponse("Por favor seleccione un destino primero", status=400)
        # Referencias
        reference = request.POST.get('reference')
        reference = estandarizar_nombre(reference)
        
        # Precio - verificar que el valor sea positivo
        price = request.POST.get('price')
        if price:
            price_number=float(price)
            if price_number<0:
                return HttpResponse("Solo se permiten registros con precios positivos", status=400)
        else:
            price=None
        # Cantidad
        weight = request.POST.get('weight')
        weight=float(weight)
        if weight<=0:
            return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
        # Orden de compra
        order = request.POST.get('order')
        order = estandarizar_nombre(order)
        # Fecha de orden de compra
        orderdate = request.POST.get('orderdate')
        orderdate = parse_date(orderdate)
        # Observaciones
        observations = request.POST.get('observations')
        observations = estandarizar_nombre(observations)
        # Unidades (Solo para la respuesta del servidor)
        unit = request.POST.get('unit')
        # Nuemero de proyecto
        nproject = request.POST.get('nproject')
        nproject = estandarizar_nombre(nproject)

         #Ubicación en Almacen
        wlocation_id = request.POST.get('wlocation')
        if not wlocation_id.isdigit():
            messages.error(request, 'Por favor seleccione una ubicación en almacén primero')
            return HttpResponse(" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        try:
            nameWlocation = Almacenamiento.objects.get(id=wlocation_id)
            wlocation = nameWlocation
        except ObjectDoesNotExist:
            wlocation = None
            messages.error(request, 'Por favor seleccione una ubicación en almacén primero')
            return HttpResponse("Por favor seleccione una ubicación en almacén primero", status=400)
        
        #Fecha de Vencimiento
        edate = request.POST.get('edate')
        # Convertir la cadena de texto en un objeto de fecha
        edate = parse_date(edate)

        # Obtener la fecha actual
        today = datetime.now().date()
        # Obtener la fecha futura (mañana)
        tomorrow = today + timedelta(days=1)

           
        
        
        # Obtener minStockControl
        minStockControl = request.POST.get('minStockControl')
        if minStockControl=="Activo":
            minStockControl=True
        elif minStockControl=="Inactivo":
            minStockControl=False
            
        #verificar que el valor sea positivo
        minstock = request.POST.get('minstock')
        if minstock=='':
            minstock=0

        minstock_number=float(minstock)
        if minstock_number<0:
            messages.error(request, 'Solo se permiten registros con stock mínimo positivo')
            return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)


        # Restar entrada anterior de Inventario
        id_inv=entrada.inventario.id
        inventario= get_object_or_404(Inventarios,id=id_inv)
        inventario_anterior=inventario.weight
        entrada_anterior=entrada.weight
        cantidad_restada=inventario_anterior-(entrada_anterior)
        inventario.weight=cantidad_restada
        inventario.save()
        
        
        # Verificar si los nuevos valores de Lab, Name, Trademark, y Reference coincidan con algún registro en el inventario
        try:
            inventario_existente = Inventarios.objects.filter(
                name=name, trademark=trademark, reference=reference, lab=lab).first()
            # si coincidenden los datos se actualiza registro
            if inventario_existente:
                # Suma al inventario exitente (ya restado cantidad anterior) el nuevo valor de cantidad en la edición
                inventario_existente.weight += int(weight)
                # Si el inventario queda menor a 0 devulve los cambios en este y retorna
                if inventario_existente.weight<0:
                    inventario_existente.weight -= int(weight)
                    inventario_existente.weight += entrada_anterior
                    inventario_existente.save()
                    return HttpResponse('No se puede realizar la edición del registro ya que esta hace que el inventario sea menor que 0', 400)

                inventario_existente.is_active= True
                inventario_existente.edate = edate
                inventario_existente.minstock = minstock
                inventario_existente.wlocation = wlocation
                inventario_existente.minStockControl = minStockControl
                inventario_existente.last_updated_by = request.user

                # Si después de realizar el proceso de actualización el inventario queda en 0 inactiva y pasa alerta
                if inventario_existente.weight == 0:
                    inventario_existente.is_active = False
                    warning=' pero la cantidad en inventario a llegado a 0, por favor verifique y comuniquese con el coordinador de área'
                    
                    # Envío de correo al coordinador del laboratorio
                    alerta='Cantidad reactivo ha llegado a cero'
                    reactivo=f'{inventario_existente.name}'
                    marca=f'{inventario_existente.trademark}'
                    referencia=f'{inventario_existente.reference}'
                    cantidad=f'{inventario_existente.weight}'
                    unidad=f'{inventario_existente.name.unit}'
                    mensaje=f'Con el proceso de edición de entrada de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cero en inventario ({cantidad} {unidad}).'
                    enviar_correo_alerta(request, alerta, reactivo, marca, referencia, cantidad, unidad, mensaje, inventario_existente)
                # Si el reactivo ya existe y NO está activo(cantidad00), poner is_active=True y sumar el peso obtenido del formulario al peso existente    
                if inventario_existente.minStockControl and inventario_existente.weight > 0 and float(inventario_existente.weight) <= float(inventario_existente.minstock):
                    warning=' pero la cantidad en inventario quedó por debajo del mínimo, por favor verifique y comuniquese con el coordinador de área'

                    # Envío de correo al coordinador del laboratorio
                    alerta='Cantidad reactivo por debajo de inventario mínimo'
                    reactivo=f'{inventario_existente.name}'
                    marca=f'{inventario_existente.trademark}'
                    referencia=f'{inventario_existente.reference}'
                    cantidad=f'{inventario_existente.weight}'
                    unidad=f'{inventario_existente.name.unit}'
                    mensaje=f'Con el proceso de edición de la entrada de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cantidades críticas ({cantidad} {unidad}), por debajo del stock mínimo ({inventario_existente.minstock} {unidad}).'
                    enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente)
                inventario_existente.save()
            
            # Crea un evento de editar inventario
            tipo_evento = 'EDITAR INVENTARIO'
            usuario_evento = request.user
            crear_evento(tipo_evento, usuario_evento)                    
            # Si no existe en la base de datos se creará un nuevo registro de inventario    
            if not inventario_existente:
                inventario = Inventarios.objects.create(
                    name=name,
                    trademark=trademark,
                    weight=weight,
                    reference=reference,
                    lab=lab,
                    wlocation=wlocation,
                    minStockControl=minStockControl,
                    minstock=minstock,
                    edate=edate,
                    created_by=request.user,  # Asignar el usuario actualmente autenticado
                    last_updated_by=request.user,  # Asignar el usuario actualmente autenticado
                    )
                # Crea un evento de crear inventario
                tipo_evento = 'CREAR INVENTARIO'
                usuario_evento = request.user
                crear_evento(tipo_evento, usuario_evento)

            

        
        except Inventarios.DoesNotExist:
            print('Hola los valores no coinciden')
            
        entrada.name = name
        entrada.trademark = trademark
        entrada.reference = reference
        entrada.weight = weight
        entrada.location = location
        entrada.order = order
        entrada.date_order = orderdate
        entrada.manager = manager
        entrada.observations = observations
        entrada.nproject = nproject
        entrada.price = price
        entrada.destination = destination
        entrada.lab = lab
        entrada.inventario=inventario
        entrada.last_updated_by=request.user

        # Guarda los cambios en la entrada existente
        entrada.save()
        mensaje=f'Se ha editado de manera exitosa la entrada del: {nReactivo} , cantidad {weight} {unit}'
        mensaje+=warning
        # Crea un evento de editar entrada
        tipo_evento = 'EDITAR ENTRADA'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        # Envía una respuesta de éxito (puedes personalizar el mensaje según tu necesidad)
        return HttpResponse(mensaje, status=200)

    #----------------------------------------------------------------------#

        
    
    
    # Contexto para incluir valores previos en el template
    
    if entrada.price:
        precio = int(entrada.price)
    elif entrada.price==0:
        precio=0
    else:
        precio =''
    # Formatear la fecha en el formato "YYYY-MM-DD"
    vdate = django_date(entrada.inventario.edate, "Y-m-d")
    if entrada.inventario.minStockControl:
        controlMinStock="Activo"
        minStockControl="checked"
    else:
        controlMinStock="Inactivo"
        minStockControl="unchecked"
    
    weight = int(entrada.weight)
    minstock = int(entrada.inventario.minstock)
    date_order=django_date(entrada.date_order, "Y-m-d")
    
    laboratorio = request.user.lab
    # Obtener la fecha de hoy
    today = date.today()
    # Calcular la fecha ayer
    yesterday = today - timedelta(days=1)
    tomorrow = today + timedelta(days=1)
    
    context = {
        'entrada': entrada,
        'precio':precio,
        'reactivos': Reactivos.objects.all(),
        'vdate':vdate,
        'controlMinStock':controlMinStock,
        'minStockControl':minStockControl,
        'weight':weight,
        'minstock':minstock,
        'date_order':date_order,

        'responsables': Responsables.objects.all(),
        'marcas': Marcas.objects.all(),
        'ubicaiones': Ubicaciones.objects.all(),
        'destinos':Destinos.objects.all(),
        'laboratorios':Laboratorios.objects.all(),
        'wubicaciones':Almacenamiento.objects.all(),
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,
        'usuarios': User.objects.all(),
        'tomorrow': tomorrow,
        'yesterday': yesterday,
        
    }

    return render(request, 'reactivos/editar_entrada.html', context)

# Esta vista edita el registro de salidas, además según los valores edita a su vez el modelo Inventarios
# de manera que una afectación en la salida, afecta directamente el inventario

@login_required
def editar_salida(request, pk):
    # Recupera la entrada o muestra una página de error 404 si no se encuentra
    salida = get_object_or_404(Salidas, pk=pk)
    if request.method == 'POST': 
        # Alerta blanco en caso de mensajes adicionales
        warning=""
        # Tomar datos y realizar validaciones

        # Laboratorio
        laboratorio = request.POST.get('lab')
        if not laboratorio:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un laboratorio válido, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)
        
        nlab = laboratorio
        try:
            namelab = Laboratorios.objects.get(name=laboratorio)
            laboratorio = namelab
        except Laboratorios.DoesNotExist:
            messages.error(request, "El Laboratorio "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            laboratorio = None
            return HttpResponse("El responsable "+nlab +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)

        # Cantidad Verificar que el peso sea positvo
        weight = request.POST.get('weight')
        weight_number=float(weight)
        if weight_number<=0:
            messages.error(request, 'Solo se permiten registros de cantidades positivas')
            return HttpResponse("Error de cantidades al insertar en la base de datos", status=400)
        
        name = request.POST.get('name')
        nReactivo = name
        try:
            nameReactivo = Reactivos.objects.get(name=name)
            name = nameReactivo
        except Reactivos.DoesNotExist:
            messages.error(request, "El reactivo "+nReactivo +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            name = None
            return HttpResponse("El reactivo "+nReactivo +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        # Ubicación y Facultad
        facultad = request.POST.get('facultad')
        facultad=get_object_or_404(Facultades, name=facultad)
        
        location = request.POST.get('location')
        nlocation = location
        try:
            nameLocation = Ubicaciones.objects.get(name=location,facultad=facultad)
            location = nameLocation
        except Ubicaciones.DoesNotExist:
            messages.error(request, "La ubicación "+nlocation +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            location = None
            return HttpResponse("La ubicación "+nlocation +" no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        # Responsable y correo
        manager = request.POST.get('manager')
        correo = request.POST.get('correo')
        nmanager = manager
        try:
            nameManager = Responsables.objects.get(name=manager, mail=correo)
            manager = nameManager
        except Responsables.DoesNotExist:
            messages.error(request, "El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.")
            manager = None
            return HttpResponse("El responsable "+nmanager +
                           " no se encuentra en la base de datos, favor crearlo primero.", status=400)
        #Marca        
        trademark_id = request.POST.get('trademark')
        if not trademark_id.isdigit():
            return HttpResponse("Por favor selrccione una marca primero.", status=400)
        try:
            nameMarca = Marcas.objects.get(id=trademark_id)
            trademark = nameMarca
        except ObjectDoesNotExist:
            trademark = None
            return HttpResponse("La Marca no se encuentra en la base de datos, favor crearlo primero.", status=400)
        
        # Referencia
        reference = request.POST.get('reference')
        if not reference:
            messages.error(request, 'No fue posible realizar su registro: no seleccionó una referencia, por favor verifique.')
            return HttpResponse("Error al consultar en la base de datos", status=400)
        
        # Destino
        destination_id = request.POST.get('destination')
        if not destination_id.isdigit():
            messages.error(request, 'No fue posible realizar su registro: no seleccionó un destino, por favor verifique.')
            return HttpResponse("Error al insertar en la base de datos", status=400)
        try:
            namedestino = Destinos.objects.get(id=destination_id)
            destination = namedestino
        except ObjectDoesNotExist:
            destination = None
            messages.error(request, 'Por favor seleccione un destino primero')
            return redirect('reactivos:registrar_salida')
        
        # Observaciones
        observations = request.POST.get('observations')
        observations = estandarizar_nombre(observations)
             
        # Verificar si el reactivo ya existe en la tabla de inventarios
        try:
            inventario_existente = Inventarios.objects.filter(name=name, trademark=trademark, reference=reference, lab=laboratorio).first()

            if inventario_existente:
                # Eliminar la salida anterior
                # Obtener salida anterior
                salida_anterior=salida.weight
                
                # Obtener inventario anterior
                inventario_anterior=salida.inventario.weight 
                
                # Eliminar Salida Anterior
                inventario_restado= inventario_anterior+salida_anterior
                
                
                # Restar del inventario anterior la salida actual
                inventario_existente.weight=inventario_restado-int(weight)
                
                # si el inventario quedó menor que 0 devolver error
                if inventario_existente.weight < 0:
                    return HttpResponse("No se puede realizar la edición del registro ya esta hace que el inventario sea menor que 0.", 400)
                 
                else:
                    inventario_existente.is_active=True
                    # Si el inventario es mayor o igual a 0 verificar si este llega a 0
                    if inventario_existente.weight==0:
                        warning=', pero esta a hecho que el inventario llegue a "Cero", por favor verifique e informe al coordinador de laboratorio'
                        inventario_existente.is_active=False

                        # Envío de correo al coordinador del laboratorio
                        alerta='Cantidad reactivo ha llegado a cero'
                        reactivo=f'{inventario_existente.name}'
                        marca=f'{inventario_existente.trademark}'
                        referencia=f'{inventario_existente.reference}'
                        cantidad=f'{inventario_existente.weight}'
                        unidad=f'{inventario_existente.name.unit}'
                        mensaje=f'Con el proceso de edición de salida de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cero en inventario ({cantidad} {unidad}).'
                        enviar_correo_alerta(request, alerta, reactivo, marca, referencia, cantidad, unidad, mensaje, inventario_existente)
                        # Si el inventario es mayor o igual a 0 verificar si está por debajo del control de stock mínimo
                    elif inventario_existente.minStockControl and inventario_existente.weight<=inventario_existente.minstock:
                        warning=', pero esta a hecho que el inventario llegue a una cantidad inferior al "Inventario Mínimo", por favor verifique e informe al coordinador de laboratorio'
                        
                        # Envío de correo al coordinador del laboratorio
                        alerta='Cantidad reactivo por debajo de inventario mínimo'
                        
                        reactivo=f'{inventario_existente.name}'
                        
                        marca=f'{inventario_existente.trademark}'
                        referencia=f'{inventario_existente.reference}'
                        cantidad=f'{inventario_existente.weight}'
                        unidad=f'{inventario_existente.name.unit}'
                        mensaje=f'Con el proceso de edición de la salida de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cantidades críticas ({cantidad} {unidad}), por debajo del stock mínimo ({inventario_existente.minstock} {unidad}).'
                        enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente)
                    # Actualizar el resto de datos del inventario
                    inventario_existente.last_updated_by=request.user
                    inventario_existente.save()  
                    # Crea un evento de editar inventario
                    tipo_evento = 'EDITAR INVENTARIO'
                    usuario_evento = request.user
                    crear_evento(tipo_evento, usuario_evento)              
            else:
                return HttpResponse("No se puede realizar la edición de la salida ya que los valor seleccionados (Laboratorio, Reactivo, Marca o referencia) no corresponden a un insumo existente en el inventario, verifique de nuevo", status=400)

        except Inventarios.DoesNotExist:
            print('Inventario no coincide')

        
        # Actualizar tomar unidades del inventario para el mensaje        
        
        unit=inventario_existente.name.unit
        # Actualizar datos de entrada 
        inventario=get_object_or_404(Inventarios, id=inventario_existente.id)    
        salida.inventario=inventario
        salida.name=name
        salida.trademark=trademark
        salida.reference=reference
        salida.weight=weight
        salida.location=location
        salida.manager=manager
        salida.observations=observations
        salida.destination=destination
        salida.lab=laboratorio
        salida.created_by=request.user  # Asignar el usuario actualmente autenticado
        salida.save()
        # Crea un evento de editar salida
        tipo_evento = 'EDITAR SALIDA'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)
        
        mensaje=f'Se ha editado de manera exitosa la salida del insumo : {salida.name} cantidad: {salida.weight} {unit}{warning}.'
        return HttpResponse(mensaje, status=200)
    #----------------------------------------------------------------------#    
    # Contexto para incluir valores previos en el template
    # Recupera la entrada o muestra una página de error 404 si no se encuentra
    salida = get_object_or_404(Salidas, pk=pk)
    
    weight = int(salida.weight)

    laboratorio = request.user.lab
    
    context = {
        'salida': salida,
        'reactivos': Reactivos.objects.all(),
        'weight':weight,
        'responsables': Responsables.objects.all(),
        'marcas': Marcas.objects.all(),
        'ubicaiones': Ubicaciones.objects.all(),
        'destinos':Destinos.objects.all(),
        'laboratorios':Laboratorios.objects.all(),
        'wubicaciones':Almacenamiento.objects.all(),
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,
        'usuarios': User.objects.all(),        
    }

    return render(request, 'reactivos/editar_salida.html', context)

# Funcionalidad para eliminar (No lo elimina, lo inactiva)registros de reactivo, 

@login_required
def eliminar_usuario(request, pk):
    # Obtén la entrada correspondiente al PK o muestra una página de error 404 si no se encuentra
    usuario = get_object_or_404(User, pk=pk)
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    usuario.is_active=False
    usuario.last_updated_by=request.user
    usuario.save()
    # Crea un evento de eliminar usuario
    tipo_evento = 'ELIMINAR USUARIO'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    
    # Construye el mensaje de éxito
    mensaje = f'Se ha eliminado a petición del usuario el registro número {pk} usuario "{usuario.first_name} {usuario.last_name}" de manera exitosa.'
    mensaje = mensaje
    
    return HttpResponse(mensaje,200)

# Funcionalidad para Activar registros de usuario, 

@login_required
def activar_usuario(request, pk):
    # Obtén la entrada correspondiente al PK o muestra una página de error 404 si no se encuentra
    usuario = get_object_or_404(User, pk=pk)
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    usuario.is_active=True
    usuario.last_updated_by=request.user
    usuario.save()

    # Crea un evento de activar usuario
    tipo_evento = 'ACTIVAR USUARIO'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    
    # Construye el mensaje de éxito
    mensaje = f'Se ha restaurado a petición del usuario el registro número {pk} usuario "{usuario.first_name} {usuario.last_name}" de manera exitosa.'
    mensaje = mensaje
    
    return HttpResponse(mensaje,200)



# Funcionalidad para eliminar (No lo elimina, lo inactiva)registros de reactivo, 

@login_required
def eliminar_reactivo(request, pk):
    # Obtén la entrada correspondiente al PK o muestra una página de error 404 si no se encuentra
    reactivo = get_object_or_404(Reactivos, pk=pk)
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    reactivo.is_active=False
    reactivo.last_updated_by=request.user
    reactivo.save()
    # Crea un evento de eliminar reactivo
    tipo_evento = 'ELIMINAR REACTIVO'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    
    # Construye el mensaje de éxito
    mensaje = f'Se ha eliminado a petición del usuario el registro número {pk} reactivo "{reactivo.name}" de manera exitosa.'
    mensaje = mensaje
    
    return HttpResponse(mensaje,200)

# Funcionalidad para activar registros de reactivo, 

@login_required
def activar_reactivo(request, pk):
    # Obtén la entrada correspondiente al PK o muestra una página de error 404 si no se encuentra
    reactivo = get_object_or_404(Reactivos, pk=pk)
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    reactivo.is_active=True
    reactivo.last_updated_by=request.user
    reactivo.save()

    # Crea un evento de activar reactivo
    tipo_evento = 'ACTIVAR REACTIVO'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    
    # Construye el mensaje de éxito
    mensaje = f'Se ha activado a petición del usuario el registro número {pk} reactivo "{reactivo.name}" de manera exitosa.'
    mensaje = mensaje
    
    return HttpResponse(mensaje,200)


# Funcionalidad para eliminar (No lo elimina, lo inactiva)registros de entrada, además de esto debe restar del inventario la cantidad sumada
# Al correspondiente resgistro, además si llega a cero debe inactivar el registro

@login_required
def eliminar_entrada(request, pk):
    # Obtén la entrada correspondiente al PK o muestra una página de error 404 si no se encuentra
    entrada = get_object_or_404(Entradas, pk=pk)
    # se usará para establecer mensaje adicional si el inventario llega a cero
    warning=''
    
    # Guarda el nombre de la entrada antes de eliminarla para el mensaje
    nombre_entrada = entrada.name

    # # Guarda cantidad y registro de inventario antes de eliminar para restar del inventario
    cantidad_entrada = entrada.weight
    id_inventario = entrada.inventario.id
    inventario = get_object_or_404(Inventarios, id=id_inventario)
    if inventario.weight<cantidad_entrada:
        return HttpResponse('No se puede eliminar el registro ya que esta acción hace que el inventario sea menor que 0',200)
    else:
        inventario.weight=inventario.weight-cantidad_entrada
        inventario.last_updated_by=request.user
        if inventario.weight==0:
            warning= ' Con esta acción el inventario a llegado a 0 por favor verifique y comuniquese con su coordinador.'
            inventario.is_active=False

            # Envío de correo al coordinador del laboratorio
            alerta='Cantidad reactivo ha llegado a cero'
            reactivo=f'{inventario.name}'
            marca=f'{inventario.trademark}'
            referencia=f'{inventario.reference}'
            cantidad=f'{inventario.weight}'
            unidad=f'{inventario.name.unit}'
            inventario_existente=inventario
            mensaje=f'Con el proceso de eliminación del registro de entrada de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cero en inventario ({cantidad} {unidad}).'
            enviar_correo_alerta(request, alerta, reactivo, marca, referencia, cantidad, unidad, mensaje, inventario_existente)
        elif inventario.minStockControl and inventario.weight>0 and inventario.weight<=inventario.minstock:
            warning= ' Con esta acción el inventario a llegado por debajo al inventario mínimo por favor verifique y comuniquese con su coordinador.'    
            # Envío de correo al coordinador del laboratorio
            alerta='Cantidad reactivo por debajo de inventario mínimo'
            reactivo=f'{inventario.name}'
            marca=f'{inventario.trademark}'
            referencia=f'{inventario.reference}'
            cantidad=f'{inventario.weight}'
            unidad=f'{inventario.name.unit}'
            inventario_existente=inventario
            mensaje=f'Con el proceso de eliminación del registro de la entrada de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cantidades críticas ({cantidad} {unidad}), por debajo del stock mínimo ({inventario.minstock} {unidad}).'
            enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente)
    

    # Crea un evento de editar inventario
    
    tipo_evento = 'EDITAR INVENTARIO'
    
    usuario_evento = request.user
    
    crear_evento(tipo_evento, usuario_evento)
    inventario.save()    
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    entrada.is_active=False
    entrada.last_updated_by=request.user
    entrada.save()
    # Crea un evento de eliminar entrada
    tipo_evento = 'ELIMINAR ENTRADA'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    # Construye el mensaje de éxito
    mensaje = f'Se ha eliminado a petición del usuario el registro número {pk} reactivo "{nombre_entrada}" de manera exitosa.'
    mensaje = mensaje+warning
    
    return HttpResponse(mensaje,200)

# Funcionalidad para eliminar (No lo elimina, lo inactiva)registros de entrada, además de esto debe restar del inventario la cantidad sumada
# Al correspondiente resgistro, además si llega a cero debe inactivar el registro

@login_required
def eliminar_salida(request, pk):
    # Obtener la salida correspondiente al PK o muestra una página de error 404 si no se encuentra
    salida = get_object_or_404(Salidas, pk=pk)
    # se usará para establecer mensaje adicional si el inventario llega a cero
    warning=''
    
    # Guarda el nombre de la entrada antes de eliminarla para el mensaje
    nombre_salida = salida.name

    # # Guarda cantidad y registro de inventario antes de eliminar para sumar al inventario
    cantidad_salida = salida.weight
    id_inventario = salida.inventario.id
    inventario = get_object_or_404(Inventarios, id=id_inventario)
    nuevo_posible_inventario=inventario.weight+cantidad_salida
    if nuevo_posible_inventario<0:
        return HttpResponse('No se puede eliminar el registro ya que esta acción hace que el inventario sea menor que 0',200)
    else:
        inventario.weight=nuevo_posible_inventario
        inventario.last_updated_by=request.user
        if inventario.weight==0:
            warning= ' Pero con esta acción el inventario a llegado a 0 por favor verifique y comuniquese con su coordinador.'
            inventario.is_active=False

            # Envío de correo al coordinador del laboratorio
            alerta='Cantidad reactivo ha llegado a cero'
            reactivo=f'{inventario.name}'
            marca=f'{inventario.trademark}'
            referencia=f'{inventario.reference}'
            cantidad=f'{inventario.weight}'
            unidad=f'{inventario.name.unit}'
            inventario_existente=inventario
            mensaje=f'Con el proceso de eliminación del registro de salida de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cero en inventario ({cantidad} {unidad}).'
            enviar_correo_alerta(request, alerta, reactivo, marca, referencia, cantidad, unidad, mensaje, inventario_existente)
        elif inventario.minStockControl and inventario.weight>0 and inventario.weight<=inventario.minstock:
            warning= ' Pero con esta acción el inventario a llegado por debajo al inventario mínimo por favor verifique y comuniquese con su coordinador.'    
            # Envío de correo al coordinador del laboratorio
            alerta='Cantidad reactivo por debajo de inventario mínimo'
            reactivo=f'{inventario.name}'
            marca=f'{inventario.trademark}'
            referencia=f'{inventario.reference}'
            cantidad=f'{inventario.weight}'
            unidad=f'{inventario.name.unit}'
            inventario_existente=inventario
            mensaje=f'Con el proceso de eliminación del registro de la salida de reactivo {reactivo}, con marca {marca} y referencia {referencia}; ha llegado a cantidades críticas ({cantidad} {unidad}), por debajo del stock mínimo ({inventario.minstock} {unidad}).'
            enviar_correo_alerta(request, alerta, reactivo, marca, referencia,cantidad,unidad,mensaje, inventario_existente)
    
    inventario.save()
    # Crea un evento de editar inventario
    tipo_evento = 'EDITAR INVENTARIO'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)    
    # Elimina el registro lo inactiva, no lo elimina y además actualiza el usuario que realiza la acción
    salida.is_active=False
    salida.last_updated_by=request.user
    salida.save()

    # Crea un evento de eliminar salida
    tipo_evento = 'ELIMINAR SALIDA'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)
    
    # Construye el mensaje de éxito
    mensaje = f'Se ha eliminado a petición del usuario el registro número {pk} reactivo "{nombre_salida}" de manera exitosa.'
    mensaje = mensaje+warning
    
    return HttpResponse(mensaje,200)

# La vista "inventario" se encarga de obtener los valores de la tabla del modelo "Inventario" y los envía al template "inventario.html"
# Además, recibe los valores filtrados desde el template, especificando qué elementos se desean mostrar en la lista. Estos valores 
# filtrados, como el nombre y la marca, se guardan en la sesión utilizando los siguientes comandos:
# request.session['filtered_name'] = name
# request.session['filtered_trademark'] = trademark
# Esto permite que los valores de filtrado puedan ser utilizados en otras vistas que realicen la exportación a formatos como Excel o 
# PDF. De esta manera, se garantiza la consistencia de los datos filtrados al exportarlos a otros formatos.

class InventarioListView(LoginRequiredMixin,ListView):
    model = Inventarios
    template_name = "reactivos/inventario.html"
    paginate_by = 10
    

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        lab = request.GET.get('lab')
        name = request.GET.get('name')
        id_r = request.GET.get('id_r')
        
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
            lab=''

        # Guardar los valores de filtrado en la sesión
        request.session['filtered_lab'] = lab
        request.session['filtered_name'] = name
        request.session['filtered_id_r'] = id_r
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Entradas para colocar un contexto de acuerdo con el filtro
        lab = self.request.GET.get('lab')
        name = self.request.GET.get('id_r')
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=''
        
        # Colocar contexto de acuerdo con el filtro seleccionado
        if name and lab:
            reactivo=get_object_or_404(Reactivos, id=name)
            lab=get_object_or_404(Laboratorios, id=lab)
            # Busca en el inventario por Reactivo y Marca y obtén la cantidad total
            inventario_entries = Inventarios.objects.filter(name=reactivo, lab=lab, is_active=True)

            # Calcula la cantidad total en inventario
            total_weight = inventario_entries.aggregate(total_weight=models.Sum('weight'))['total_weight'] or 0

            context['reactivo_filtrado']=reactivo.name
            context['cantidad_filtrada']=total_weight
            context['unidad_filtrada']=reactivo.unit.name
        
        
        elif name:
            reactivo=get_object_or_404(Reactivos, id=name)
            
            # Busca en el inventario por Reactivo y Marca y obtén la cantidad total
            inventario_entries = Inventarios.objects.filter(name=reactivo, is_active=True)

            # Calcula la cantidad total en inventario
            total_weight = inventario_entries.aggregate(total_weight=models.Sum('weight'))['total_weight'] or 0

            context['reactivo_filtrado']=reactivo.name
            context['cantidad_filtrada']=total_weight
            context['unidad_filtrada']=reactivo.unit.name
            
        
        else:
            context['reactivo_filtrado']=''
            context['cantidad_filtrada']=''
            context['unidad_filtrada']=''


        unique_labs_ids = Inventarios.objects.values('lab').distinct()
        unique_labs = Laboratorios.objects.filter(id__in=unique_labs_ids)

        unique_names_ids = Inventarios.objects.values('name').distinct()
        unique_names = Reactivos.objects.filter(id__in=unique_names_ids)

        unique_trademarks_ids = Inventarios.objects.values(
            'trademark').distinct()
        unique_trademarks = Marcas.objects.filter(id__in=unique_trademarks_ids)
               
        laboratorio = self.request.user.lab
        
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        

        context['unique_labs'] = unique_labs
        context['unique_names'] = unique_names
        context['unique_trademarks'] = unique_trademarks
        

        

        # Obtener la lista de inventarios
        inventarios = context['object_list']

        # Recorrer los inventarios y cambiar el formato de la fecha
        for inventario in inventarios:
            # Obtener la fecha de vencimiento
            edate = inventario.edate
            if edate:
                # Cambiar el formato de fecha de inglés a formato dd/mm/aaaa
                edate = edate.strftime('%d/%m/%Y')

            # Actualizar la fecha en el inventario
            inventario.edate = edate

        context['object_list'] = inventarios

        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        lab = self.request.GET.get('lab')
        id_reactivo = self.request.GET.get('id_r')
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=''
             
        # Definir Queryset para filtrado de visulización
             
        if lab and id_reactivo:
            queryset = queryset.filter(name=id_reactivo, lab=lab, is_active=True)
        elif lab:
            queryset = queryset.filter(lab=lab, is_active=True)


        elif id_reactivo:
            queryset = queryset.filter(name=id_reactivo, is_active=True)

        
        
        else:
            queryset = queryset.filter(is_active=True)
            
        queryset = queryset.order_by('id')
        return queryset
    


#Listado Usuarios
class ListadoUsuarios(LoginRequiredMixin, ListView):
    model=User
    template_name='usuarios/listar_usuarios.html'

    def get_queryset(self):
        return self.model.objects.filter(is_active=True)
    


    
# Muestra el listado de salidas
class EntradasListView(LoginRequiredMixin,ListView):
    model = Entradas
    template_name = "reactivos/listado_entradas.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        lab = request.GET.get('lab')
        name = request.GET.get('name')
        id_reagent = request.GET.get('id_r')
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')  
        
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
            lab=''

        # Guardar los valores de filtrado en la sesión
        request.session['filtered_lab'] = lab
        request.session['filtered_name'] = name
        request.session['filtered_id_reagent'] = id_reagent
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Calcular la fecha hace un mes hacia atrás
        one_month_ago = today - timedelta(days=30)

        # Agregar la fecha al contexto
        context['one_month_ago'] = one_month_ago

        # Agregar la fecha de hoy al contexto
        context['today'] = today

        unique_labs_ids = Entradas.objects.values('lab').distinct()
        unique_labs = Laboratorios.objects.filter(id__in=unique_labs_ids)

        unique_names_ids = Entradas.objects.values('name').distinct()
        unique_names = Reactivos.objects.filter(id__in=unique_names_ids)

        unique_locations_ids = Entradas.objects.values(
            'location').distinct()
        
        unique_locations = Ubicaciones.objects.filter(id__in=unique_locations_ids)
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        context['shools'] = Facultades.objects.all()
        context['destinations'] = Destinos.objects.all()
        context['created_bys'] = UserModel.objects.all()
        

        context['unique_labs'] = unique_labs
        context['unique_names'] = unique_names
        context['unique_locations'] = unique_locations
        

        

        # Obtener la lista de inventarios
        entradas = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha
        
        for entrada in entradas:
            # Obtener la fecha de vencimiento
            date_order = entrada.date_order
            if date_order==None:
                date_order=''
            else:
                # Cambiar el formato de fecha de inglés a formato dd/mm/aaaa
                date_order = date_order.strftime('%d/%m/%Y')

            # Actualizar la fecha en el inventario
            entrada.date_order = date_order
        
        context['object_list'] = entradas
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        lab = self.request.GET.get('lab')
        id_reagent = self.request.GET.get('id_r')
        
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

                
        

        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=None
        
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(date_create__gte=start_date)
        if end_date:
            queryset = queryset.filter(date_create__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)

             
        
        if lab and id_reagent:
            queryset = queryset.filter(name=id_reagent, lab=lab, is_active=True)
        elif lab:
            queryset = queryset.filter(lab=lab, is_active=True)
        elif id_reagent:
            queryset = queryset.filter(name=id_reagent, is_active=True)
        else:
            queryset = queryset.filter(is_active=True)

            
        queryset = queryset.order_by('id')
        return queryset
    
# Muestra el listado de salidas
class SalidasListView(LoginRequiredMixin,ListView):
    model = Salidas
    template_name = "reactivos/listado_salidas.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        lab = request.GET.get('lab')
        name = request.GET.get('name')
        id_reagent = request.GET.get('id_r')
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')  
        
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
            lab=''

        # Guardar los valores de filtrado en la sesión
        request.session['filtered_lab'] = lab
        request.session['filtered_name'] = name
        request.session['filtered_id_reagent'] = id_reagent
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Calcular la fecha hace un mes hacia atrás
        one_month_ago = today - timedelta(days=30)

        # Agregar la fecha al contexto
        context['one_month_ago'] = one_month_ago

        # Agregar la fecha de hoy al contexto
        context['today'] = today

        unique_labs_ids = Entradas.objects.values('lab').distinct()
        unique_labs = Laboratorios.objects.filter(id__in=unique_labs_ids)

        unique_names_ids = Entradas.objects.values('name').distinct()
        unique_names = Reactivos.objects.filter(id__in=unique_names_ids)

        unique_locations_ids = Entradas.objects.values(
            'location').distinct()
        
        unique_locations = Ubicaciones.objects.filter(id__in=unique_locations_ids)
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        context['shools'] = Facultades.objects.all()
        context['destinations'] = Destinos.objects.all()
        context['created_bys'] = UserModel.objects.all()
        

        context['unique_labs'] = unique_labs
        context['unique_names'] = unique_names
        context['unique_locations'] = unique_locations
        

        

        # Obtener la lista de inventarios
        entradas = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha
        
               
        context['object_list'] = entradas
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        lab = self.request.GET.get('lab')
        id_reagent = self.request.GET.get('id_r')
        
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

                
        

        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=None
        
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(date_create__gte=start_date)
        if end_date:
            queryset = queryset.filter(date_create__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)

             
        
        if lab and id_reagent:
            queryset = queryset.filter(name=id_reagent, lab=lab, is_active=True)
        elif lab:
            queryset = queryset.filter(lab=lab, is_active=True)
        elif id_reagent:
            queryset = queryset.filter(name=id_reagent, is_active=True)
        else:
            queryset = queryset.filter(is_active=True)

            
        queryset = queryset.order_by('id')
        return queryset
    
# Muestra el listado de usuarios
class UsuariosListView(LoginRequiredMixin,ListView):
    model = User
    template_name = "usuarios/listado_usuarios.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        lab = request.GET.get('lab')
        
        id_user = request.GET.get('id_user')
        
        
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
            lab=''

        # Guardar los valores de filtrado en la sesión
        request.session['filtered_lab'] = lab
        
        request.session['filtered_id'] = id_user      
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        
        
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['roles'] = Rol.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        

        # Obtener la lista de inventarios
        usuarios = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha
        
               
        context['object_list'] = usuarios
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        lab = self.request.GET.get('lab')
        
        user_id = self.request.GET.get('id_user')
        
        # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
        if lab=='0':
             lab=None

                
        if lab:
            queryset = queryset.filter(lab=lab)
        if user_id:
            queryset = queryset.filter(id=user_id)
                    
        queryset = queryset.order_by('id')
        return queryset
# Listado de eventos

class EventosListView(LoginRequiredMixin,ListView):
    model = Eventos
    template_name = "admin/listado_eventos.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        id_user = request.GET.get('id_user')
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        
        

        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_id'] = id_user
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
           
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Agregar la fecha de hoy al contexto
        context['today'] = today
        
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['roles'] = Rol.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        

        # Obtener la lista de inventarios
        usuarios = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha
        
               
        context['object_list'] = usuarios
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        
        user_id = self.request.GET.get('id_user')
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
             
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass
        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(fecha_evento__gte=start_date)
        if end_date:
            queryset = queryset.filter(fecha_evento__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(fecha_evento__gte=start_date,fecha_evento__lte=end_date)
        
        if user_id:
            queryset = queryset.filter(usuario_evento=user_id)
                    
        queryset = queryset.order_by('-id')
        return queryset

# Muestra el listado de reactivos
class ReactivosListView(LoginRequiredMixin,ListView):
    model = Reactivos
    template_name = "reactivos/listado_reactivos.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR', 'TECNICO'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        name = self.request.GET.get('name')
        
        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_name'] = name
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        laboratorio = self.request.user.lab
        
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        
        # Obtener la lista de inventarios
        reactivos = context['object_list']
        if hasattr(self, 'no_results_message'):
            context['no_results_message'] = self.no_results_message
               
        context['object_list'] = reactivos
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        name = self.request.GET.get('name')
        if name:
            queryset = queryset.filter(name=name)
            
        queryset = queryset.order_by('id')
        # Verificar si no se encontraron resultados
        if not queryset.exists():
            self.no_results_message = 'No hay reactivos que coincidan con el criterio de búsqueda: "'+name+'"'
        return queryset


#Listado Usuarios
class ListadoUsuarios(LoginRequiredMixin, ListView):
    model=User
    template_name='usuarios/listar_usuarios.html'

    def get_queryset(self):
        return self.model.objects.filter(is_active=True)


#Crear Usuarios
class CrearUsuario(LoginRequiredMixin, CreateView):
    model=User
    form_class=FormularioUsuario
    template_name='usuarios/crear_usuario.html'
    success_url='reactivos:index'
    # Validador de grupos que pueden acceder
    # Sobreescribir el método dispatch para aplicar el decorador
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['laboratorio'] = self.request.user.lab
        context['labname'] = self.request.user.lab.name
        context['usuarios'] = User.objects.all()
        context['roles'] = Rol.objects.all()  
        context['laboratorios'] = Laboratorios.objects.all()  
        return context
    
    def send_confirmation_email(self, user):
        protocol = 'https' if self.request.is_secure() else 'http'
        domain = self.request.get_host()
        # Codificar el correo electrónico en base64 y agregarlo en la URL
        encoded_email = base64.urlsafe_b64encode(user.email.encode()).decode()
        reset_link = reverse('reactivos:password_reset') + '?' + urlencode({'email': encoded_email})
        reset_url = f"{protocol}://{domain}{reset_link}"
        home_url = f"{protocol}://{domain}"

        subject = _('Bienvenido a la Gestión de Insumos Químicos')
        if user.acceptDataProcessing:
            aceptapolitica='Acepta'
        else:
            aceptapolitica='No acepta'

        context = {
            'user': user,
            'reset_url': reset_url,
            'aceptapolitica':aceptapolitica,
            'home_url':home_url,
        }
        message = render_to_string('registration/registro_exitoso_email.html', context)
        plain_message = strip_tags(message)
        from_email = None  # Agrega el correo electrónico desde el cual se enviará el mensaje
        recipient_list = [user.email]
        # envío de correo electrónico en segundo plano
        # Archivo adjunto nulo
        attach_path=None
        # Crear un hilo y ejecutar enviar_correo en segundo plano
        correo_thread = threading.Thread(
        target=enviar_correo, args=(recipient_list, subject, message, attach_path),)
        correo_thread.start()

    def form_valid(self, form):

        def has_required_password_conditions(password):
            # Mínimo 8 caracteres, al menos una mayúscula, un número y un carácter especial
            password_pattern = r"^(?=.*[A-Z])(?=.*\d)(?=.*[@$!%*?&])[A-Za-z\d@$!%*?&]{8,}$"
            return re.match(password_pattern, password) is not None
        
        # Verificar si el id_number ya existe en la base de datos
        id_number = form.cleaned_data['id_number']

        if User.objects.filter(id_number=id_number).exists():
            messages.error(self.request, f"No es posible crear el usuario {form.cleaned_data['username']} ya que su número de identificación {id_number} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")
        
        # Verificar si el phone_number ya existe en la base de datos
        phone_number = form.cleaned_data['phone_number']

        if User.objects.filter(phone_number=phone_number).exists():
            messages.error(self.request, f"No es posible crear el usuario {form.cleaned_data['username']} ya que su número de teléfono {phone_number} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")
        
        # Verificar si el email ya existe en la base de datos
        email = form.cleaned_data['email']
        

        if User.objects.filter(email=email).exists():
            messages.error(self.request, f"No es posible crear el usuario {form.cleaned_data['username']} ya que su correo electrónico {email} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")
        
        # Verificar si el username ya existe en la base de datos
        username = form.cleaned_data['username']
        

        if User.objects.filter(username=username).exists():
            messages.error(self.request, f"No es posible crear el usuario {form.cleaned_data['username']} ya que su nombre de usuario {username} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")

        # Validación personalizada para contraseñas
        password1 = form.cleaned_data.get('password1')
        password2 = form.cleaned_data.get('password2')
        
        if password1 != password2:
            messages.error(self.request, "No se puedo crear el usuario porque las contraseñas no coinciden.")
            return HttpResponse('Contraseña no cumple', status=400)

        if not has_required_password_conditions(password1):
            messages.error(self.request, "No se puedo crear el usuario porque la contraseña no satisface los requisitos de la política de contraseñas: Mínimo 8 caracteres, al menos una letra mayúscula, un número y un caracter especial")
            
            return HttpResponse('Contraseña no cumple', status=400)
        
        
        # Agrega los campos adicionales al usuario antes de guardarlo en la base de datos

        user = form.save(commit=False)
        user.acceptDataProcessing = form.cleaned_data['acceptDataProcessing']
        user.rol = form.cleaned_data['rol']
        user.lab = form.cleaned_data['lab']
        user.id_number = form.cleaned_data['id_number']
        user.phone_number = form.cleaned_data['phone_number']
        user.email = form.cleaned_data['email']
        user.username = form.cleaned_data['username']
        # Asignar el usuario actual al campo user_create
        user.user_create = self.request.user
        # Asignar el usuario actual al campo last_updated_by
        user.last_updated_by = self.request.user
        
        # Guarda el usuario en la base de datos
        user.save()
        
        # Establece el objeto creado para que se pueda usar en la redirección
        self.object = user
        
     
        # Enviar correo electrónico de confirmación
        protocol = 'https' if self.request.is_secure() else 'http'
        domain = self.request.get_host()
        # Codificar el correo electrónico en base64 y agregarlo en la URL
        encoded_email = base64.urlsafe_b64encode(user.email.encode()).decode()
        reset_link = reverse('reactivos:password_reset') + '?' + urlencode({'email': encoded_email})
        reset_url = f"{protocol}://{domain}{reset_link}"
        home_url = f"{protocol}://{domain}"

        subject = _('Bienvenido a la Gestión de Insumos Químicos')
        if user.acceptDataProcessing:
            aceptapolitica='Acepta'
        else:
            aceptapolitica='No acepta'

        context = {
            'user': user,
            'reset_url': reset_url,
            'aceptapolitica':aceptapolitica,
            'home_url':home_url,
        }
        message = render_to_string('registration/registro_exitoso_email.html', context)
        plain_message = strip_tags(message)
        from_email = None  # Agrega el correo electrónico desde el cual se enviará el mensaje
        recipient_list = [user.email]
        # envío de correo electrónico en segundo plano
        # Archivo adjunto nulo
        attach_path=None
        # Crear un hilo y ejecutar enviar_correo en segundo plano
        correo_thread = threading.Thread(
        target=enviar_correo, args=(recipient_list, subject, message, attach_path),)
        correo_thread.start()

        # Agregar mensaje de éxito
        # Crea un evento de crear usuario
        tipo_evento = 'CREAR USUARIO'
        usuario_evento = self.request.user
        crear_evento(tipo_evento, usuario_evento)
        messages.success(self.request, f"Se ha creado exitosamente el usuario {user.first_name} {user.last_name} y se ha enviado un correo electrónico de confirmación a {user.email}.")

        return HttpResponse('Operación exitosa', status=200)
    
# Vista para editar usuarios
@method_decorator(login_required, name='dispatch')
class EditarUsuario(UpdateView):
    model = User
    form_class = FormularioUsuario
    template_name = 'usuarios/editar_usuario.html'
    success_url='reactivos:editar_usuario'
    
    

    # Validador de grupos que pueden acceder
    # Sobreescribir el método dispatch para aplicar el decorador
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        
        
        # Verificar si el id_number ya existe en la base de datos
        id_number = form.cleaned_data['id_number']
        user = self.get_object()
        

        if User.objects.filter(id_number=id_number).exclude(pk=user.pk).exists():
            messages.error(self.request, f"No es posible editar el usuario {user.username} ya que su número de identificación {id_number} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")

        # Verificar si el phone_number ya existe en la base de datos
        phone_number = form.cleaned_data['phone_number']

        if User.objects.filter(phone_number=phone_number).exclude(pk=user.pk).exists():
            messages.error(self.request, f"No es posible editar el usuario {user.username} ya que su número de teléfono {phone_number} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")

        # Verificar si el email ya existe en la base de datos
        email = form.cleaned_data['email']

        if User.objects.filter(email=email).exclude(pk=user.pk).exists():
            messages.error(self.request, f"No es posible editar el usuario {user.username} ya que su correo electrónico {email} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")

        # Verificar si el username ya existe en la base de datos
        username = form.cleaned_data['username']

        if User.objects.filter(username=username).exclude(pk=user.pk).exists():
            messages.error(self.request, f"No es posible editar el usuario {user.username} ya que su nombre de usuario {username} ya existe en la base de datos.")
            return HttpResponseBadRequest("Ya existe un registro en la base de datos")

        # Aplicar los cambios al usuario y guardarlo
        user = form.save()
        print(f'{user.first_name}')
        # Agregar mensaje de éxito
        messages.success(self.request, f"Se ha editado exitosamente el usuario {user.first_name} {user.last_name}.")

        return HttpResponse('Operación exitosa', status=200)

    
    
    def get_object(self, queryset=None):
        # Obtener el valor de pk desde la URL
        pk = self.kwargs.get('pk')

        # Obtener el usuario que se está editando o devolver 404 si no existe
        return get_object_or_404(User, pk=pk)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        #  Aquí deberías decodificar la contraseña y pasarla como valor inicial al formulario
        user = self.get_object()
        context['user'] = user
        context['laboratorio'] = self.request.user.lab
        context['labname'] = self.request.user.lab.name
        context['usuarios'] = User.objects.all()
        context['roles'] = Rol.objects.all()  
        context['laboratorios'] = Laboratorios.objects.all()  
        return context
    
# La vista "editar_reactivo" se encarga de gestionar la edición  de un registro reactivo en la db. Esta vista toma los datos del formulario 
# existente en el template "editar_reactivo.html" y realiza las operaciones necesarias en la base de datos para editar 
# la información del reactivo. Esto puede incluir la validación de los datos ingresados, la edición de un nuevo registro 
# en la tabla correspondiente y cualquier otra gestión requerida para asegurar la integridad de los datos en la base de datos.
@login_required
def editar_usuario(request, pk):
    usuario = get_object_or_404(User, pk=pk)
    if request.method == 'POST':
        id_number = request.POST.get('id_number')
        phone_number = request.POST.get('phone_number')
        email = request.POST.get('email')
        username = request.POST.get('username')
        first_name = request.POST.get('first_name')
        first_name=estandarizar_nombre(first_name)
        last_name = request.POST.get('last_name')
        last_name=estandarizar_nombre(last_name)
        rol = request.POST.get('rol')
        lab = request.POST.get('lab')

        # Verificación de unicidad de número de identificación
        if User.objects.filter(id_number=id_number).exclude(pk=usuario.pk).exists():
            mensaje=f"No es posible editar el usuario {usuario.username} ya que su número de identificación {id_number} ya existe en la base de datos."
            return HttpResponseBadRequest(mensaje)
        
        # Verificación de unicidad de número telefónico
        if User.objects.filter(phone_number=phone_number).exclude(pk=usuario.pk).exists():
            mensaje =f"No es posible editar el usuario {usuario.username} ya que su número de teléfono {phone_number} ya existe en la base de datos."
            return HttpResponseBadRequest(mensaje)
        
        # Verificación de unicidad de correo electrónico
        if User.objects.filter(email=email).exclude(pk=usuario.pk).exists():
            mensaje =f"No es posible editar el usuario {usuario.username} ya que su correo electrónio {email} ya existe en la base de datos."
            return HttpResponseBadRequest(mensaje)
        
        # Verificación de unicidad de nombre de usuario
        if User.objects.filter(username=username).exclude(pk=usuario.pk).exists():
            mensaje =f"No es posible editar el usuario {usuario.username} ya que su nombre de usuario {username} ya existe en la base de datos."
            return HttpResponseBadRequest(mensaje)
        
        # Instancia de rol
        rol = get_object_or_404(Rol, id=rol)

        # Instancia de lab
        lab = get_object_or_404(Laboratorios, id=lab)
        
        usuario.id_number=id_number
        usuario.phone_number=phone_number
        usuario.email=email
        usuario.username=username
        usuario.first_name=first_name
        usuario.last_name=last_name
        usuario.rol=rol
        usuario.lab=lab
        usuario.last_updated_by=request.user
        usuario.save()

        # Crea un evento de editar usuario
        tipo_evento = 'EDITAR USUARIO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        mensaje = f"Se ha editado exitosamente el usuario {usuario.first_name} {usuario.last_name}."
        return HttpResponse(mensaje,200)    
    ############# Contexto ##########
    laboratorio = request.user.lab
    
    context = {
        'usuario':usuario,
        'laboratorio':laboratorio,
        'laboratorios': Laboratorios.objects.all(),
        'roles': Rol.objects.all(),
    }
    
    return render(request, 'usuarios/editar_usuario.html', context)


# Muestra el listado de solicitudes
class SolicitudesListView(LoginRequiredMixin,ListView):
    model = Solicitudes
    template_name = "solicitudes/listado_solicitudes.html"
    paginate_by = 10
    
    
    @check_group_permission(groups_required=['ADMINISTRADOR'])
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    def get(self, request, *args, **kwargs):
        # Obtener el número de registros por página de la sesión del usuario
        per_page = request.session.get('per_page')
        if per_page:
            self.paginate_by = int(per_page)
        else:
            self.paginate_by = 10  # Valor predeterminado si no hay variable de sesión

        # Obtener los parámetros de filtrado
        
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')  
        
        

        # Guardar los valores de filtrado en la sesión
        
        request.session['filtered_start_date'] = start_date
        request.session['filtered_end_date'] = end_date
        

        return super().get(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)

        # Obtener la fecha de hoy
        today = date.today()
        # Calcular la fecha hace un mes hacia atrás
        one_month_ago = today - timedelta(days=30)

        # Agregar la fecha al contexto
        context['one_month_ago'] = one_month_ago

        # Agregar la fecha de hoy al contexto
        context['today'] = today
        laboratorio = self.request.user.lab
        
    
        context['usuarios'] = User.objects.all()
        context['laboratorio'] = laboratorio
        context['laboratorios'] = Laboratorios.objects.all()
        
        # Obtener la lista de inventarios
        entradas = context['object_list']
        # Recorrer los entradas y cambiar el formato de la fecha        
               
        context['object_list'] = entradas
        return context
    
    def get_queryset(self):
        queryset = super().get_queryset()
        # Obtener las fechas de inicio y fin de la solicitud GET
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')
        
        # Validar y convertir las fechas
        try:
            if start_date:
                start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
            if end_date:
                end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
        except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
            pass

        
        # Realiza la filtración de acuerdo a las fechas
        if start_date:
            queryset = queryset.filter(date_create__gte=start_date)
        if end_date:
            queryset = queryset.filter(date_create__lte=end_date)
        elif start_date and end_date:
            queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)
        else:
            queryset = queryset.filter(is_active=True)
            
        queryset = queryset.order_by('id')
        return queryset



#-------------------------------------------------------------------------------------------------------------------------------------
# Vista para la creación del detalle del reactivo, hasta el momento solo tiene contexto el reactivo, pero se le puede poner lo necesario
@login_required
def detalle_reactivo(request, pk):

    inventario = get_object_or_404(Inventarios, pk=pk)
    laboratorio = request.user.lab
    
    context = {
        'usuarios': User.objects.all(),
        'laboratorio': laboratorio,

        'inventario': inventario
    }
    return render(request, 'reactivos/detalle_reactivo.html', context)


# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda
class GuardarPerPageView(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_lab = request.session.get('filtered_lab')
        filtered_name = request.session.get('filtered_name')
        filtered_id_r = request.session.get('filtered_id_r')
        
        url = reverse('reactivos:inventario')
        params = {}
        if filtered_lab:
            params['lab'] = filtered_lab
        if filtered_name:
            params['name'] = filtered_name
        if filtered_id_r:
            params['id_r'] = filtered_id_r
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda    
class GuardarPerPageViewIn(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_lab = request.session.get('filtered_lab')
        filtered_name = request.session.get('filtered_name')
        filtered_id_reagent = request.session.get('filtered_id_reagent')
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        
        url = reverse('reactivos:listado_entradas')
        params = {}
        if filtered_lab:
            params['lab'] = filtered_lab
        if filtered_name:
            params['name'] = filtered_name

        if filtered_id_reagent:
            params['id_r'] = filtered_id_reagent
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    
# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda    
class GuardarPerPageViewEvent(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        filtered_id = request.session.get('filtered_id')
        
        
        
        url = reverse('reactivos:listado_eventos')
        params = {}
        if filtered_id:
            params['id_user'] = filtered_id
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)

# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda    
class GuardarPerPageViewOut(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_lab = request.session.get('filtered_lab')
        filtered_name = request.session.get('filtered_name')
        filtered_id_reagent = request.session.get('filtered_id_reagent')
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        
        url = reverse('reactivos:listado_salidas')
        params = {}
        if filtered_lab:
            params['lab'] = filtered_lab
        if filtered_name:
            params['name'] = filtered_name

        if filtered_id_reagent:
            params['id_r'] = filtered_id_reagent
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)


# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda

class GuardarPerPageViewSolicitud(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        
        url = reverse('reactivos:listado_solicitudes')
        params = {}
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    
# Paginación de solicitudes externas
class GuardarPerPageViewSolicitudExterna(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_start_date = request.session.get('filtered_start_date')
        filtered_end_date = request.session.get('filtered_end_date')
        filtered_lab = request.session.get('filtered_lab')
        filtered_keyword = request.session.get('filtered_keyword')
        
        url = reverse('reactivos:listado_solicitudes_externas')
        params = {}
        
        if filtered_start_date:
            params['start_date'] = filtered_start_date
        if filtered_end_date:
            params['end_date'] = filtered_end_date
        if filtered_lab:
            params['lab'] = filtered_lab
        if filtered_keyword:
            params['keyword'] = filtered_keyword
        
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    
# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda

class GuardarPerPageViewReactivo(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_name = request.session.get('filtered_name')
        
        url = reverse('reactivos:listado_reactivos')
        params = {}
        if filtered_name:
            params['name'] = filtered_name
        
        if params:
            url += '?' + urlencode(params)

        return redirect(url)
# Guarda los valores de búsqueda del listado para cuando se efectue un cambio en la paginación se conserven 
# estos valores y se sostenga los criterios de búsqueda    
class GuardarPerPageViewUser(LoginRequiredMixin,View):
    def get(self, request, *args, **kwargs):
        per_page = kwargs.get('per_page')
        request.session['per_page'] = per_page

        # Redirigir a la página de inventario con los parámetros de filtrado actuales
        filtered_lab = request.session.get('filtered_lab')
        
        filtered_id = request.session.get('filtered_id')
        
        url = reverse('reactivos:listado_usuarios')
        params = {}
        if filtered_lab:
            params['lab'] = filtered_lab
        
        if filtered_id:
            params['id_user'] = filtered_id

        if params:
            url += '?' + urlencode(params)

        return redirect(url)
    


    
# La vista "crear_unidades" se encarga de gestionar la creación de unidades. Esta vista toma los datos del formulario 
# existente en el template "crear_unidades.html" y realiza las operaciones necesarias en la base de datos utilizando 
# el modelo "Unidades". El objetivo es garantizar la unicidad de los registros, lo que implica verificar si la unidad
# ya existe en la base de datos antes de crearla. Si la unidad es única, se crea un nuevo registro en la tabla 
# correspondiente utilizando el modelo "Unidades". Si la unidad ya existe, se muestra un mensaje de error o se toma la 
# acción apropiada según los requisitos del sistema.

class CrearRoles(LoginRequiredMixin, View):
    template_name = 'reactivos/crear_roles.html'

    @check_group_permission(groups_required=['SUPERUSUARIO'])
    def get(self, request, *args, **kwargs):
        
        context = {
            'usuarios': User.objects.all(),
            'laboratorio': self.request.user.lab,
            
        }
        return render(request, self.template_name, context)

    @check_group_permission(groups_required=['SUPERUSUARIO'])
    def post(self, request, *args, **kwargs):
        name = request.POST.get('name')
        name=estandarizar_nombre(name)

        if Rol.objects.filter(name=name).exists():
            rol = Rol.objects.get(name=name)
            rol_id = rol.id
            messages.error(
                request, 'Ya existe el rol con nombre ' + name + ' id: ' + str(rol_id))
            return HttpResponse('Error al insertar en la base de datos', status=400)

        rol = Rol.objects.create(
            name=name,
            user_create=request.user,
            last_updated_by=request.user,
        )
        rol_id = rol.id
        # Crea un evento de crear rol
        tipo_evento = 'CREAR ROL'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        messages.success(
            request, 'Se ha creado exitosamente el rol con nombre ' + name + ' id: ' + str(rol_id))

        context = {'rol_id': rol.id, 'rol_name': rol.name}
        return HttpResponse('Operación exitosa', status=200)
    

# Devuelve valores de name, trademark y reference para ser insertados los select correspondientes en el template Inventarios al modificar 
# el select name

class NamesTrademarksAndReferencesByLabAPI(LoginRequiredMixin,View):
    def get(self, request):
        lab = request.GET.get('lab')

        inventarios = Inventarios.objects.all()

        if lab:
            if lab!='0':
                inventarios = inventarios.filter(lab=lab)

        names = inventarios.values('name', 'name__name').distinct().order_by('name__name')
        trademarks = inventarios.values('trademark', 'trademark__name').distinct().order_by('trademark__name')
        references = inventarios.values('reference').distinct().order_by('reference')

        names_trademarks_and_references_list = {
            'names': list(names),
            'trademarks': list(trademarks),
            'references': list(references)
        }
        

        return JsonResponse(names_trademarks_and_references_list, safe=False)
    

# Devuelve valores de name, trademark y reference para ser insertados los select correspondientes en el template Inventarios al modificar 
# el select name

class SelectOptionsByLabIN(LoginRequiredMixin,View):
    def get(self, request):
        lab = request.GET.get('lab')

        entradas = Entradas.objects.all()

        if lab:
            if lab!='0':
                entradas = entradas.filter(lab=lab)

        names = entradas.values('name', 'name__name').distinct().order_by('name__name')
        locations = entradas.values('location', 'location__name','location__facultad__name').distinct().order_by('location__name')
        destinations = entradas.values('destination','destination__name').distinct().order_by('destination__name')
        created_bys = entradas.values('created_by','created_by__first_name','created_by__last_name').distinct().order_by('created_by__first_name')

        select_option_list = {
            'names': list(names),
            'locations': list(locations),
            'destinations': list(destinations),
            'created_bys': list(created_bys)
        }

        return JsonResponse(select_option_list, safe=False)

# Devuelve valores de name, trademark y reference para ser insertados los select correspondientes en el template Inventarios al modificar 
# el select name

class SelectOptionsByLabOUT(LoginRequiredMixin,View):
    def get(self, request):
        lab = request.GET.get('lab')

        salidas = Salidas.objects.all()

        if lab:
            if lab!='0':
                salidas = salidas.filter(lab=lab)

        names = salidas.values('name', 'name__name').distinct().order_by('name__name')
        locations = salidas.values('location', 'location__name','location__facultad__name').distinct().order_by('location__name')
        destinations = salidas.values('destination','destination__name').distinct().order_by('destination__name')
        created_bys = salidas.values('created_by','created_by__first_name','created_by__last_name').distinct().order_by('created_by__first_name')

        select_option_list = {
            'names': list(names),
            'locations': list(locations),
            'destinations': list(destinations),
            'created_bys': list(created_bys)
        }

        return JsonResponse(select_option_list, safe=False)

    
# Devuelve valores de trademark y reference para ser insertados los select correspondientes en el template Inventarios al modificar 
# name de reactivo

class TrademarksByLabAndNameAPI(LoginRequiredMixin, View):
    def get(self, request):
        name = request.GET.get('name')
        lab = request.GET.get('lab')
        

        inventarios = Inventarios.objects.all()

        if name:
            # Verificar si el valor de name es un número
            if name.isdigit():
                reactivo = get_object_or_404(Reactivos, id=int(name))
                name = reactivo.id
                inventarios = inventarios.filter(name=name)
            else:
                reactivo = get_object_or_404(Reactivos, name=name)
                name = reactivo.id
                inventarios = inventarios.filter(name=name)

        if lab:
            # Verificar si el valor de lab es un número
            if lab.isdigit():
                if lab == '0':
                    
                    inventarios = Inventarios.objects.all()
                else:
                    laboratorio = get_object_or_404(Laboratorios, id=int(lab))
                    lab = laboratorio.id
                    inventarios = inventarios.filter(lab=lab)
            else:
                lab_obj = get_object_or_404(Laboratorios, name=lab)
                lab = lab_obj.id
                inventarios = inventarios.filter(lab=lab)
                
         
        trademarks = inventarios.values('trademark', 'trademark__name').distinct()
        trademarks_list = list(trademarks)
        return JsonResponse(trademarks_list, safe=False)
    
    

class ReferencesByLabAndNameAPI(LoginRequiredMixin, View):
    def get(self, request):
        name = request.GET.get('name')
        lab = request.GET.get('lab')

        inventarios = Inventarios.objects.all()

        if name:
            # Verificar si el valor de name es un número
            if name.isdigit():
                reactivo = get_object_or_404(Reactivos, id=int(name))
                name = reactivo.id
            else:
                reactivo = get_object_or_404(Reactivos, name=name)
                name = reactivo.id

        if lab:
            # Verificar si el valor de lab es un número
            if lab.isdigit():
                laboratorio = get_object_or_404(Laboratorios, id=int(lab))
                lab = laboratorio.id
            else:
                lab = get_object_or_404(Laboratorios, name=lab)
                lab = lab.id

        if name:
            inventarios = inventarios.filter(name=name)

        if lab:
            inventarios = inventarios.filter(lab=lab)

        references = inventarios.values('reference').distinct()
        references_list = list(references)
        

        return JsonResponse(references_list, safe=False)




    
# Devuelve valores de reference para ser insertados los select correspondientes en el template Inventarios al modificar 
# trademark de reactivo

class ReferencesByTrademarkAPI(LoginRequiredMixin,View):
    def get(self, request):
        lab = request.GET.get('lab')
        name = request.GET.get('name')
        trademark = request.GET.get('trademark')

        inventarios = Inventarios.objects.all()
        if lab:
            # Verificar si el valor de lab es un número
            if lab.isdigit():
                laboratorio = get_object_or_404(Laboratorios, id=int(lab))
                lab = laboratorio.id
            else:
                lab = get_object_or_404(Laboratorios, name=lab)
                lab = lab.id
        if name:
            # Verificar si el valor de name es un número
            if name.isdigit():
                reactivo = get_object_or_404(Reactivos, id=int(name))
                name = reactivo.id
            else:
                reactivo = get_object_or_404(Reactivos, name=name)
                name = reactivo.id

        if name:
            inventarios = inventarios.filter(name=name, lab=lab, trademark=trademark)

        references = inventarios.values('reference').distinct()
        references_list = list(references)

        return JsonResponse(references_list, safe=False)
    
# Devuelve al template los valores únicos de wlocation según el nombre del reactivo en la tabla del modelo Almacenamiento

class WlocationsAPI(LoginRequiredMixin,View):
    def get(self, request):
        lab = request.GET.get('lab')
        
        if lab:
            almacenamiento = Almacenamiento.objects.filter(lab__name=lab)
            wlocation_list = almacenamiento.values('id','name').distinct()
            # Agregar la opción "Seleccione" al principio de la lista
            wlocation_list = [{'id': '', 'name': 'Seleccione'}] + list(wlocation_list)
            return JsonResponse(list(wlocation_list), safe=False)

        return JsonResponse([], safe=False)

# Utilizando los valores filtrados en el template inventario.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Inventarios. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    lab = request.session.get('filtered_lab')
    id_reactivo = request.session.get('filtered_id_r')
    
    queryset = Inventarios.objects.all()
    #Filtra según los valores previos de filtro en los selectores

    # si el valor de lab viene de sesión superusuario o ADMINISTRADOR lab=0 cambiar por lab=''
    if lab=='0':
         lab=''
         
    # Definir Queryset para filtrado de visulización
         
    if lab and id_reactivo:
        queryset = queryset.filter(name=id_reactivo, lab=lab, is_active=True)
    elif lab:
        queryset = queryset.filter(lab=lab, is_active=True)
    elif id_reactivo:
        queryset = queryset.filter(name=id_reactivo, is_active=True)          
    else:
        queryset = queryset.filter(is_active=True)

    queryset = queryset.order_by('id')

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    # sheet.merge_cells('C1:F1')

    sheet['D1'] = 'Inventario de insumos'
    sheet['D2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Código'
    sheet['C4'] = 'CAS'
    sheet['D4'] = 'Reactivo'
    sheet['E4'] = 'Marca'
    sheet['F4'] = 'Referencia'
    sheet['G4'] = 'Cantidad'
    sheet['H4'] = 'Unidades'
    sheet['I4'] = 'Inventario mínimo'
    sheet['J4'] = 'Ubicación'
    sheet['K4'] = 'Laboratorio'
    sheet['L4'] = 'Vencimiento'
    sheet['M4'] = 'Registrado por'
    sheet['N4'] = 'Fecha y hora de Registro'
    sheet['O4'] = 'Actualizado por'
    sheet['P4'] = 'Fecha y hora de última Actualización'

    # Establecer la altura de la fila 1 y 2 a 30
    sheet.row_dimensions[1].height = 35
    sheet.row_dimensions[2].height = 35

    # Establecer estilo de celda para A1

    cell_A1 = sheet['D1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 7
    sheet.column_dimensions['A'].width = 7

    # Establecer el ancho de la columna B a 9
    sheet.column_dimensions['B'].width = 9

    # Establecer el ancho de la columna C a 13
    sheet.column_dimensions['C'].width = 13

    # Establecer el ancho de la columna D a 28
    sheet.column_dimensions['D'].width = 28

    # Establecer el ancho de la columna E a 13
    sheet.column_dimensions['E'].width = 13

    # Establecer el ancho de la columna F a 10
    sheet.column_dimensions['F'].width = 10

    # Establecer el ancho de la columna G a 8
    sheet.column_dimensions['G'].width = 8

    # Establecer el ancho de la columna H a 5
    sheet.column_dimensions['H'].width = 5

    # Establecer el ancho de la columna I a 10
    sheet.column_dimensions['I'].width = 10

    # Establecer el ancho de la columna J a 17
    sheet.column_dimensions['J'].width = 17
    # Establecer el ancho de la columna K a 13
    sheet.column_dimensions['K'].width = 13
    
    # Establecer el ancho de la columna L a 13
    sheet.column_dimensions['L'].width = 13
    # Establecer el ancho de la columna M a 23
    sheet.column_dimensions['M'].width = 23
    # Establecer el ancho de la columna N a 19
    sheet.column_dimensions['N'].width = 19
    # Establecer el ancho de la columna O a 23
    sheet.column_dimensions['O'].width = 23
    # Establecer el ancho de la columna P a 19
    sheet.column_dimensions['P'].width = 19

    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 17):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = item.name.code
        sheet.cell(row=row, column=3).value = item.name.cas
        sheet.cell(row=row, column=4).value = item.name.name
        sheet.cell(row=row, column=5).value = item.trademark.name
        sheet.cell(row=row, column=6).value = item.reference
        sheet.cell(row=row, column=7).value = item.weight
        sheet.cell(row=row, column=8).value = item.name.unit.name
        sheet.cell(row=row, column=9).value = f'{int(item.minstock)} {item.name.unit.name}'
        sheet.cell(row=row, column=10).value = item.wlocation.name
        sheet.cell(row=row, column=11).value = item.lab.name
        sheet.cell(row=row, column=12).value = item.edate
        sheet.cell(row=row, column=13).value = item.created_by.first_name+' '+item.created_by.last_name
        sheet.cell(row=row, column=14).value = str((item.date_create).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=15).value = item.last_updated_by.first_name+' '+item.last_updated_by.last_name
        sheet.cell(row=row, column=16).value = str((item.last_update).strftime('%d/%m/%Y %H:%M:%S'))

        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 17):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 16
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Inventario_Reactivos.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template listado_entradas.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Inventarios. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_input(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    lab = request.session.get('filtered_lab')
    id_reagent = request.session.get('filtered_id_reagent')
    
    start_date = request.session.get('filtered_start_date')
    end_date = request.session.get('filtered_end_date')
    
    
    # Validar y convertir las fechas
    try:
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
        pass

    queryset = Entradas.objects.all()
    #Filtra según los valores previos de filtro en los selectores

        
    # Realiza la filtración de acuerdo a las fechas
    if start_date:
        queryset = queryset.filter(date_create__gte=start_date)
    if end_date:
        queryset = queryset.filter(date_create__lte=end_date)
    elif start_date and end_date:
        queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)
    
    if lab and id_reagent:
        queryset = queryset.filter(name=id_reagent, lab=lab, is_active=True)
    elif lab:
        queryset = queryset.filter(lab=lab, is_active=True)
    elif id_reagent:
        queryset = queryset.filter(name=id_reagent, is_active=True)
    else:
        queryset = queryset.filter(is_active=True)
    queryset = queryset.order_by('id')
        

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    sheet.merge_cells('C1:F1')

    sheet['C1'] = 'Listado de movimientos de Entrada'
    sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Fecha y hora de registro'
    sheet['C4'] = 'Nombre del reactivo'
    sheet['D4'] = 'Código'
    sheet['E4'] = 'CAS'
    sheet['F4'] = 'Marca'
    sheet['G4'] = 'Referencia'
    sheet['H4'] = 'Cantidad'
    sheet['I4'] = 'Unidad'
    sheet['J4'] = 'Laboratorio'
    sheet['K4'] = 'Orden de Compra'
    sheet['L4'] = 'Fecha de orden'
    sheet['M4'] = 'Proyecto'
    sheet['N4'] = 'Valor de la compra'
    sheet['O4'] = 'Destino'
    sheet['P4'] = 'Responsable'
    sheet['Q4'] = 'Asignatura'
    sheet['R4'] = 'Facultad'
    sheet['S4'] = 'Registrado por'
    sheet['T4'] = 'Actualizado por'
    sheet['U4'] = 'Fecha y hora de última Actualización'
    sheet['V4'] = 'Observaciones'
    
    

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1

    cell_A1 = sheet['C1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 13
    sheet.column_dimensions['A'].width = 13

    # Establecer el ancho de la columna B a 23
    sheet.column_dimensions['B'].width = 23

    # Establecer el ancho de la columna C a 40
    sheet.column_dimensions['C'].width = 40

    # Establecer el ancho de la columna D a 9
    sheet.column_dimensions['D'].width = 9

    # Establecer el ancho de la columna E a 10
    sheet.column_dimensions['E'].width = 10

    # Establecer el ancho de la columna F a 14
    sheet.column_dimensions['F'].width = 14

    # Establecer el ancho de la columna G a 12
    sheet.column_dimensions['G'].width = 12

    # Establecer el ancho de la columna H a 10
    sheet.column_dimensions['H'].width = 10

    # Establecer el ancho de la columna I a 9
    sheet.column_dimensions['I'].width = 9

    # Establecer el ancho de la columna J a 51
    sheet.column_dimensions['J'].width = 51

    # Establecer el ancho de la columna K a 18
    sheet.column_dimensions['K'].width = 18

    # Establecer el ancho de la columna L a 16
    sheet.column_dimensions['L'].width = 16

    # Establecer el ancho de la columna M a 10
    sheet.column_dimensions['M'].width = 10

    # Establecer el ancho de la columna N a 19
    sheet.column_dimensions['N'].width = 19

    # Establecer el ancho de la columna O a 14
    sheet.column_dimensions['O'].width = 14

    # Establecer el ancho de la columna P a 28
    sheet.column_dimensions['P'].width = 28

    # Establecer el ancho de la columna Q a 37
    sheet.column_dimensions['Q'].width = 37

    # Establecer el ancho de la columna R a 36
    sheet.column_dimensions['R'].width = 36

    # Establecer el ancho de la columna S a 29
    sheet.column_dimensions['S'].width = 29

    # Establecer el ancho de la columna T a 29
    sheet.column_dimensions['T'].width = 29

    # Establecer el ancho de la columna U a 34
    sheet.column_dimensions['U'].width = 34

    # Establecer el ancho de la columna V a 60
    sheet.column_dimensions['V'].width = 60

    # Crear un estilo con formato de moneda
    currency_style = NamedStyle(name='currency', number_format='_("$"* #,##0.00_);_("$"* \(#,##0.00\);_("$"* "-"??_);_(@_)')
    

    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 23):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        if not item.date_order:
            item.date_order=''
        if item.location:
            subject=item.location.name
            school=item.location.facultad.name
        else:
            subject=''
            school=''
        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = str((item.date_create).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=3).value = item.name.name
        sheet.cell(row=row, column=4).value = item.name.code
        sheet.cell(row=row, column=5).value = item.name.cas
        sheet.cell(row=row, column=6).value = item.trademark.name
        sheet.cell(row=row, column=7).value = item.reference
        sheet.cell(row=row, column=8).value = item.weight
        sheet.cell(row=row, column=9).value = item.name.unit.name
        sheet.cell(row=row, column=10).value = item.lab.name
        sheet.cell(row=row, column=11).value = item.order
        sheet.cell(row=row, column=12).value = str(item.date_order)
        sheet.cell(row=row, column=13).value = item.nproject
        # Aplicar el estilo a la celda
        sheet.cell(row=row, column=14).style = currency_style
        sheet.cell(row=row, column=14).value = item.price
        sheet.cell(row=row, column=15).value = item.destination.name
        sheet.cell(row=row, column=16).value = item.manager.name
        sheet.cell(row=row, column=17).value = subject
        sheet.cell(row=row, column=18).value = school
        sheet.cell(row=row, column=19).value = item.created_by.first_name+' '+item.created_by.last_name
        sheet.cell(row=row, column=20).value = item.last_updated_by.first_name+' '+item.last_updated_by.last_name
        sheet.cell(row=row, column=21).value = str((item.last_update).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=22).value = item.observations

               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 23):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 22
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Registro_Entradas.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template listado_entradas.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Inventarios. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_event(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    
    start_date = request.session.get('filtered_start_date')
    end_date = request.session.get('filtered_end_date')
    user_id = request.session.get('filtered_id')
    
    
    # Validar y convertir las fechas
    try:
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
        pass

    queryset = Eventos.objects.all()
    #Filtra según los valores previos de filtro en los selectores

    # Realiza la filtración de acuerdo a las fechas
    if start_date:
        queryset = queryset.filter(fecha_evento__gte=start_date)
    if end_date:
        queryset = queryset.filter(fecha_evento__lte=end_date)
    elif start_date and end_date:
        queryset = queryset.filter(fecha_evento__gte=start_date,fecha_evento__lte=end_date)
    
    if user_id:
        queryset = queryset.filter(usuario_evento=user_id)
                
    queryset = queryset.order_by('-id')
        

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    sheet.merge_cells('C1:F1')

    sheet['C1'] = 'Listado de eventos'
    sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Fecha y hora de evento'
    sheet['C4'] = 'Tipo de evento'
    sheet['D4'] = 'Usuario'    

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1

    cell_A1 = sheet['C1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 6
    sheet.column_dimensions['A'].width = 6

    # Establecer el ancho de la columna B a 25
    sheet.column_dimensions['B'].width = 25

    # Establecer el ancho de la columna C a 34
    sheet.column_dimensions['C'].width = 34

    # Establecer el ancho de la columna D a 18
    sheet.column_dimensions['D'].width = 18

    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 5):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        
        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = str((item.fecha_evento).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=3).value = item.tipo_evento.name
        sheet.cell(row=row, column=4).value = f'{item.usuario_evento.first_name} {item.usuario_evento.last_name}'

               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 5):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 4
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Eventos.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template listado_salidas.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Inventarios. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_output(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    lab = request.session.get('filtered_lab')
    id_reagent = request.session.get('filtered_id_reagent')
    
    start_date = request.session.get('filtered_start_date')
    end_date = request.session.get('filtered_end_date')
    
    
    # Validar y convertir las fechas
    try:
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
        pass

    queryset = Salidas.objects.all()
    #Filtra según los valores previos de filtro en los selectores

        
    # Realiza la filtración de acuerdo a las fechas
    if start_date:
        queryset = queryset.filter(date_create__gte=start_date)
    if end_date:
        queryset = queryset.filter(date_create__lte=end_date)
    elif start_date and end_date:
        queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)
    
    if lab and id_reagent:
        queryset = queryset.filter(name=id_reagent, lab=lab, is_active=True)
    elif lab:
        queryset = queryset.filter(lab=lab, is_active=True)
    elif id_reagent:
        queryset = queryset.filter(name=id_reagent, is_active=True)
    else:
        queryset = queryset.filter(is_active=True)
    queryset = queryset.order_by('id')
    
        

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    sheet.merge_cells('C1:F1')

    sheet['C1'] = 'Listado de movimientos de Salida'
    sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Fecha y hora de registro'
    sheet['C4'] = 'Nombre del reactivo'
    sheet['D4'] = 'Código'
    sheet['E4'] = 'CAS'
    sheet['F4'] = 'Marca'
    sheet['G4'] = 'Referencia'
    sheet['H4'] = 'Cantidad'
    sheet['I4'] = 'Unidad'
    sheet['J4'] = 'Laboratorio'
    sheet['K4'] = 'Destino'
    sheet['L4'] = 'Responsable'
    sheet['M4'] = 'Asignatura'
    sheet['N4'] = 'Facultad'
    sheet['O4'] = 'Registrado por'
    sheet['P4'] = 'Actualizado por'
    sheet['Q4'] = 'Fecha y hora de última Actualización'
    sheet['R4'] = 'Observaciones'
    
    

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1

    cell_A1 = sheet['C1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 13
    sheet.column_dimensions['A'].width = 13

    # Establecer el ancho de la columna B a 23
    sheet.column_dimensions['B'].width = 23

    # Establecer el ancho de la columna C a 40
    sheet.column_dimensions['C'].width = 40

    # Establecer el ancho de la columna D a 9
    sheet.column_dimensions['D'].width = 9

    # Establecer el ancho de la columna E a 10
    sheet.column_dimensions['E'].width = 10

    # Establecer el ancho de la columna F a 14
    sheet.column_dimensions['F'].width = 14

    # Establecer el ancho de la columna G a 12
    sheet.column_dimensions['G'].width = 12

    # Establecer el ancho de la columna H a 10
    sheet.column_dimensions['H'].width = 10

    # Establecer el ancho de la columna I a 9
    sheet.column_dimensions['I'].width = 9

    # Establecer el ancho de la columna J a 35
    sheet.column_dimensions['J'].width = 35

    # Establecer el ancho de la columna K a 18
    sheet.column_dimensions['K'].width = 18

    # Establecer el ancho de la columna L a 28
    sheet.column_dimensions['L'].width = 28

    # Establecer el ancho de la columna M a 26
    sheet.column_dimensions['M'].width = 26

    # Establecer el ancho de la columna N a 36
    sheet.column_dimensions['N'].width = 36

    # Establecer el ancho de la columna O a 29
    sheet.column_dimensions['O'].width = 29

    # Establecer el ancho de la columna P a 29
    sheet.column_dimensions['P'].width = 29

    # Establecer el ancho de la columna Q a 34
    sheet.column_dimensions['Q'].width = 34

    # Establecer el ancho de la columna R a 45
    sheet.column_dimensions['R'].width = 45

    # Establecer el ancho de la columna S a 29
    sheet.column_dimensions['S'].width = 29
    
    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 19):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        
        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = str((item.date_create).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=3).value = item.name.name
        sheet.cell(row=row, column=4).value = item.name.code
        sheet.cell(row=row, column=5).value = item.name.cas
        sheet.cell(row=row, column=6).value = item.trademark.name
        sheet.cell(row=row, column=7).value = item.reference
        sheet.cell(row=row, column=8).value = item.weight
        sheet.cell(row=row, column=9).value = item.name.unit.name
        sheet.cell(row=row, column=10).value = item.lab.name
        sheet.cell(row=row, column=11).value = item.destination.name
        sheet.cell(row=row, column=12).value = item.manager.name
        sheet.cell(row=row, column=13).value = item.location.name
        sheet.cell(row=row, column=14).value = item.location.facultad.name
        sheet.cell(row=row, column=15).value = item.created_by.first_name+' '+item.created_by.last_name
        sheet.cell(row=row, column=16).value = item.last_updated_by.first_name+' '+item.last_updated_by.last_name
        sheet.cell(row=row, column=17).value = str((item.last_update).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=18).value = item.observations

               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 19):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 18
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Registro_Salidas.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template listado_solicitudes.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Solicitudes. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite

@login_required
def export_to_excel_solicitud(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario    
    start_date = request.session.get('filtered_start_date')
    end_date = request.session.get('filtered_end_date')
    
    
    # Validar y convertir las fechas
    try:
        if start_date:
            start_date = make_aware(datetime.strptime(start_date, '%Y-%m-%d'))
        if end_date:
            end_date = make_aware(datetime.strptime(end_date, '%Y-%m-%d').replace(hour=23, minute=59, second=59))
    except ValueError:
            # Manejar errores de formato de fecha aquí si es necesario
        pass

    queryset = Solicitudes.objects.all()
    #Filtra según los valores previos de filtro en los selectores
        
    # Realiza la filtración de acuerdo a las fechas
    if start_date:
        queryset = queryset.filter(date_create__gte=start_date)
    if end_date:
        queryset = queryset.filter(date_create__lte=end_date)
    elif start_date and end_date:
        queryset = queryset.filter(date_create__gte=start_date,date_create__lte=end_date)
    
    queryset = queryset.order_by('id')
        
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    sheet.merge_cells('C1:F1')

    sheet['C1'] = 'Listado de solicitudes'
    sheet['C2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Fecha y hora de registro'
    sheet['C4'] = 'Tipo de solicitud'
    sheet['D4'] = 'Asunto'
    sheet['E4'] = 'Mensaje'
    sheet['F4'] = 'Archivos Adjuntos'
    sheet['G4'] = 'Usuario que registra'
    sheet['H4'] = 'Laboratorio al que pertenece'
    sheet['I4'] = 'Correo Usuario'
    sheet['J4'] = 'Teléfono Usuario'
    sheet['K4'] = 'Estado de la solicitud'
    sheet['L4'] = 'Fecha de respuesta'
    sheet['M4'] = 'Respuesta'
    sheet['N4'] = 'Usuario que responde'

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1

    cell_A1 = sheet['C1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 13
    sheet.column_dimensions['A'].width = 13

    # Establecer el ancho de la columna B a 23
    sheet.column_dimensions['B'].width = 23

    # Establecer el ancho de la columna C a 30
    sheet.column_dimensions['C'].width = 30

    # Establecer el ancho de la columna D a 30
    sheet.column_dimensions['D'].width = 30

    # Establecer el ancho de la columna E a 30
    sheet.column_dimensions['E'].width = 30

    # Establecer el ancho de la columna F a 11
    sheet.column_dimensions['F'].width = 11

    # Establecer el ancho de la columna G a 22
    sheet.column_dimensions['G'].width = 22

    # Establecer el ancho de la columna H a 36
    sheet.column_dimensions['H'].width = 36

    # Establecer el ancho de la columna I a 24
    sheet.column_dimensions['I'].width = 24

    # Establecer el ancho de la columna J a 17
    sheet.column_dimensions['J'].width = 17

    # Establecer el ancho de la columna K a 21
    sheet.column_dimensions['K'].width = 21

    # Establecer el ancho de la columna L a 19
    sheet.column_dimensions['L'].width = 19

    # Establecer el ancho de la columna M a 30
    sheet.column_dimensions['M'].width = 30

    # Establecer el ancho de la columna N a 22
    sheet.column_dimensions['N'].width = 22


    # Define una alineación que tenga la vertical en la parte superior y la horizontal a la izquierda
    # Crear un objeto de alineación
    alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")

    
    
    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 15):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:

        if item.tramitado:
            usuario_tramita=f'{item.usuario_tramita.first_name} {item.usuario_tramita.last_name}'
            fecha_tramite=str((item.fecha_tramite).strftime('%d/%m/%Y %H:%M:%S'))
            observaciones=item.observaciones
            estado_tramite='Tramitado'
        else:
            usuario_tramita=''
            fecha_tramite=''
            observaciones=''
            estado_tramite='Pendiente trámite'

        if item.archivos_adjuntos:
            protocol = 'https' if request.is_secure() else 'http'
            domain = request.get_host()
            url = f'{protocol}://{domain}{item.archivos_adjuntos.url}'

            # Establecer un estilo para la celda
            estilo_celda = sheet.cell(row=row, column=6)
            estilo_celda.value = "Descargar"

            # Crear un objeto Hyperlink con la URL
            hyperlink = Hyperlink(target=url, ref=f'A{row}')

            # Aplicar el hipervínculo a la celda
            estilo_celda.hyperlink = hyperlink
        else:
            sheet.cell(row=row, column=6).value = ""

        sheet.cell(row=row, column=1).value = '{:04d}'.format(item.id)
        sheet.cell(row=row, column=2).value = str((item.date_create).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=3).value = str(item.tipo_solicitud)
        sheet.cell(row=row, column=4).value = item.name
        sheet.cell(row=row, column=5).value = item.mensaje
        # sheet.cell(row=row, column=6).value = Descargar
        sheet.cell(row=row, column=7).value = f'{item.created_by.first_name} {item.created_by.last_name}'
        sheet.cell(row=row, column=8).value = item.created_by.lab.name
        sheet.cell(row=row, column=9).value = str(item.created_by.email)
        sheet.cell(row=row, column=10).value =item.created_by.phone_number
        sheet.cell(row=row, column=11).value = estado_tramite
        sheet.cell(row=row, column=12).value = fecha_tramite
        sheet.cell(row=row, column=13).value = observaciones
        sheet.cell(row=row, column=14).value = usuario_tramita

               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 15):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).alignment = alignment

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 14
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_Solicitudes.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response


# Utilizando los valores filtrados en el template listado_usuarios.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo User. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_user(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    
    lab = request.session.get('filtered_lab')
    user_id = request.session.get('filtered_id')
    
    queryset = User.objects.all()
    # Filtra según los valores previos de filtro en los selectores
    
    if lab=='0':
             lab=None

    if lab:
            queryset = queryset.filter(lab=lab)
    if user_id:
            queryset = queryset.filter(id=user_id)
                
    queryset = queryset.order_by('id')
        

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    #sheet.merge_cells('C1:F1')

    sheet['D1'] = 'Listado de Usuarios'
    sheet['D2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Id usuario'
    sheet['B4'] = 'Nombres'
    sheet['C4'] = 'Apellidos'
    sheet['D4'] = 'Identificación'
    sheet['E4'] = 'Correo Electrónico'
    sheet['F4'] = 'Teléfono'
    sheet['G4'] = 'Laboratorio al que pertenece'
    sheet['H4'] = 'Rol de usuario'
    sheet['I4'] = 'Último Acceso'
    sheet['J4'] = 'Acepta tratamiento de datos personales'
    sheet['K4'] = 'Activo'
    sheet['L4'] = 'Fecha de creación'
    sheet['M4'] = 'Creado por'
    sheet['N4'] = 'Fecha de última actualización'
    sheet['O4'] = 'Última actualización por'

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1
    cell_A1 = sheet['D1']
    cell_A1.font = Font(bold=True, size=15)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 9
    sheet.column_dimensions['A'].width = 9

    # Establecer el ancho de la columna B a 21
    sheet.column_dimensions['B'].width = 21

    # Establecer el ancho de la columna C a 21
    sheet.column_dimensions['C'].width = 21

    # Establecer el ancho de la columna D a 16
    sheet.column_dimensions['D'].width = 16

    # Establecer el ancho de la columna E a 29
    sheet.column_dimensions['E'].width = 29

    # Establecer el ancho de la columna F a 12
    sheet.column_dimensions['F'].width = 12

    # Establecer el ancho de la columna G a 38
    sheet.column_dimensions['G'].width = 38

    # Establecer el ancho de la columna H a 15
    sheet.column_dimensions['H'].width = 15

    # Establecer el ancho de la columna I a 18
    sheet.column_dimensions['I'].width = 18

    # Establecer el ancho de la columna J a 13
    sheet.column_dimensions['J'].width = 13

    # Establecer el ancho de la columna K a 8
    sheet.column_dimensions['K'].width = 8

    # Establecer el ancho de la columna L a 19
    sheet.column_dimensions['L'].width = 19

    # Establecer el ancho de la columna M a 29
    sheet.column_dimensions['M'].width = 29

    # Establecer el ancho de la columna N a 19
    sheet.column_dimensions['N'].width = 19

    # Establecer el ancho de la columna O a 29
    sheet.column_dimensions['O'].width = 29
    
    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 16):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        # Establecer acepta procesamientod e datos
        if item.acceptDataProcessing:
            item.acceptDataProcessing='Acepta'
        else:
            item.acceptDataProcessing='No acepta'
        # Establecer el is_active como un valor más comprensible que True y False
        if item.is_active==True:
            item.is_active='Activo'
        else:
            item.is_active='Inactivo'  

        # Establecer fecha de acceso
        if item.last_login:
            item.last_login=item.last_login.strftime('%d/%m/%Y %H:%M:%S')
        else:
            item.last_login='No access'
        
        if item.user_create:
            creator=f'{item.user_create.first_name} {item.user_create.last_name}'
        else:
            creator=''
              

        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = item.first_name
        sheet.cell(row=row, column=3).value = item.last_name
        sheet.cell(row=row, column=4).value = item.id_number
        sheet.cell(row=row, column=5).value = item.email
        sheet.cell(row=row, column=6).value = item.phone_number
        sheet.cell(row=row, column=7).value = item.lab.name
        sheet.cell(row=row, column=8).value = item.rol.name
        sheet.cell(row=row, column=9).value = item.last_login
        sheet.cell(row=row, column=10).value = item.acceptDataProcessing
        sheet.cell(row=row, column=11).value = item.is_active
        sheet.cell(row=row, column=12).value = item.date_joined.strftime('%d/%m/%Y %H:%M:%S')
        sheet.cell(row=row, column=13).value = creator
        sheet.cell(row=row, column=14).value = str((item.last_update).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=15).value = item.last_updated_by.first_name+' '+item.last_updated_by.last_name
        
               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 16):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 15
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_de_Usuarios.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response


# Utilizando los valores filtrados en el template listado_reactivos.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Reactivos. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_react(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    
    name = request.session.get('filtered_name')
    
    queryset = Reactivos.objects.all()
    #Filtra según los valores previos de filtro en los selectores
    # 
    if name:
        queryset = queryset.filter(name=name)
            
    queryset = queryset.order_by('id')
    # Verificar si no se encontraron resultados
    if not queryset.exists():
        print("No hay criterios de búsqueda que coincidan")
        

    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    #sheet.merge_cells('C1:F1')

    sheet['D1'] = 'Listado de Reactivos '
    sheet['D2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Consecutivo'
    sheet['B4'] = 'Código'
    sheet['C4'] = 'Cas'
    sheet['D4'] = 'Nombre'
    sheet['E4'] = 'Almacenamiento Interno'
    sheet['F4'] = 'Clase de almacenamiento'
    sheet['G4'] = 'Estado'
    sheet['H4'] = 'Unidades'
    sheet['I4'] = 'Activo'
    sheet['J4'] = 'Fecha de creación'
    sheet['K4'] = 'Creado por'
    sheet['L4'] = 'Fecha de última actualización por'
    sheet['M4'] = 'Última actualización por'

    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1
    cell_A1 = sheet['D1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    # Establecer el ancho de la columna A a 9
    sheet.column_dimensions['A'].width = 9

    # Establecer el ancho de la columna B a 13
    sheet.column_dimensions['B'].width = 13

    # Establecer el ancho de la columna C a 16
    sheet.column_dimensions['C'].width = 16

    # Establecer el ancho de la columna D a 45
    sheet.column_dimensions['D'].width = 45

    # Establecer el ancho de la columna E a 24
    sheet.column_dimensions['E'].width = 24

    # Establecer el ancho de la columna F a 68
    sheet.column_dimensions['F'].width = 68

    # Establecer el ancho de la columna G a 9
    sheet.column_dimensions['G'].width = 9

    # Establecer el ancho de la columna H a 6
    sheet.column_dimensions['H'].width = 6

    # Establecer el ancho de la columna I a 9
    sheet.column_dimensions['I'].width = 9

    # Establecer el ancho de la columna J a 18
    sheet.column_dimensions['J'].width = 18

    # Establecer el ancho de la columna K a 23
    sheet.column_dimensions['K'].width = 23

    # Establecer el ancho de la columna L a 18
    sheet.column_dimensions['L'].width = 18

    # Establecer el ancho de la columna M a 23
    sheet.column_dimensions['M'].width = 23
    
    row = 4
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 14):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font

    row = 5
    for item in queryset:
        
        # Establecer el is_active como un valor más comprensible que True y False
        if item.is_active==True:
            item.is_active='Activo'
        else:
            item.is_active='Inactivo'        

        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = item.code
        sheet.cell(row=row, column=3).value = item.cas
        sheet.cell(row=row, column=4).value = item.name
        sheet.cell(row=row, column=5).value = str(item.almacenamiento_interno)
        sheet.cell(row=row, column=6).value = item.clase_almacenamiento.name+' '+item.clase_almacenamiento.description
        sheet.cell(row=row, column=7).value =  str(item.state)
        sheet.cell(row=row, column=8).value =  str(item.unit)
        sheet.cell(row=row, column=9).value =  item.is_active
        sheet.cell(row=row, column=10).value = str((item.date_create).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=11).value = item.created_by.first_name+' '+item.created_by.last_name
        sheet.cell(row=row, column=12).value = str((item.last_update).strftime('%d/%m/%Y %H:%M:%S'))
        sheet.cell(row=row, column=13).value = item.last_updated_by.first_name+' '+item.last_updated_by.last_name
               
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 14):
            sheet.cell(row=row, column=col).border = thin_border

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 13
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_de_Reactivos.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template listado_reactivos.html, y guardados en los datos de sesión, se crea el archivo de Excel 
# correspondiente e se introducen los valores desde la tabla del modelo Reactivos. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila, y se añaden filtros 
# a los encabezados en caso de que el usuario lo solicite
@login_required
def export_to_excel_lab(request):
    
    queryset = Laboratorios.objects.all()
    
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Ruta al archivo de imagen del logotipo

    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen y procesarla con pillow
    pil_image = PILImage.open(logo_path)

    # Crear un objeto Image de openpyxl a partir de la imagen procesada
    image = ExcelImage(pil_image)

    # Anclar la imagen a la celda A1
    sheet.add_image(image, 'A1')

    # Obtener la fecha actual
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M:%S")

    # Unificar las celdas A1, B1, C1 y D1
    #sheet.merge_cells('C1:F1')

    sheet['D1'] = 'Listado de Laboratorios'
    sheet['D2'] = 'Fecha de Creación: '+fecha_creacion
    sheet['A4'] = 'Id'
    sheet['B4'] = 'Laboratorio'
    sheet['C4'] = 'Administradores'
    sheet['D4'] = 'Teléfonos'
    sheet['E4'] = 'Correos'
    sheet['F4'] = 'Coordinadores'
    sheet['G4'] = 'Teléfonos'
    sheet['H4'] = 'Correos'
    sheet['I4'] = 'Técnicos'
    sheet['J4'] = 'Teléfonos'
    sheet['K4'] = 'Correos'
    # Establecer la altura de la fila 1 y 2 a 30 y fila 3 a 25
    sheet.row_dimensions[1].height = 30
    sheet.row_dimensions[2].height = 30
    sheet.row_dimensions[3].height = 25

    # Establecer estilo de celda para A1
    cell_A1 = sheet['D1']
    cell_A1.font = Font(bold=True, size=16)

    # Configurar los estilos de borde
    thin_border = Border(left=Side(style='thin'), right=Side(
    style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

    # Establecer el estilo de las celdas A2:D3
    bold_font = Font(bold=True)

    

    # Establecer el ancho de la columna A a 4
    sheet.column_dimensions['A'].width = 4

    # Establecer el ancho de la columna B a 52
    sheet.column_dimensions['B'].width = 52

    # Establecer el ancho de la columna C a 32
    sheet.column_dimensions['C'].width = 32
       
    # Establecer el ancho de la columna D a 13
    sheet.column_dimensions['D'].width = 13

    # Establecer el ancho de la columna E a 32
    sheet.column_dimensions['E'].width = 32

    # Establecer el ancho de la columna F a 32
    sheet.column_dimensions['F'].width = 32

    # Establecer el ancho de la columna G a 13
    sheet.column_dimensions['G'].width = 13

    # Establecer el ancho de la columna H a 32
    sheet.column_dimensions['H'].width = 32

    # Establecer el ancho de la columna I a 32
    sheet.column_dimensions['I'].width = 32

    # Establecer el ancho de la columna j a 13
    sheet.column_dimensions['J'].width = 13

    # Establecer el ancho de la columna K a 32
    sheet.column_dimensions['K'].width = 32


    row = 4
    # Define una alineación que tenga la vertical en la parte superior y la horizontal a la izquierda
    # Crear un objeto de alineación
    alignment = Alignment(wrap_text=True, vertical="top", horizontal="left")
    # Aplicar el estilo de borde a las celdas de la fila actual
    for col in range(1, 12):
        sheet.cell(row=row, column=col).border = thin_border
        sheet.cell(row=row, column=col).font = bold_font
        

    

    row = 5
    for item in queryset:
        # Obtén los usuarios que pertenecen a este laboratorio con rol adinistrador
        rol = get_object_or_404(Rol, name='ADMINISTRADOR')

        # Obtén una lista de coordinadores que cumplan con las condiciones
        administradores = User.objects.filter(lab=item.id, rol=rol.id, is_active=True)

        # Inicializa una lista para almacenar los nombres de los coordinadores
        administradores_nombres = []
        administradores_telefonos = []
        administradores_correos = []

        for administrador in administradores:
            administrador_nombre = f"{administrador.first_name} {administrador.last_name}"
            administrador_telefono = f"{administrador.phone_number}"
            administrador_correo = f"{administrador.email}"


            administradores_nombres.append(administrador_nombre)
            administradores_telefonos.append(administrador_telefono)
            administradores_correos.append(administrador_correo)
            
        # Verifica si la lista de coordinadores está vacía
        if not administradores_nombres:
            administradores_names = ''
            administradores_phones = ''
            administradores_emails = ''
        else:
            # Formatea la lista de coordinadores como una cadena con saltos de línea
            administradores_names = ', \n'.join(administradores_nombres)
            administradores_phones = ', \n'.join(administradores_telefonos)
            administradores_emails = ', \n'.join(administradores_correos)
        # -----------------------------------------------------------------------
        # Obtén los usuarios que pertenecen a este laboratorio con rol coordinador
        rol = get_object_or_404(Rol, name='COORDINADOR')

        # Obtén una lista de coordinadores que cumplan con las condiciones
        coordinadores = User.objects.filter(lab=item.id, rol=rol.id, is_active=True)

        # Inicializa una lista para almacenar los nombres de los coordinadores
        coordinadores_nombres = []
        coordinadores_telefonos = []
        coordinadores_correos = []

        for coordinador in coordinadores:
            coordinador_nombre = f"{coordinador.first_name} {coordinador.last_name}"
            coordinador_telefono = f"{coordinador.phone_number}"
            coordinador_correo = f"{coordinador.email}"


            coordinadores_nombres.append(coordinador_nombre)
            coordinadores_telefonos.append(coordinador_telefono)
            coordinadores_correos.append(coordinador_correo)
            
        # Verifica si la lista de coordinadores está vacía
        if not coordinadores_nombres:
            coordinadores_names = ''
            coordinadores_phones = ''
            coordinadores_emails = ''
        else:
            # Formatea la lista de coordinadores como una cadena con saltos de línea
            coordinadores_names = ', \n'.join(coordinadores_nombres)
            coordinadores_phones = ', \n'.join(coordinadores_telefonos)
            coordinadores_emails = ', \n'.join(coordinadores_correos)
        
        # -----------------------------------------------------------------------
        # Obtén los usuarios que pertenecen a este laboratorio con rol ADMINISTRADOR
        rol = get_object_or_404(Rol, name='TECNICO')

        # Obtén una lista de tecnicos que cumplan con las condiciones
        tecnicos = User.objects.filter(lab=item.id, rol=rol.id, is_active=True)

        # Inicializa una lista para almacenar los nombres de los coordinadores
        tecnicos_nombres = []
        tecnicos_telefonos = []
        tecnicos_correos = []

        for tecnico in tecnicos:
            tecnico_nombre = f"{tecnico.first_name} {tecnico.last_name}"
            tecnico_telefono = f"{tecnico.phone_number}"
            tecnico_correo = f"{tecnico.email}"


            tecnicos_nombres.append(tecnico_nombre)
            tecnicos_telefonos.append(tecnico_telefono)
            tecnicos_correos.append(tecnico_correo)
            
        # Verifica si la lista de coordinadores está vacía
        if not tecnicos_nombres:
            tecnicos_names = ''
            tecnicos_phones = ''
            tecnicos_emails = ''
        else:
            # Formatea la lista de coordinadores como una cadena con saltos de línea
            tecnicos_names = ', \n'.join(tecnicos_nombres)
            tecnicos_phones = ', \n'.join(tecnicos_telefonos)
            tecnicos_emails = ', \n'.join(tecnicos_correos)

        
        # Establecer el alto de la fila actual a 50
        sheet.row_dimensions[row].height = 50   

        sheet.cell(row=row, column=1).value = item.id
        sheet.cell(row=row, column=2).value = item.name
        sheet.cell(row=row, column=3).value = str(administradores_names)
        sheet.cell(row=row, column=4).value = str(administradores_phones)
        sheet.cell(row=row, column=5).value = str(administradores_emails)
        sheet.cell(row=row, column=6).value = str(coordinadores_names)
        sheet.cell(row=row, column=7).value = str(coordinadores_phones)
        sheet.cell(row=row, column=8).value = str(coordinadores_emails)
        sheet.cell(row=row, column=9).value = str(tecnicos_names)
        sheet.cell(row=row, column=10).value = str(tecnicos_phones)
        sheet.cell(row=row, column=11).value = str(tecnicos_emails)
              
        # Aplicar el estilo de borde a las celdas de la fila actual
        for col in range(1, 12):
            sheet.cell(row=row, column=col).border = thin_border
            sheet.cell(row=row, column=col).alignment = alignment

        row += 1

    # Obtén el rango de las columnas de la tabla
    start_column = 1
    end_column = 11
    start_row = 4
    end_row = row - 1

    # Convertir los números de las columnas en letras de columna
    start_column_letter = get_column_letter(start_column)
    end_column_letter = get_column_letter(end_column)

    # Rango de la tabla con el formato "A4:I{n}", donde n es el número de filas en la tabla
    table_range = f"{start_column_letter}{start_row}:{end_column_letter}{end_row}"

    # Agregar filtros solo a las columnas de la tabla
    sheet.auto_filter.ref = table_range

    # Establecer fondo blanco desde la celda A1 hasta el final de la tabla

    fill = PatternFill(fill_type="solid", fgColor=WHITE)
    start_cell = sheet['A1']
    end_column_letter = get_column_letter(end_column+1)
    end_row = row+1
    end_cell = sheet[end_column_letter + str(end_row)]
    table_range = start_cell.coordinate + ':' + end_cell.coordinate

    for row in sheet[table_range]:
        for cell in row:
            cell.fill = fill

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=Listado_de_Laboratorios.xlsx'

    workbook.save(response)
    # crear evento descarga de archivos
    tipo_evento = 'DESCARGA DE ARCHIVOS'
    usuario_evento = request.user
    crear_evento(tipo_evento, usuario_evento)

    return response

# Utilizando los valores filtrados en el template inventario.html, y guardados en los datos de sesión, se crea el archivo PDF 
# correspondiente e se introducen los valores desde la tabla del modelo Inventarios. Además, se aplican formatos a los encabezados, se 
# coloca un título, la fecha de creación y el logo. También se ajustan los anchos de columna y las alturas de fila
@login_required
def export_to_pdf(request):
    # Obtener los valores filtrados almacenados en la sesión del usuario
    lab = request.session.get('filtered_lab')
    name = request.session.get('filtered_name')
    trademark = request.session.get('filtered_trademark')

    queryset = Inventarios.objects.all()

    if lab and name and trademark:
        queryset = queryset.filter(lab=lab, name=name, trademark=trademark, is_active=True)
    elif lab and name:
        queryset = queryset.filter(lab=lab, name=name, is_active=True)
    elif lab and trademark:
        queryset = queryset.filter(lab=lab, trademark=trademark, is_active=True)
    elif name and trademark:
        queryset = queryset.filter(name=name, trademark=trademark, is_active=True)
    elif lab:
        queryset = queryset.filter(lab=lab, is_active=True)
    elif name:
        queryset = queryset.filter(name=name, is_active=True)
    elif trademark:
        queryset = queryset.filter(trademark=trademark, is_active=True)
    else:
        queryset = queryset.filter(is_active=True)
    context = {
        'object_list': queryset
    }

    template = get_template('reactivos/inventario.html')
    context = {
        'object_list': queryset
    }

    html = template.render(context)
    buffer = BytesIO()

    # Crear el archivo PDF
    p = canvas.Canvas(buffer, pagesize=landscape(letter))

    # Establecer el tamaño y posición del encabezado
    header_width = 600
    header_height = 40
    header_x = 40
    header_y = 580

    # Cargar la imagen del logotipo
    logo_path = finders.find('inventarioreac/Images/escudoUnal_black.png')

    # Cargar la imagen del logotipo y redimensionarla
    # Cargar la imagen del logotipo
    p.drawInlineImage(logo_path, header_x + 10,
                      header_y - 50, width=120, height=70)

    # Dibujar el título del informe con la fecha
    fecha_creacion = datetime.now().strftime("%d/%m/%Y %H:%M")
    title_text = f"Inventario de Insumos \n{fecha_creacion}"
    p.setFont("Helvetica-Bold", 14)

    # Dividir eltexto del título en líneas más cortas
    lines = title_text.split('\n')

    # Calcular la altura total del texto del título
    title_height = len(lines) * 5

    # Calcular la posición vertical inicial del título
    title_y = header_y - title_height

    # Dibujar cada línea del texto del título justificada
    for line in lines:
        p.drawRightString(header_x - 50 + header_width, title_y, line)
        title_y -= 20

    # Establecer el tamaño y posición de la tabla
    table_x = header_x 
    table_y = header_y - header_height - 30

    # Dibujar la tabla en el PDF
    p.setFont("Helvetica-Bold", 10)  # Configurar la fuente en negrita

    # Ajustar el espaciado inferior de la línea o el espaciado superior del texto para los encabezados
    # Espaciado adicional entre la línea y el texto inferior de los encabezados
    line_space_bottom_header = 16

    # Dibujar línea debajo de los encabezados
    # Establecer el color de la línea (más tenue)
    p.setStrokeColorRGB(0.8, 0.8, 0.8)
    p.setLineWidth(0.5)  # Establecer el grosor de la línea (más delgada)
    p.line(table_x-30, table_y - line_space_bottom_header+8, table_x +
           748, table_y - line_space_bottom_header+8)  # Dibujar línea

    p.drawString(table_x - 30, table_y, "Código")
    p.drawString(table_x + 11, table_y, "CAS")
    p.drawString(table_x + 50, table_y, "Reactivo")
    p.drawString(table_x + 235, table_y, "Marca")
    p.drawString(table_x + 318, table_y, "Referencia")
    p.drawString(table_x + 375, table_y, "Cant.")
    p.drawString(table_x + 420, table_y, "Ud.")
    p.drawString(table_x + 440, table_y, "Ubicación")
    p.drawString(table_x + 530, table_y, "Laboratorio")
    p.drawString(table_x + 687, table_y, "Vencimiento")

    p.setFont("Helvetica", 12)  # Restaurar la configuración de la fuente

    row_height = 20
    max_rows = 18  # Máximo número de filas por página
    current_row = 0
    page_number = 1

    for item in queryset:
        # Ajustar el espaciado inferior de la línea o el espaciado superior del texto
        line_space_bottom = 16  # Espaciado adicional entre la línea y el texto inferior
        text_space_top = 6  # Espaciado adicional entre el texto superior y la línea

        if current_row == max_rows:
            # Se alcanzó el límite de filas en la página actual, crear una nueva página
            p.showPage()
            p.setFont("Helvetica-Bold", 10)  # Configurar la fuente en negrita

            # Dibujar línea debajo de los encabezados en la nueva página
            p.setStrokeColorRGB(0.8, 0.8, 0.8)
            p.setLineWidth(0.5)
            p.line(table_x-30, table_y - line_space_bottom_header+8, table_x +
                   748, table_y - line_space_bottom_header+8)  # Dibujar línea

            p.drawString(table_x - 30, table_y, "Código")
            p.drawString(table_x + 11, table_y, "CAS")
            p.drawString(table_x + 50, table_y, "Reactivo")
            p.drawString(table_x + 235, table_y, "Marca")
            p.drawString(table_x + 318, table_y, "Referencia")
            p.drawString(table_x + 375, table_y, "Cant.")
            p.drawString(table_x + 420, table_y, "Ud.")
            p.drawString(table_x + 440, table_y, "Ubicación")
            p.drawString(table_x + 530, table_y, "Laboratorio")
            p.drawString(table_x + 687, table_y, "Vencimiento")

            p.setFont("Helvetica", 10)  # Restaurar la configuración de la fuente

            current_row = 0  # Reiniciar el contador de filas
            page_number += 1  # Incrementar el número de página

        # Dibujar línea debajo de la fila
        p.setStrokeColorRGB(0.8, 0.8, 0.8)
        p.setLineWidth(0.5)
        p.line(table_x-30, table_y - (line_space_bottom + (row_height * current_row))-16, table_x +
               748, table_y - (line_space_bottom + (row_height * current_row))-16) # Dibujar línea

        # Ajustar la posición del texto
        row_y = table_y - (text_space_top + (row_height * current_row))-20

        # Dibujar los contenidos de texto
        p.setFont("Helvetica", 10)
        p.drawString(table_x - 30, row_y, str(item.name.code))
        p.drawString(table_x + 11, row_y, str(item.name.cas))
        p.drawString(table_x + 50, row_y, str(item.name.name))
        p.drawString(table_x + 235, row_y, str(item.trademark.name))
        p.drawString(table_x + 318, row_y, str(item.reference))
        # Formatear a 1 posición decimal
        p.drawString(table_x + 375, row_y, "{:.1f}".format(item.weight))
        p.drawString(table_x + 420, row_y, str(item.name.unit.name))
        p.drawString(table_x + 440, row_y, str(item.wlocation.name))
        p.drawString(table_x + 530, row_y, str(item.lab.name))
        p.drawString(table_x + 687, row_y, str(item.edate))

        current_row += 1

    p.showPage()
    p.save()

    # Obtener el contenido del archivo PDF generado
    pdf_file = buffer.getvalue()
    buffer.close()

    # Enviar el archivo PDF como respuesta al navegador
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = 'attachment; filename="Inventario de Reactivos.pdf"'

    response.write(pdf_file)

    return response




# Toma los valores enviados desde el template registrar_entrada.html o registrar_salida.html, consulta en la tabla reactivos y devuelve
# los valores de code, cas, state, y unit, para que se actualicen en los campos de entrada de los formualrios correspondientes
@login_required
def get_value(request):
    if request.META.get('HTTP_X_REQUESTED_WITH') == 'XMLHttpRequest':
        # Si la solicitud es una solicitud AJAX, procesar la solicitud y devolver una respuesta JSON
        value_selected = request.GET.get('value_selected')
        try:
            # Intentar obtener el valor correspondiente de la base de datos
            reactivo = Reactivos.objects.get(name=value_selected)
            cas = reactivo.cas
            codigo = reactivo.code
            liquid = reactivo.state.name
            nombre_unit = reactivo.unit.name  # Obtener el nombre de la unidad
        except Reactivos.DoesNotExist:
            # Si el reactivo no existe, devolver un valor por defecto
            cas = ""
            codigo = ""
            liquid = ''
            nombre_unit = ""

        return JsonResponse({
            'cas': cas,
            'codigo': codigo,
            'liquid': liquid,
            'nombre_unit': nombre_unit  # Pasar el nombre de la unidad
        })
    else:
        # Si la solicitud no es una solicitud AJAX, devolver una respuesta HTTP 400 Bad Request
        return HttpResponseBadRequest()

# Devuelve los valores de la tabla Reactivos según lo escrito en el campo name del formulario registrar_salida.html en forma de una 
# lista de autocompletado
@login_required
def autocomplete(request):
    term = request.GET.get('term', '')
    reactivos = Reactivos.objects.filter((Q(name__icontains=term) | Q(
        code__icontains=term) | Q(cas__icontains=term)),is_active=True)[:10]
    results = []
    for reactivo in reactivos:
        result = {
            'id': reactivo.id,
            'name': reactivo.name,
            'code': reactivo.code,
            'cas': reactivo.cas
        }
        results.append(result)

    return JsonResponse(results, safe=False)


# Autocompleta reactivos en el inventario
class AutocompleteReactivosAPI(LoginRequiredMixin,View):
    def get(self, request):
        term = request.GET.get('term', '')
        lab = request.GET.get('lab', '')
        if lab=='0':
            inventarios = Inventarios.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term), is_active=True
        ).order_by('name').distinct('name')[:10]
        else:
            inventarios = Inventarios.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term),
            lab=lab, is_active=True
        ).order_by('name').distinct('name')[:10]
        results = []
        for inventario in inventarios:
            result = {
                'id': inventario.name.id,
                'name': inventario.name.name,
                'code': inventario.name.code,
                'cas': inventario.name.cas,
                
            }
            results.append(result)

        return JsonResponse(results, safe=False)

# Autocompleta reactivos en listado de entradas
class AutocompleteReactivosInAPI(LoginRequiredMixin,View):
    def get(self, request):
        term = request.GET.get('term', '')
        lab = request.GET.get('lab', '')
        if lab=='0':
            entradas = Entradas.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term), is_active=True
        ).order_by('name').distinct('name')[:10]
        else:
            entradas = Entradas.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term),
            lab=lab, is_active=True
        ).order_by('name').distinct('name')[:10]
        results = []
        for entrada in entradas:
            result = {
                'id': entrada.name.id,
                'name': entrada.name.name,
                'code': entrada.name.code,
                'cas': entrada.name.cas,
                
            }
            results.append(result)

        return JsonResponse(results, safe=False)

# Autocompleta reactivos en listado de salidas
class AutocompleteReactivosOutAPI(LoginRequiredMixin,View):
    def get(self, request):
        term = request.GET.get('term', '')
        lab = request.GET.get('lab', '')
        if lab=='0':
            entradas = Salidas.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term), is_active=True
        ).order_by('name').distinct('name')[:10]
        else:
            entradas = Salidas.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term),
            lab=lab, is_active=True
        ).order_by('name').distinct('name')[:10]
        results = []
        for entrada in entradas:
            result = {
                'id': entrada.name.id,
                'name': entrada.name.name,
                'code': entrada.name.code,
                'cas': entrada.name.cas,
                
            }
            results.append(result)

        return JsonResponse(results, safe=False)
    

# Devuelve los valores de la tabla Inventarios según lo escrito en el campo name del formulario registrar_salida.html en forma de una 
# lista de autocompletado

class AutocompleteOutAPI(LoginRequiredMixin,View):
    def get(self, request):
        term = request.GET.get('term', '')
        lab = request.GET.get('lab', '')
       
        inventarios = Inventarios.objects.filter(
            Q(name__name__icontains=term) | Q(name__code__icontains=term) | Q(name__cas__icontains=term),
            lab__name__icontains=lab,
            weight__gt=0, is_active=True
        ).order_by('name').distinct('name')[:10]

        results = []
        for inventario in inventarios:
            result = {
                'id': inventario.name.id,
                'name': inventario.name.name,
                'code': inventario.name.code,
                'cas': inventario.name.cas,
            }
            results.append(result)

        return JsonResponse(results, safe=False)
# Devuelve los valores de la tabla Usuarios según lo escrito en el campo name de la vista listado de usuarios en forma de una 
# lista de autocompletado

class AutocompleteUserAPI(LoginRequiredMixin,View):
    def get(self, request):
        term = request.GET.get('term', '')
        lab = request.GET.get('lab', '')
        if not lab:
            lab='0'
        if lab=='0':       
            usuarios = User.objects.filter(
                Q(first_name__icontains=term) | Q(last_name__icontains=term) | Q(id_number__icontains=term) | Q(phone_number__icontains=term) | Q(email__icontains=term) | Q(username__icontains=term)
            ).order_by('first_name')[:10]
        else:
            usuarios = User.objects.filter(
                Q(first_name__icontains=term) | Q(last_name__icontains=term) | Q(id_number__icontains=term) | Q(phone_number__icontains=term) | Q(email__icontains=term) | Q(username__icontains=term),
                lab=lab
            ).order_by('first_name')[:10]

        results = []
        for usuario in usuarios:
            result = {
                'first_name': usuario.first_name,
                'last_name': usuario.last_name,
                'email': usuario.email,
                'id_user': usuario.id,
            }
            results.append(result)
        return JsonResponse(results, safe=False)
    

# Devuelve los valores de la tabla Ubicaciones según lo escrito en el campo name del formulario registrar_salida.html en forma de 
# una lista de autocompletado
@login_required
def autocomplete_location(request):
    term = request.GET.get('term', '')
    ubicaciones = Ubicaciones.objects.filter(Q(name__icontains=term))[:10]
    results = []
    for ubicacion in ubicaciones:
        results.append({'name': ubicacion.name,'facultad':ubicacion.facultad.name })

    return JsonResponse(results, safe=False)



# Devuelve los valores de la tabla Responsables según lo escrito en el campo name del formulario registrar_salida.html en forma de 
# una lista de autocompletado
@login_required

def autocomplete_manager(request):
    term = request.GET.get('term', '')
    responsables = Responsables.objects.filter(Q(name__icontains=term) | Q(phone__icontains=term) | Q(mail__icontains=term) | Q(cc__icontains=term))[:10]
    results = []
    for responsable in responsables:
        results.append({'name': responsable.name, 'mail': responsable.mail})
    return JsonResponse(results, safe=False)

from django.http import JsonResponse

# Envía el Stock de un reactivo actual dependiendo de los valores de lab, name, marca y referencia como información al usuario 
# sepa cuanto retirar
@login_required
def obtener_stock(request):
    if request.method == "GET":
        lab_name = request.GET.get("lab")
        name = request.GET.get("name")
        trademark = request.GET.get("trademark")
        reference = request.GET.get("reference")
        
        # Obtener los IDs correspondientes a partir de los nombres
        lab = get_object_or_404(Laboratorios, name=lab_name)
        reactivos = get_object_or_404(Reactivos, name=name)
        
        # Realizar la lógica para obtener el stock en base a los parámetros recibidos

        # Supongamos que obtienes el stock de la base de datos o de alguna otra fuente de datos
        stock = Inventarios.objects.filter(lab=lab, name=reactivos, trademark=trademark, reference=reference).values("weight").first()
       
        # Devolver la respuesta en formato JSON
        return JsonResponse({"stock": stock})
    
#Estandariza la escritura en la base de datos que sea mayúscula, sin tildes ni nigún tipo de caracter esecial, reemplaza Ñ por N
def estandarizar_nombre(nombre):
    nombre = nombre.upper()  # Convertir a mayúsculas
    nombre = re.sub('[áÁ]', 'A', nombre)  # Reemplazar á y Á por A
    nombre = re.sub('[éÉ]', 'E', nombre)  # Reemplazar é y É por E
    nombre = re.sub('[íÍ]', 'I', nombre)  # Reemplazar í y Í por I
    nombre = re.sub('[óÓ]', 'O', nombre)  # Reemplazar ó y Ó por O
    nombre = re.sub('[úÚ]', 'U', nombre)  # Reemplazar ú y Ú por U
    nombre = re.sub('[ñÑ]', 'N', nombre)  # Reemplazar ñ y Ñ por N
    nombre = re.sub('[^A-Za-z0-9@ .,()%_-]', '', nombre)  # Eliminar caracteres especiales excepto números y espacios
    nombre = nombre.strip()  # Eliminar espacios al principio y al final
    return nombre


class OcultarReactivoView(LoginRequiredMixin, View):
    template_name = 'reactivos/ocultar_visibilidad.html'  # Asegúrate de tener este template creado

    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def get(self, request, pk):
        usuario = get_object_or_404(User, pk=pk)
        context = {
            'usuario': usuario,
        }
        return render(request, self.template_name, context)
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def post(self, request, pk):
        
        # Obtiene el objeto de inventario
        try:
            inventory = Inventarios.objects.get(pk=pk)
        except Inventarios.DoesNotExist:
            mensaje = f'El reactivo no existe.'
            return HttpResponse(mensaje,200)

        # Realiza la actualización
        inventory.visibility = False
        inventory.last_updated_by = request.user
        inventory.save()

        # Crea un evento de ELIMINAR USUARIO
        tipo_evento = 'OCULTAR REACTIVO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        # Construye el mensaje de éxito
        mensaje = f'El reactivo "{inventory.name}" se ha ocultado de manera exitosa.'
        
        return HttpResponse(mensaje,200)
    
class MostrarReactivoView(LoginRequiredMixin, View):
    template_name = 'reactivos/mostrar_visibilidad.html'  # Asegúrate de tener este template creado

    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def get(self, request, pk):
        usuario = get_object_or_404(User, pk=pk)
        context = {
            'usuario': usuario,
        }
        return render(request, self.template_name, context)
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def post(self, request, pk):
        
        # Obtiene el objeto de inventario
        try:
            inventory = Inventarios.objects.get(pk=pk)
        except Inventarios.DoesNotExist:
            mensaje = f'El reactivo no existe.'
            return HttpResponse(mensaje,200)

        # Realiza la actualización
        inventory.visibility = True
        inventory.last_updated_by = request.user
        inventory.save()

        # Crea un evento de ELIMINAR USUARIO
        tipo_evento = 'MOSTRAR REACTIVO'
        usuario_evento = request.user
        crear_evento(tipo_evento, usuario_evento)

        # Construye el mensaje de éxito
        mensaje = f'Se a activado la visibilidad del reactivo "{inventory.name}" de manera exitosa.'
        
        return HttpResponse(mensaje,200)

class RevisarDisponibilidadView(LoginRequiredMixin, View):
    template_name = 'reactivos/revisar_disponibilidad.html'  # Asegúrate de tener este template creado

    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR'])
    def get(self, request, pk):
        usuario = get_object_or_404(User, pk=pk)
        context = {
            'usuario': usuario,
        }
        return render(request, self.template_name, context)
    
    @check_group_permission(groups_required=['COORDINADOR', 'ADMINISTRADOR', 'TECNICO'])
    def post(self, request, pk):
        # Inicializar el diccionario para realizar el seguimiento del total por laboratorio
        total_por_laboratorio = defaultdict(Decimal)
        # Incializar el mensaje
        mensaje=''
        # Obtener el nombre del reactivo
        reactivo = get_object_or_404(Reactivos, pk=pk)
        # Obtener el objeto de inventario
        try:
            inventories = Inventarios.objects.filter(name=pk, is_active=True, visibility=True)
        except Inventarios.DoesNotExist:
            mensaje = f'No hay existencia del reactivo en inventario.'
            return HttpResponse(mensaje,200)

        # Calcular el total por cada laboratorio
        for inventory in inventories:
            if inventory.lab:
                total_por_laboratorio[inventory.lab.name] += inventory.weight

        # Imprimir el total por cada laboratorio
        for lab_name, total_quantity in total_por_laboratorio.items():
            units = inventories.filter(lab__name=lab_name).first().name.unit
            # Mensaje como html para que se pueda ver como sweet alert
            mensaje+=f'<li><strong>Laboratorio:</strong> {lab_name}</li><li><strong>Cantidad:</strong> {total_quantity} {units}</li><br>'
        
        mensaje=f'<div class="card" style="text-align: left;"><div class="card-body"><h5><strong>Existencia en Laboratorios - {reactivo.name}</strong></h5><br><ul class="list-unstyled">{mensaje}<a href="/export2xlsxlab/">Ver información de Laboratorios</a></ul></div></div>'
        

        return HttpResponse(mensaje, 200)



import warnings
from urllib.parse import urlparse, urlunparse



from django.conf import settings

# Avoid shadowing the login() and logout() views below.
from django.contrib.auth import REDIRECT_FIELD_NAME, get_user_model
from django.contrib.auth import login as auth_login
from django.contrib.auth import logout as auth_logout
from django.contrib.auth import update_session_auth_hash
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import (
    AuthenticationForm,
    PasswordChangeForm,
    PasswordResetForm,
    SetPasswordForm,
)



class RedirectURLMixin:
    next_page = None
    redirect_field_name = REDIRECT_FIELD_NAME
    success_url_allowed_hosts = set()

    def get_success_url(self):
        return self.get_redirect_url() or self.get_default_redirect_url()

    def get_redirect_url(self):
        """Return the user-originating redirect URL if it's safe."""
        redirect_to = self.request.POST.get(
            self.redirect_field_name, self.request.GET.get(self.redirect_field_name)
        )
        url_is_safe = url_has_allowed_host_and_scheme(
            url=redirect_to,
            allowed_hosts=self.get_success_url_allowed_hosts(),
            require_https=self.request.is_secure(),
        )
        return redirect_to if url_is_safe else ""

    def get_success_url_allowed_hosts(self):
        return {self.request.get_host(), *self.success_url_allowed_hosts}

    def get_default_redirect_url(self):
        """Return the default redirect URL."""
        if self.next_page:
            return resolve_url(self.next_page)
        raise ImproperlyConfigured("No URL to redirect to. Provide a next_page.")


class LoginView(RedirectURLMixin, FormView):
    """
    Display the login form and handle the login action.
    """

    form_class = AuthenticationForm
    authentication_form = None
    template_name = "registration/login.html"
    redirect_authenticated_user = False
    extra_context = None

    @method_decorator(sensitive_post_parameters())
    @method_decorator(csrf_protect)
    @method_decorator(never_cache)
    def dispatch(self, request, *args, **kwargs):
        if self.redirect_authenticated_user and self.request.user.is_authenticated:
            redirect_to = self.get_success_url()
            if redirect_to == self.request.path:
                raise ValueError(
                    "Redirection loop for authenticated user detected. Check that "
                    "your LOGIN_REDIRECT_URL doesn't point to a login page."
                )
            
            return HttpResponseRedirect(redirect_to)
        return super().dispatch(request, *args, **kwargs)

    def get_default_redirect_url(self):
        """Return the default redirect URL."""
        if self.next_page:
            return resolve_url(self.next_page)
        else:
            return resolve_url(settings.LOGIN_REDIRECT_URL)

    def get_form_class(self):
        return self.authentication_form or self.form_class

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["request"] = self.request
        return kwargs

    def form_valid(self, form):
        """Security check complete. Log the user in."""
        auth_login(self.request, form.get_user())
        # Crea un evento de inicio de sesión
        tipo_evento = 'INICIO DE SESION'
        usuario_evento = self.request.user
        crear_evento(tipo_evento, usuario_evento)
        return HttpResponseRedirect(self.get_success_url())

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_site = get_current_site(self.request)
        context.update(
            {
                self.redirect_field_name: self.get_redirect_url(),
                "site": current_site,
                "site_name": current_site.name,
                **(self.extra_context or {}),
            }
        )
        return context


class LogoutView(RedirectURLMixin, TemplateView):
    """
    Log out the user and display the 'You are logged out' message.
    """

    # RemovedInDjango50Warning: when the deprecation ends, remove "get" and
    # "head" from http_method_names.
    http_method_names = ["get", "head", "post", "options"]
    template_name = "registration/logged_out_reactivos.html"
    extra_context = None

    # RemovedInDjango50Warning: when the deprecation ends, move
    # @method_decorator(csrf_protect) from post() to dispatch().
    @method_decorator(never_cache)
    def dispatch(self, request, *args, **kwargs):
        return super().dispatch(request, *args, **kwargs)

    @method_decorator(csrf_protect)
    def post(self, request, *args, **kwargs):
        # Crea un evento de cierre de sesión
        tipo_evento = 'CIERRE DE SESION'
        usuario_evento = self.request.user
        crear_evento(tipo_evento, usuario_evento)

        """Logout may be done via POST."""
        auth_logout(request)
        redirect_to = self.get_success_url()
        if redirect_to != request.get_full_path():
            # Redirect to target page once the session has been cleared.
            return HttpResponseRedirect(redirect_to)
        return super().get(request, *args, **kwargs)

    # RemovedInDjango50Warning.
    get = post

    def get_default_redirect_url(self):
        """Return the default redirect URL."""
        if self.next_page:
            return resolve_url(self.next_page)
        elif settings.LOGOUT_REDIRECT_URL:
            return resolve_url(settings.LOGOUT_REDIRECT_URL)
        else:
            return self.request.path

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        current_site = get_current_site(self.request)
        context.update(
            {
                "site": current_site,
                "site_name": current_site.name,
                "title": _("Logged out"),
                "subtitle": None,
                **(self.extra_context or {}),
            }
        )
        return context


def logout_then_login(request, login_url=None):
    """
    Log out the user if they are logged in. Then redirect to the login page.
    """
    login_url = resolve_url(login_url or settings.LOGIN_URL)
    return LogoutView.as_view(next_page=login_url)(request)


def redirect_to_login(next, login_url=None, redirect_field_name=REDIRECT_FIELD_NAME):
    """
    Redirect the user to the login page, passing the given 'next' page.
    """
    resolved_url = resolve_url(login_url or settings.LOGIN_URL)

    login_url_parts = list(urlparse(resolved_url))
    if redirect_field_name:
        querystring = QueryDict(login_url_parts[4], mutable=True)
        querystring[redirect_field_name] = next
        login_url_parts[4] = querystring.urlencode(safe="/")

    return HttpResponseRedirect(urlunparse(login_url_parts))


# Class-based password reset views
# - PasswordResetView sends the mail
# - PasswordResetDoneView shows a success message for the above
# - PasswordResetConfirmView checks the link the user clicked and
#   prompts for a new password
# - PasswordResetCompleteView shows a success message for the above


class PasswordContextMixin:
    extra_context = None

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.update(
            {"title": self.title, "subtitle": None, **(self.extra_context or {})}
        )
        return context


class PasswordResetView(PasswordContextMixin, FormView):
    email_template_name = "registration/password_reset_email_reactivos.html"
    extra_email_context = None
    
    from_email = None
    html_email_template_name = None
    subject_template_name = "registration/password_reset_subject.txt"
    success_url = reverse_lazy("reactivos:password_reset_done_reactivos")
    template_name = "registration/password_reset_formulario.html"
    title = _("Password reset")
    token_generator = default_token_generator
    

    @method_decorator(csrf_protect)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        opts = {
            "use_https": self.request.is_secure(),
            "token_generator": self.token_generator,
            "from_email": self.from_email,
            "email_template_name": self.email_template_name,
            "subject_template_name": self.subject_template_name,
            "request": self.request,
            "html_email_template_name": self.html_email_template_name,
            "extra_email_context": self.extra_email_context,
        }
        form.save(**opts)
        return super().form_valid(form)
    
class CustomPasswordResetView(PasswordResetView):
    form_class = CustomPasswordResetForm
    email_template_name = "registration/password_reset_email_reactivos.html"
    extra_email_context = None
    
    from_email = None
    html_email_template_name = None
    subject_template_name = "registration/password_reset_subject.txt"
    success_url = reverse_lazy("reactivos:password_reset_done_reactivos")
    template_name = "registration/password_reset_formulario.html"
    title = _("Password reset")
    token_generator = default_token_generator
    

    @method_decorator(csrf_protect)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def form_valid(self, form):
        opts = {
            "use_https": self.request.is_secure(),
            "token_generator": self.token_generator,
            "from_email": self.from_email,
            "email_template_name": self.email_template_name,
            "subject_template_name": self.subject_template_name,
            "request": self.request,
            "html_email_template_name": self.html_email_template_name,
            "extra_email_context": self.extra_email_context,
        }
        form.save(**opts)
        return super().form_valid(form)


INTERNAL_RESET_SESSION_TOKEN = "_password_reset_token"


class PasswordResetDoneView(PasswordContextMixin, TemplateView):
    template_name = "registration/password_reset_done_reactivos.html"
    title = _("Password reset sent")


class PasswordResetConfirmView(PasswordContextMixin, FormView):
    form_class = SetPasswordForm
    post_reset_login = False
    post_reset_login_backend = None
    reset_url_token = "set-password"
    success_url = reverse_lazy("reactivos:password_reset_complete")
    template_name = "registration/password_reset_confirmacion.html"
    title = _("Enter new password")
    token_generator = default_token_generator

    @method_decorator(sensitive_post_parameters())
    @method_decorator(never_cache)
    def dispatch(self, *args, **kwargs):
        if "uidb64" not in kwargs or "token" not in kwargs:
            raise ImproperlyConfigured(
                "The URL path must contain 'uidb64' and 'token' parameters."
            )

        self.validlink = False
        self.user = self.get_user(kwargs["uidb64"])

        if self.user is not None:
            token = kwargs["token"]
            if token == self.reset_url_token:
                session_token = self.request.session.get(INTERNAL_RESET_SESSION_TOKEN)
                if self.token_generator.check_token(self.user, session_token):
                    # If the token is valid, display the password reset form.
                    self.validlink = True
                    return super().dispatch(*args, **kwargs)
            else:
                if self.token_generator.check_token(self.user, token):
                    # Store the token in the session and redirect to the
                    # password reset form at a URL without the token. That
                    # avoids the possibility of leaking the token in the
                    # HTTP Referer header.
                    self.request.session[INTERNAL_RESET_SESSION_TOKEN] = token
                    redirect_url = self.request.path.replace(
                        token, self.reset_url_token
                    )
                    return HttpResponseRedirect(redirect_url)

        # Display the "Password reset unsuccessful" page.
        return self.render_to_response(self.get_context_data())

    def get_user(self, uidb64):
        try:
            # urlsafe_base64_decode() decodes to bytestring
            uid = urlsafe_base64_decode(uidb64).decode()
            user = UserModel._default_manager.get(pk=uid)
        except (
            TypeError,
            ValueError,
            OverflowError,
            UserModel.DoesNotExist,
            ValidationError,
        ):
            user = None
        return user

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.user
        return kwargs

    def form_valid(self, form):
        user = form.save()
        del self.request.session[INTERNAL_RESET_SESSION_TOKEN]
        if self.post_reset_login:
            auth_login(self.request, user, self.post_reset_login_backend)
        return super().form_valid(form)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.validlink:
            context["validlink"] = True
        else:
            context.update(
                {
                    "form": None,
                    "title": _("Password reset unsuccessful"),
                    "validlink": False,
                }
            )
        return context




class PasswordResetCompleteView(PasswordContextMixin, TemplateView):
    template_name = "registration/password_reset_complete_reactivos.html"
    title = _("Password reset complete")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["login_url"] = resolve_url(settings.LOGIN_URL)
        return context


class PasswordChangeView(PasswordContextMixin, FormView):
    form_class = PasswordChangeForm
    success_url = reverse_lazy("reactivos:password_change_done_reactivos")
    template_name = "registration/password_change_form_reactivos.html"
    title = _("Password change")

    @method_decorator(sensitive_post_parameters())
    @method_decorator(csrf_protect)
    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)

    def get_form_kwargs(self):
        kwargs = super().get_form_kwargs()
        kwargs["user"] = self.request.user
        return kwargs

    def form_valid(self, form):
        form.save()
        # Updating the password logs out all other sessions for the user
        # except the current one.
        # Crea un evento de cambio de contraseña
        tipo_evento = 'CAMBIAR CONTRASENA'
        usuario_evento = self.request.user
        crear_evento(tipo_evento, usuario_evento)

        update_session_auth_hash(self.request, form.user)
        return super().form_valid(form)


class PasswordChangeDoneView(PasswordContextMixin, TemplateView):
    template_name = "registration/password_change_done_reactivos.html"
    title = _("Password change successful")

    @method_decorator(login_required)
    def dispatch(self, *args, **kwargs):
        return super().dispatch(*args, **kwargs)
